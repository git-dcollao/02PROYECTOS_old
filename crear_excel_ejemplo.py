#!/usr/bin/env python3
"""
Script para crear un archivo Excel de ejemplo para pruebas de subida de control
"""
import sys
import os
from datetime import datetime, timedelta

# Agregar el directorio raíz al path
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl no está instalado. Ejecute: pip install openpyxl")
    sys.exit(1)

def crear_excel_ejemplo():
    """Crea un archivo Excel de ejemplo para subida de control"""
    
    print("🚀 Creando archivo Excel de ejemplo...")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Actividades"
    
    # Definir encabezados para el formato de procesar-proyecto-xlsx
    encabezados = [
        "Proyecto",
        "EDT",
        "Modo de tarea", 
        "Nombre de tarea",
        "Duración",
        "Comienzo",
        "Fin",
        "Predecesoras",
        "Nombres de los recursos",
        "% completado",
        "Nivel de esquema",
        "Días Corrido"
    ]
    
    # Escribir encabezados con formato
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=encabezado)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Datos de ejemplo que simulan el formato español con fechas problemáticas y recursos realistas
    fecha_base = datetime(2010, 1, 29)
    
    datos_ejemplo = [
        ["Proyecto Test", "1", "Automática", "Proyecto Principal", "5 días", "vie 29-01-10 9:00", "jue 04-02-10 15:00", "", "Juan Pérez", "0%", 1, "5 días"],
        ["", "1.1", "Automática", "Planificación del proyecto", "3 días", "vie 29-01-10 9:00", "mar 02-02-10 17:00", "", "Juan Pérez;Ana García", "100%", 2, "3 días"],
        ["", "1.2", "Automática", "Análisis de requerimientos", "2 días", "mié 03-02-10 9:00", "jue 04-02-10 15:00", "1.1", "Maria Rodriguez[60%];Carlos López[40%]", "80%", 2, "2 días"],
        ["", "1.3", "Automática", "Diseño técnico", "4 días", "lun 08-02-10 9:00", "jue 11-02-10 17:00", "1.2", "Roberto Silva,Luis Martín", "60%", 2, "4 días"],
        ["", "2", "Automática", "Fase de Desarrollo", "10 días", "vie 12-02-10 9:00", "jue 25-02-10 17:00", "1", "Equipo Dev", "0%", 1, "10 días"],
        ["", "2.1", "Automática", "Desarrollo módulo 1", "6 días", "vie 12-02-10 9:00", "vie 19-02-10 17:00", "1.3", "Pedro Morales[80%];Sofia Ramírez[20%]", "40%", 2, "6 días"],
        ["", "2.2", "Automática", "Desarrollo módulo 2", "6 días", "lun 15-02-10 9:00", "lun 22-02-10 17:00", "1.3", "Diego Torres;Lucia Fernández", "30%", 2, "6 días"],
        ["", "3", "Automática", "Fase de Pruebas", "5 días", "mar 23-02-10 9:00", "lun 01-03-10 17:00", "2", "QA Team", "0%", 1, "5 días"],
        ["", "3.1", "Automática", "Pruebas unitarias", "2 días", "mar 23-02-10 9:00", "mié 24-02-10 17:00", "2.1", "Carmen Jiménez;Miguel Ruiz[50%]", "20%", 2, "2 días"],
        ["", "3.2", "Automática", "Pruebas integración", "3 días", "jue 25-02-10 9:00", "lun 01-03-10 17:00", "2.2,3.1", "Elena Castro[70%],Fernando Ortega[30%]", "10%", 2, "3 días"],
    ]
    
    # Escribir datos (sin formateo especial de fechas ya que son strings)
    for fila, datos in enumerate(datos_ejemplo, 2):
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=fila, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for col in range(1, len(encabezados) + 1):
        column_letter = get_column_letter(col)
        if col == 4:  # Nombre de tarea
            ws.column_dimensions[column_letter].width = 30
        elif col in [6, 7]:  # Comienzo, Fin
            ws.column_dimensions[column_letter].width = 20
        elif col == 9:  # Nombres de los recursos
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = 15
    
    # Guardar archivo
    nombre_archivo = f"ejemplo_control_actividades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_archivo = os.path.join(os.path.dirname(__file__), nombre_archivo)
    
    try:
        wb.save(ruta_archivo)
        print(f"✅ Archivo creado exitosamente: {nombre_archivo}")
        print(f"📍 Ubicación: {ruta_archivo}")
        print(f"📊 Datos: {len(datos_ejemplo)} actividades de ejemplo")
        return ruta_archivo
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {str(e)}")
        return None

if __name__ == '__main__':
    archivo = crear_excel_ejemplo()
    if archivo:
        print("🎉 Proceso completado exitosamente")
        print("\n📋 Instrucciones:")
        print("1. Abre el archivo Excel generado")
        print("2. Modifica los datos según sea necesario") 
        print("3. Guarda el archivo")
        print("4. Súbelo usando el botón 'Subir control' en la aplicación")
    else:
        print("💥 Proceso falló")
        sys.exit(1)
