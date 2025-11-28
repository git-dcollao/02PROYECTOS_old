#!/usr/bin/env python3
"""
Script para crear un archivo Excel de prueba específico para probar el filtro de proyectos
"""
import sys
import os
from datetime import datetime, timedelta

# Agregar el directorio raíz al path
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl no está instalado. Ejecute: pip install openpyxl")
    sys.exit(1)

def crear_excel_prueba_filtro():
    """Crea un archivo Excel para probar el filtro de nombres duplicados"""
    
    print("🚀 Creando archivo Excel de prueba para filtro...")
    
    # Nombre del archivo que se generará
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo_sin_extension = f"prueba_filtro_{timestamp}"
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Actividades"
    
    # Definir encabezados
    encabezados = [
        "Proyecto",
        "EDT", 
        "Modo de tarea",
        "Nombre de tarea",
        "Duración",
        "Comienzo",
        "Fin",
        "Predecesoras",
        "Nombres de los recursos",
        "% completado",
        "Nivel de esquema",
        "Días Corrido"
    ]
    
    # Escribir encabezados con formato
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=encabezado)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Datos de ejemplo - incluye un proyecto principal con el mismo nombre del archivo (que debe ser excluido)
    # y otro proyecto principal diferente (que debe ser incluido en la lista de proyectos para asignación)
    datos_ejemplo = [
        # Proyecto 1: Este debe ser EXCLUIDO porque su nombre coincide con el archivo
        ["Test Filter", "1", "Automática", nombre_archivo_sin_extension, "10 días", "vie 29-01-10 9:00", "jue 11-02-10 15:00", "", "Juan Pérez", "0%", 1, "10 días"],
        ["", "1.1", "Automática", "Actividad del proyecto duplicado", "3 días", "vie 29-01-10 9:00", "mar 02-02-10 17:00", "", "Juan Pérez", "100%", 2, "3 días"],
        ["", "1.2", "Automática", "Segunda actividad duplicado", "2 días", "mié 03-02-10 9:00", "jue 04-02-10 15:00", "1.1", "Ana García", "80%", 2, "2 días"],
        
        # Proyecto 2: Este debe ser INCLUIDO porque su nombre es diferente al archivo
        ["", "2", "Automática", "Proyecto Diferente que Debe Aparecer", "8 días", "vie 12-02-10 9:00", "mar 23-02-10 17:00", "", "Equipo Dev", "0%", 1, "8 días"],
        ["", "2.1", "Automática", "Desarrollo módulo principal", "5 días", "vie 12-02-10 9:00", "jue 18-02-10 17:00", "", "Pedro Morales", "50%", 2, "5 días"],
        ["", "2.2", "Automática", "Integración de módulos", "3 días", "vie 19-02-10 9:00", "mar 23-02-10 17:00", "2.1", "Sofia Ramírez", "30%", 2, "3 días"],
        
        # Proyecto 3: Otro proyecto INCLUIDO con nombre diferente
        ["", "3", "Automática", "Proyecto Final Test", "5 días", "mié 24-02-10 9:00", "mar 02-03-10 17:00", "2", "QA Team", "0%", 1, "5 días"],
        ["", "3.1", "Automática", "Pruebas finales", "2 días", "mié 24-02-10 9:00", "jue 25-02-10 17:00", "2.2", "Carmen Jiménez", "20%", 2, "2 días"],
        ["", "3.2", "Automática", "Validación completa", "3 días", "vie 26-02-10 9:00", "mar 02-03-10 17:00", "3.1", "Miguel Ruiz", "10%", 2, "3 días"],
    ]
    
    # Escribir datos
    for fila, datos in enumerate(datos_ejemplo, 2):
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=fila, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for col in range(1, len(encabezados) + 1):
        column_letter = get_column_letter(col)
        if col == 4:  # Nombre de tarea
            ws.column_dimensions[column_letter].width = 40
        elif col in [6, 7]:  # Comienzo, Fin
            ws.column_dimensions[column_letter].width = 20
        elif col == 9:  # Nombres de los recursos
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = 15
    
    # Guardar archivo
    nombre_archivo_completo = f"{nombre_archivo_sin_extension}.xlsx"
    ruta_archivo = os.path.join(os.path.dirname(__file__), nombre_archivo_completo)
    
    try:
        wb.save(ruta_archivo)
        print(f"✅ Archivo creado exitosamente: {nombre_archivo_completo}")
        print(f"📍 Ubicación: {ruta_archivo}")
        print(f"📊 Datos: {len(datos_ejemplo)} actividades de ejemplo")
        print(f"\n🔍 Prueba del filtro:")
        print(f"   ⚠️  Proyecto que DEBE SER EXCLUIDO: '{nombre_archivo_sin_extension}'")
        print(f"   ✅ Proyecto que DEBE APARECER: 'Proyecto Diferente que Debe Aparecer'")
        print(f"   ✅ Proyecto que DEBE APARECER: 'Proyecto Final Test'")
        return ruta_archivo
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {str(e)}")
        return None

if __name__ == '__main__':
    archivo = crear_excel_prueba_filtro()
    if archivo:
        print("\n🎉 Proceso completado exitosamente")
        print("\n📋 Instrucciones de prueba:")
        print("1. Ve a la aplicación web en http://127.0.0.1:5050")
        print("2. Sube el archivo Excel generado")
        print("3. Verifica que en la lista de proyectos para asignación aparezcan SOLO:")
        print("   - 'Proyecto Diferente que Debe Aparecer'")
        print("   - 'Proyecto Final Test'")
        print("4. NO debe aparecer el proyecto con el mismo nombre del archivo")
    else:
        print("💥 Proceso falló")
        sys.exit(1)
