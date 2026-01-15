from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app, make_response, session
from flask_login import current_user, login_required
from sqlalchemy.orm import joinedload
from app.models import (
    Requerimiento, TipoRecinto, Recinto, Sector, Trabajador, 
    Financiamiento, Especialidad, 
    Equipo, Tipologia, Fase, TipoProyecto, Estado, Prioridad, Grupo, Area, db,  # Agregar Prioridad, Area y Fase
    requerimiento_trabajador_especialidad, EquipoTrabajo, GanttArchivo, ActividadProyecto, AvanceActividad, HistorialAvanceActividad, HistorialControl,  # Agregar HistorialControl
    AdministradorRecinto, TrabajadorRecinto,  # Agregar modelos para gestión de permisos por recintos
    ObservacionRequerimiento,  # Nuevo modelo para observaciones
    ActividadGantt  # Modelo para carta Gantt (actividades del Gantt)
)
from datetime import datetime
from werkzeug.utils import secure_filename
import os
import pandas as pd
import openpyxl
from io import BytesIO
import json
import re
import logging
import uuid
import uuid
import locale

from werkzeug.exceptions import BadRequest
import time

# Configurar logger
logger = logging.getLogger(__name__)

def user_has_admin_permissions(user):
    """
    Verifica si un usuario tiene permisos administrativos.
    Esto incluye SUPERADMIN del enum UserRole o roles personalizados de administrador.
    """
    # Verificar si es SUPERADMIN del sistema
    if hasattr(user, 'rol') and user.rol:
        rol_name = user.rol.name if hasattr(user.rol, 'name') else str(user.rol)
        if rol_name and rol_name == 'SUPERADMIN':
            return True
    
    # Verificar si tiene un custom role de administrador
    if hasattr(user, 'custom_role_id') and user.custom_role_id:
        from app.models import CustomRole
        custom_role = CustomRole.query.get(user.custom_role_id)
        if custom_role and custom_role.name.upper() in ['ADMIN', 'SUPERADMIN']:
            return True
    
    return False

def parsear_fecha_espanol(fecha_str):
    """
    Parsea fechas en formato español como 'vie 29-01-10 9:00'
    """
    if not fecha_str:
        return None
        
    fecha_str = str(fecha_str).strip()
    
    # Diccionario de días en español a inglés
    dias_semana = {
        'lun': 'Mon', 'mar': 'Tue', 'mié': 'Wed', 'jue': 'Thu', 
        'vie': 'Fri', 'sáb': 'Sat', 'dom': 'Sun',
        'mie': 'Wed', 'sab': 'Sat'  # Variaciones sin acentos
    }
    
    try:
        # Formato: "vie 29-01-10 9:00"
        if ' ' in fecha_str and '-' in fecha_str:
            partes = fecha_str.split()
            if len(partes) >= 2:
                fecha_parte = partes[1]  # "29-01-10"
                
                # Extraer día, mes, año
                if '-' in fecha_parte:
                    fecha_components = fecha_parte.split('-')
                    if len(fecha_components) == 3:
                        dia = int(fecha_components[0])
                        mes = int(fecha_components[1])
                        año = int(fecha_components[2])
                        
                        # Ajustar año de 2 dígitos a 4 dígitos
                        if año < 50:
                            año += 2000
                        elif año < 100:
                            año += 1900
                        
                        return datetime(año, mes, dia).date()
        
        # Si no coincide con el formato esperado, intentar otros parseos
        # Intentar formato ISO
        try:
            return datetime.fromisoformat(fecha_str).date()
        except:
            pass
            
        # Intentar formato estándar
        try:
            return datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except:
            pass
            
        # Intentar formato dd/mm/yyyy
        try:
            return datetime.strptime(fecha_str, '%d/%m/%Y').date()
        except:
            pass
            
        print(f"⚠️ No se pudo parsear la fecha: {fecha_str}")
        return None
        
    except Exception as e:
        print(f"⚠️ Error al parsear fecha '{fecha_str}': {e}")
        return None

# Elimina la definición del modelo GanttArchivo aquí.
# El modelo GanttArchivo ya está definido en app\models.py.
# Si necesitas usarlo, importa desde app.models:
from app.models import GanttArchivo

controllers_bp = Blueprint('controllers', __name__)

# ==================================================================================
# Rutas CRUD para Recintos
@controllers_bp.route('/recintos', endpoint='ruta_recintos')
@login_required
def recintos():
    # Filtrar recintos por sector del usuario actual
    user_sector_id = current_user.sector_id
    if user_sector_id:
        # Obtener recintos del mismo sector a través de la relación tiporecinto->sector
        recintos = Recinto.query.join(TipoRecinto).filter(TipoRecinto.id_sector == user_sector_id).all()
        # También filtrar tipos de recinto por sector
        tiposrecintos = TipoRecinto.query.filter_by(id_sector=user_sector_id).all()
    else:
        # Si el usuario no tiene sector, mostrar todos (para SUPERADMIN)
        recintos = Recinto.query.all()
        tiposrecintos = TipoRecinto.query.all()
    
    return render_template('recintos.html', recintos=recintos, tiposrecintos=tiposrecintos)

# Agregar endpoint para obtener recintos en formato JSON filtrados por sector
@controllers_bp.route('/get_recintos', methods=['GET'])
@login_required
def get_recintos():
    # Filtrar recintos por sector del usuario actual
    user_sector_id = current_user.sector_id
    if user_sector_id:
        recintos = Recinto.query.join(TipoRecinto).filter(TipoRecinto.id_sector == user_sector_id).all()
    else:
        # Si el usuario no tiene sector, mostrar todos (para SUPERADMIN)
        recintos = Recinto.query.all()
    return jsonify([{'id': e.id, 'nombre': e.nombre} for e in recintos])

# Agregar esta nueva ruta para obtener recintos por tipo de recinto
@controllers_bp.route('/get_recintos_by_tipo/<int:tiporecinto_id>')
def get_recintos_by_tipo(tiporecinto_id):
    # Obtener todos los recintos del tipo especificado
    recintos = Recinto.query.filter_by(id_tiporecinto=tiporecinto_id).all()
    
    # Log de debug
    print(f"🔍 Debug - Buscando recintos para tipo_recinto_id: {tiporecinto_id}")
    print(f"🔍 Debug - Recintos encontrados: {len(recintos)}")
    for recinto in recintos:
        print(f"   - ID: {recinto.id}, Nombre: {recinto.nombre}")
    
    return jsonify([{
        'id': recinto.id,
        'nombre': recinto.nombre
    } for recinto in recintos])

@controllers_bp.route('/add_recinto', methods=['POST'], endpoint='add_recinto')
@login_required
def add_recinto():
    try:
        # Verificar token de formulario para prevenir envíos duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_recinto_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_recintos'))
        
        nombre = request.form['nombre']
        descripcion = request.form.get('descripcion', '')  # valor por defecto vacío
        id_tiporecinto = request.form.get('id_tiporecinto', '1')  # valor por defecto 1
        
        nuevo_recinto = Recinto(
            nombre=nombre,
            descripcion=descripcion,
            id_tiporecinto=id_tiporecinto  
        )
        db.session.add(nuevo_recinto)
        db.session.commit()
        
        # Guardar token para prevenir reenvíos
        if form_token:
            session['last_add_recinto_form_token'] = form_token
        
        flash('Recinto agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar recinto: {e}')
    return redirect(url_for('controllers.ruta_recintos'))

@controllers_bp.route('/update_recinto/<int:id>', methods=['POST'], endpoint='update_recinto')
@login_required
def update_recinto(id):
    try:
        # Verificar token de formulario para prevenir envíos duplicados
        form_token = request.form.get('form_token')
        session_token = session.get(f'last_edit_recinto_{id}_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_recintos'))
        
        recinto = Recinto.query.get_or_404(id)
        recinto.nombre = request.form['nombre']
        recinto.descripcion = request.form['descripcion']
        recinto.id_tiporecinto = request.form['id_tiporecinto']
        db.session.commit()
        
        # Guardar token para prevenir reenvíos
        if form_token:
            session[f'last_edit_recinto_{id}_form_token'] = form_token
        
        flash('Recinto actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar recinto: {e}')
    return redirect(url_for('controllers.ruta_recintos'))

@controllers_bp.route('/eliminar_recinto/<int:id>', methods=['POST'], endpoint='eliminar_recinto')
def eliminar_recinto(id):
    try:
        recinto = Recinto.query.get_or_404(id)
        db.session.delete(recinto)
        db.session.commit()
        flash('Recinto eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar recinto: {e}')
    return redirect(url_for('controllers.ruta_recintos'))

# ==================================================================================
# Rutas CRUD para Tipo de Recintos
@controllers_bp.route('/tiposrecintos', endpoint='ruta_tiposrecintos')
@login_required
def tiposrecintos():
    # Filtrar tipos de recinto por sector del usuario actual
    user_sector_id = current_user.sector_id
    if user_sector_id:
        tiposrecintos = TipoRecinto.query.filter_by(id_sector=user_sector_id).all()
        # Solo mostrar el sector del usuario
        sectores = Sector.query.filter_by(id=user_sector_id).all()
    else:
        # Si el usuario no tiene sector, mostrar todos (para SUPERADMIN)
        tiposrecintos = TipoRecinto.query.all()
        sectores = Sector.query.all()
    return render_template('tiposrecintos.html', tiposrecintos=tiposrecintos, sectores=sectores)

# Agregar endpoint para obtener tipos de recinto en formato JSON filtrados por sector
@controllers_bp.route('/get_tiposrecintos', methods=['GET'])
@login_required
def get_tiposrecintos():
    # Filtrar tipos de recinto por sector del usuario actual
    user_sector_id = current_user.sector_id
    if user_sector_id:
        tiposrecintos = TipoRecinto.query.filter_by(id_sector=user_sector_id).all()
    else:
        # Si el usuario no tiene sector, mostrar todos (para SUPERADMIN)
        tiposrecintos = TipoRecinto.query.all()
    return jsonify([{'id': e.id, 'nombre': e.nombre} for e in tiposrecintos])

# Agregar esta nueva ruta para obtener tipos de recinto por sector
@controllers_bp.route('/get_tiposrecintos_by_sector/<int:sector_id>')
def get_tiposrecintos_by_sector(sector_id):
    # Obtener todos los tipos de recinto del sector especificado
    tipos_recinto = TipoRecinto.query.filter_by(id_sector=sector_id).all()
    
    # Log de debug
    print(f"🔍 Debug - Buscando tipos de recinto para sector_id: {sector_id}")
    print(f"🔍 Debug - Tipos encontrados: {len(tipos_recinto)}")
    for tipo in tipos_recinto:
        print(f"   - ID: {tipo.id}, Nombre: {tipo.nombre}")
    
    return jsonify([{
        'id': tipo.id,
        'nombre': tipo.nombre
    } for tipo in tipos_recinto])

@controllers_bp.route('/add_tiporecinto', methods=['POST'], endpoint='add_tiporecinto')
@login_required
def add_tiporecinto():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_tiporecinto_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_tiposrecintos'))
        
        nombre = request.form['name']
        descripcion = request.form.get('descripcion', '')  # valor por defecto vacío
        id_sector = request.form.get('id_sector', '1')  # valor por defecto 1
        
        nuevo_tiporecinto = TipoRecinto(
            nombre=nombre,
            descripcion=descripcion,
            id_sector=id_sector  
        )
        db.session.add(nuevo_tiporecinto)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_tiporecinto_form_token'] = form_token
        
        flash('Tipo de Recinto agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar recinto: {e}')
    return redirect(url_for('controllers.ruta_tiposrecintos'))

@controllers_bp.route('/update_tiporecinto/<int:id>', methods=['POST'], endpoint='update_tiporecinto')
@login_required
def update_tiporecinto(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_tiporecinto_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_tiposrecintos'))
        
        tiporecinto = TipoRecinto.query.get_or_404(id)
        tiporecinto.nombre = request.form['name']
        tiporecinto.descripcion = request.form['descripcion']
        tiporecinto.id_sector = request.form['id_sector']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_tiporecinto_form_token'] = form_token
        
        flash('Tipo de Recinto actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar tipo de recinto: {e}')
    return redirect(url_for('controllers.ruta_tiposrecintos'))

@controllers_bp.route('/eliminar_tiporecinto/<int:id>', methods=['POST'], endpoint='eliminar_tiporecinto')
@login_required
def eliminar_tiporecinto(id):
    try:
        recinto = TipoRecinto.query.get_or_404(id)
        db.session.delete(recinto)
        db.session.commit()
        flash('Tipo de Recinto eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar tipo de recinto: {e}')
    return redirect(url_for('controllers.ruta_tiposrecintos'))

# ==================================================================================
# Rutas CRUD para Sector
@controllers_bp.route('/sectores', endpoint='ruta_sectores')
@login_required
def sectores():
    sectores = Sector.query.all()
    return render_template('sector.html', sectores=sectores)

@controllers_bp.route('/add_sector', methods=['POST'], endpoint='add_sector')
@login_required
def add_sector():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_sector_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_sectores'))
        
        nombre = request.form['nombre']
        nuevo_sector = Sector(nombre=nombre)
        db.session.add(nuevo_sector)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_sector_form_token'] = form_token
        
        flash('Sector agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar sector: {e}')
    return redirect(url_for('controllers.ruta_sectores'))

@controllers_bp.route('/update_sector/<int:id>', methods=['POST'], endpoint='update_sector')
@login_required
def update_sector(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_sector_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_sectores'))
        
        sector = Sector.query.get_or_404(id)
        sector.nombre = request.form['nombre']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_sector_form_token'] = form_token
        
        flash('Sector actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar sector: {e}')
    return redirect(url_for('controllers.ruta_sectores'))

@controllers_bp.route('/eliminar_sector/<int:id>', methods=['POST'], endpoint='eliminar_sector')
@login_required
def eliminar_sector(id):
    try:
        sector = Sector.query.get_or_404(id)
        db.session.delete(sector)
        db.session.commit()
        flash('Sector eliminada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar sector: {e}')
    return redirect(url_for('controllers.ruta_sectores'))


# ==================================================================================
# Rutas CRUD para Trabajador 
@controllers_bp.route('/trabajadores', endpoint='ruta_trabajadores')
@login_required
def trabajadores():
    from app.models import Sector, Recinto, TipoRecinto, CustomRole, AdministradorRecinto
    
    # Verificar el tipo de usuario y permisos
    is_superadmin = False
    is_administrador = False
    
    # Verificar si es SUPERADMIN del sistema
    if hasattr(current_user, 'rol') and current_user.rol:
        rol_name = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol)
        if rol_name and rol_name == 'SUPERADMIN':
            is_superadmin = True
    
    # Verificar si tiene rol personalizado ADMIN
    if hasattr(current_user, 'custom_role_id') and current_user.custom_role_id:
        custom_role = CustomRole.query.get(current_user.custom_role_id)
        if custom_role and custom_role.name.upper() == 'ADMIN':
            is_administrador = True
    
    # Filtrar trabajadores según permisos
    if is_superadmin:
        # SUPERADMIN ve todos los trabajadores
        trabajadores = Trabajador.query.options(
            db.joinedload(Trabajador.sector),
            db.joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector),
            db.joinedload(Trabajador.custom_role)
        ).all()
    elif is_administrador:
        # ADMINISTRADOR ve trabajadores del mismo sector
        user_sector_id = current_user.sector_id
        
        if user_sector_id:
            # Filtrar trabajadores por el mismo sector
            trabajadores = Trabajador.query.options(
                db.joinedload(Trabajador.sector),
                db.joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector),
                db.joinedload(Trabajador.custom_role)
            ).filter(Trabajador.sector_id == user_sector_id).all()
        else:
            # Si no tiene sector asignado, solo se ve a sí mismo
            trabajadores = Trabajador.query.options(
                db.joinedload(Trabajador.sector),
                db.joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector),
                db.joinedload(Trabajador.custom_role)
            ).filter(Trabajador.id == current_user.id).all()
    else:
        # Otros usuarios solo ven trabajadores del mismo recinto
        user_recinto_id = current_user.recinto_id
        if user_recinto_id:
            trabajadores = Trabajador.query.options(
                db.joinedload(Trabajador.sector),
                db.joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector),
                db.joinedload(Trabajador.custom_role)
            ).filter(Trabajador.recinto_id == user_recinto_id).all()
        else:
            # Si el usuario no tiene recinto asignado, no ve ningún trabajador excepto él mismo
            trabajadores = Trabajador.query.options(
                db.joinedload(Trabajador.sector),
                db.joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector),
                db.joinedload(Trabajador.custom_role)
            ).filter(Trabajador.id == current_user.id).all()
    
    # Obtener sectores y recintos disponibles según permisos
    if is_superadmin:
        # SUPERADMIN ve todos los sectores y recintos
        sectores = Sector.query.filter_by(activo=True).all()
        recintos = Recinto.query.join(TipoRecinto).filter(
            Recinto.activo == True,
            TipoRecinto.activo == True
        ).all()
    elif is_administrador:
        # ADMINISTRADOR ve sectores y recintos de su sector
        user_sector_id = current_user.sector_id
        
        if user_sector_id:
            # Mostrar solo su sector
            sectores = Sector.query.filter(
                Sector.id == user_sector_id,
                Sector.activo == True
            ).all()
            
            # Mostrar recintos del mismo sector
            recintos = Recinto.query.join(TipoRecinto).filter(
                TipoRecinto.id_sector == user_sector_id,
                Recinto.activo == True,
                TipoRecinto.activo == True
            ).all()
        else:
            sectores = []
            recintos = []
    else:
        # Otros usuarios solo ven sectores y recintos relacionados con su recinto
        user_recinto_id = current_user.recinto_id
        if user_recinto_id and current_user.recinto:
            # Obtener el sector del recinto del usuario
            user_sector_id = current_user.recinto.tiporecinto.id_sector if current_user.recinto.tiporecinto else None
            if user_sector_id:
                sectores = Sector.query.filter_by(id=user_sector_id, activo=True).all()
                # Obtener recintos del mismo tipo que el usuario
                tipo_recinto_id = current_user.recinto.tiporecinto.id if current_user.recinto.tiporecinto else None
                if tipo_recinto_id:
                    recintos = Recinto.query.filter_by(
                        id_tiporecinto=tipo_recinto_id,
                        activo=True
                    ).all()
                else:
                    recintos = []
            else:
                sectores = []
                recintos = []
        else:
            sectores = []
            recintos = []
    
    # Calcular estadísticas
    total_trabajadores = len(trabajadores)
    trabajadores_con_sector = len([t for t in trabajadores if t.sector_id])
    trabajadores_sin_sector = total_trabajadores - trabajadores_con_sector
    trabajadores_con_recinto = len([t for t in trabajadores if t.recinto_id])
    sectores_diferentes = len(set([t.sector_id for t in trabajadores if t.sector_id]))
    
    estadisticas = {
        'total_trabajadores': total_trabajadores,
        'trabajadores_con_sector': trabajadores_con_sector,
        'trabajadores_sin_sector': trabajadores_sin_sector,
        'trabajadores_con_recinto': trabajadores_con_recinto,
        'sectores_gestionables': len(sectores),
        'sectores_diferentes': sectores_diferentes
    }
    
    # Obtener roles personalizados disponibles
    roles_personalizados = CustomRole.query.filter_by(active=True).all()
    
    # Verificar permisos para crear nuevos trabajadores
    puede_crear = is_superadmin or is_administrador
    
    return render_template('trabajadores.html', 
                         trabajadores=trabajadores, 
                         sectores=sectores,
                         recintos=recintos,
                         roles_personalizados=roles_personalizados,
                         estadisticas=estadisticas,
                         puede_crear=puede_crear)

@controllers_bp.route('/admin/trabajadores', methods=['GET'], endpoint='admin_trabajadores')
@login_required
def admin_trabajadores():
    """Página de administración avanzada de trabajadores - Solo SUPERADMIN"""
    from app.models import Sector, Recinto, TipoRecinto
    
    # Verificar que sea SUPERADMIN
    user_role = current_user.rol.value if hasattr(current_user.rol, 'value') else str(current_user.rol)
    if user_role != 'superadmin':
        flash('Acceso denegado. Solo SUPERADMIN puede acceder a esta página.', 'error')
        return redirect(url_for('controllers.ruta_trabajadores'))
    
    # SUPERADMIN puede ver todos los trabajadores
    trabajadores = Trabajador.query.options(
        db.joinedload(Trabajador.sector),
        db.joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector)
    ).all()
    
    # SUPERADMIN puede ver todos los sectores y recintos
    sectores = Sector.query.filter_by(activo=True).all()
    recintos = Recinto.query.join(TipoRecinto).filter(
        Recinto.activo == True,
        TipoRecinto.activo == True
    ).all()
    
    # Obtener estadísticas completas para SUPERADMIN
    total_trabajadores = len(trabajadores)
    trabajadores_con_sector = len([t for t in trabajadores if t.sector_id])
    trabajadores_sin_sector = total_trabajadores - trabajadores_con_sector
    trabajadores_con_recinto = len([t for t in trabajadores if t.recinto_id])
    sectores_diferentes = len(set([t.sector_id for t in trabajadores if t.sector_id]))
    
    estadisticas = {
        'total_trabajadores': total_trabajadores,
        'trabajadores_con_sector': trabajadores_con_sector,
        'trabajadores_sin_sector': trabajadores_sin_sector,
        'trabajadores_con_recinto': trabajadores_con_recinto,
        'sectores_totales': len(sectores),
        'sectores_diferentes': sectores_diferentes,
        'por_sector': {}
    }
    
    # Estadísticas por sector
    for sector in sectores:
        count = len([t for t in trabajadores if t.sector_id == sector.id])
        if count > 0:
            estadisticas['por_sector'][sector.nombre] = count
    
    return render_template('trabajadores_admin.html', 
                         trabajadores=trabajadores, 
                         sectores=sectores,
                         recintos=recintos,
                         estadisticas=estadisticas)

@controllers_bp.route('/add_trabajador', methods=['POST'], endpoint='add_trabajador')
@login_required
def add_trabajador():
    try:
        # Verificar si el usuario tiene permisos administrativos
        has_admin_permissions = user_has_admin_permissions(current_user)
        
        nombre = request.form['name']
        rut = request.form.get('rut', '').strip()
        profesion = request.form.get('profesion', '')
        nombrecorto = request.form.get('nombrecorto', '')
        sector_id = request.form.get('sector_id')
        recinto_id = request.form.get('recinto_id')
        
        # Campos de autenticación (solo si se seleccionó crear credenciales)
        crear_credenciales = request.form.get('crear_credenciales') == 'on'
        email = request.form.get('email', '').strip() if crear_credenciales else None
        password = request.form.get('password', '').strip() if crear_credenciales else None
        rol_tipo = request.form.get('rol_tipo') if crear_credenciales else None
        custom_role_id = request.form.get('custom_role_id') if crear_credenciales else None
        
        # Validar permisos para crear trabajador en el recinto especificado
        from app.utils.area_permissions import puede_crear_trabajador_en_recinto
        
        recinto_objetivo = int(recinto_id) if recinto_id else None
        
        # Verificar permisos específicos para administradores con recintos asignados
        from app.models import CustomRole, AdministradorRecinto
        
        # Verificar si es administrador con recintos asignados
        is_administrador = False
        if hasattr(current_user, 'custom_role_id') and current_user.custom_role_id:
            custom_role = CustomRole.query.get(current_user.custom_role_id)
            if custom_role and custom_role.name.upper() == 'ADMINISTRADOR':
                is_administrador = True
        
        if is_administrador and recinto_objetivo:
            # Verificar que el administrador tenga acceso al recinto donde quiere crear el trabajador
            if not AdministradorRecinto.tiene_acceso_recinto(current_user.id, recinto_objetivo):
                flash('No tiene permisos para crear trabajadores en el recinto seleccionado', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        elif recinto_objetivo and not puede_crear_trabajador_en_recinto(current_user, recinto_objetivo):
            flash('No tiene permisos para crear trabajadores en el recinto seleccionado', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar restricciones de recinto según tipo de usuario
        is_superadmin = False
        if hasattr(current_user, 'rol') and current_user.rol:
            rol_name = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol)
            if rol_name and rol_name == 'SUPERADMIN':
                is_superadmin = True
        
        # Si es administrador (no superadmin), validar que el recinto esté en sus asignaciones
        if is_administrador and not is_superadmin:
            if not recinto_objetivo:
                flash('Debe seleccionar un recinto para crear el trabajador', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            # El recinto ya fue validado arriba en la sección de permisos
            # Obtener el sector del recinto seleccionado
            recinto = Recinto.query.get(recinto_objetivo)
            if recinto and recinto.tiporecinto:
                sector_id = str(recinto.tiporecinto.id_sector)
            else:
                flash('Error: El recinto seleccionado no tiene sector asignado', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
                
        elif not is_superadmin and not is_administrador:
            # Si no es administrador ni superadmin, forzar que el trabajador sea del mismo recinto
            user_recinto_id = current_user.recinto_id
            if not user_recinto_id:
                flash('No tiene permisos para crear trabajadores (no tiene recinto asignado)', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            # Forzar el recinto del usuario
            recinto_id = str(user_recinto_id)
            
            # Obtener el sector del recinto del usuario
            if current_user.recinto and current_user.recinto.tiporecinto:
                sector_id = str(current_user.recinto.tiporecinto.id_sector)
            else:
                flash('Error: Su recinto no tiene sector asignado', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar que se haya seleccionado un sector
        if not sector_id:
            flash('Debe seleccionar un sector para el trabajador', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar RUT
        if not rut:
            flash('El RUT es obligatorio', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Formatear y validar RUT
        if not Trabajador.validate_rut(rut):
            flash('El RUT ingresado no es válido', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        rut_formateado = Trabajador.format_rut(rut)
        
        # Verificar si ya existe un trabajador con este RUT
        trabajador_existente = Trabajador.query.filter_by(rut=rut_formateado).first()
        if trabajador_existente:
            flash('Ya existe un trabajador con este RUT', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar campos de autenticación si se seleccionó crear credenciales
        if crear_credenciales:
            if not email:
                flash('El email es obligatorio para crear credenciales de acceso', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            if not password:
                flash('La contraseña es obligatoria para crear credenciales de acceso', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            if len(password) < 8:
                flash('La contraseña debe tener al menos 8 caracteres', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            # Verificar si ya existe un trabajador con este email
            trabajador_existente_email = Trabajador.query.filter_by(email=email).first()
            if trabajador_existente_email:
                flash('Ya existe un trabajador con este email', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar que el recinto pertenezca al sector seleccionado
        if recinto_id:
            recinto = Recinto.query.get(recinto_id)
            if not recinto:
                flash('El recinto seleccionado no existe', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            if recinto.sector_id != int(sector_id):
                flash('El recinto seleccionado no pertenece al sector elegido', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Determinar el rol a asignar (solo si se crean credenciales)
        rol_sistema = None
        rol_personalizado_id = None
        
        if crear_credenciales:
            # Obtener rol del usuario actual para validar permisos
            user_role = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol) if current_user.rol else None
            
            if rol_tipo == 'system' and user_role and user_role.upper() == 'SUPERADMIN':
                # Solo SUPERADMIN puede asignar rol de sistema
                from app.models import UserRole
                rol_sistema = UserRole.SUPERADMIN
            elif rol_tipo == 'custom' and custom_role_id:
                # Validar que el rol personalizado existe
                from app.models import CustomRole
                rol_personalizado = CustomRole.query.filter_by(id=custom_role_id, active=True).first()
                if rol_personalizado:
                    rol_personalizado_id = int(custom_role_id)
                else:
                    flash('El rol personalizado seleccionado no es válido', 'error')
                    return redirect(url_for('controllers.ruta_trabajadores'))
            elif rol_tipo == 'custom' and not custom_role_id:
                flash('Debe seleccionar un rol personalizado', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Crear el trabajador con sector, recinto y rol
        nuevo_trabajador = Trabajador(
            nombre=nombre,
            rut=rut_formateado,
            profesion=profesion,
            nombrecorto=nombrecorto,
            email=email,
            sector_id=int(sector_id),
            recinto_id=int(recinto_id) if recinto_id else None,
            rol=rol_sistema,
            custom_role_id=rol_personalizado_id
        )
        
        # Asignar contraseña solo si se crean credenciales
        if crear_credenciales and password:
            nuevo_trabajador.password = password
        
        db.session.add(nuevo_trabajador)
        db.session.commit()
        
        mensaje = f'Trabajador "{nombre}" agregado exitosamente'
        if nuevo_trabajador.recinto:
            mensaje += f' en recinto {nuevo_trabajador.recinto.nombre} (sector {nuevo_trabajador.sector.nombre})'
        else:
            mensaje += f' en sector {nuevo_trabajador.sector.nombre}'
        
        if crear_credenciales:
            rol_nombre = nuevo_trabajador.rol.display_name if nuevo_trabajador.rol else (nuevo_trabajador.custom_role.name if nuevo_trabajador.custom_role else 'Sin rol')
            mensaje += f'. Credenciales creadas con rol: {rol_nombre}'
        
        flash(mensaje, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar trabajador: {e}', 'error')
    return redirect(url_for('controllers.ruta_trabajadores'))

@controllers_bp.route('/admin/add_trabajador', methods=['POST'], endpoint='add_trabajador_admin')
@login_required
def add_trabajador_admin():
    """Agregar trabajador desde panel de administración - Solo SUPERADMIN"""
    
    # Verificar que sea SUPERADMIN
    user_role = current_user.rol.value if hasattr(current_user.rol, 'value') else str(current_user.rol)
    if user_role != 'superadmin':
        flash('Acceso denegado. Solo SUPERADMIN puede realizar esta acción.', 'error')
        return redirect(url_for('controllers.ruta_trabajadores'))
    
    try:
        nombre = request.form['name']
        rut = request.form.get('rut', '').strip()
        profesion = request.form.get('profesion', '')
        email = request.form.get('email', '')
        telefono = request.form.get('telefono', '')
        nombrecorto = request.form.get('nombrecorto', '')
        password = request.form.get('password', 'Maho#2024')
        sector_id = request.form.get('sector_id')
        recinto_id = request.form.get('recinto_id')
        
        # Validar que se haya seleccionado un sector
        if not sector_id:
            flash('Debe seleccionar un sector para el trabajador', 'error')
            return redirect(url_for('controllers.admin_trabajadores'))
        
        # Validar RUT
        if not rut:
            flash('El RUT es obligatorio', 'error')
            return redirect(url_for('controllers.admin_trabajadores'))
        
        # Formatear y validar RUT
        rut_clean = rut.replace('.', '').replace('-', '').replace(' ', '').upper()
        if len(rut_clean) < 8 or not rut_clean[:-1].isdigit():
            flash('El RUT ingresado no tiene un formato válido (mínimo 8 caracteres, números + dígito verificador)', 'error')
            return redirect(url_for('controllers.admin_trabajadores'))
        
        rut_formateado = Trabajador.format_rut(rut)
        
        # Verificar si ya existe un trabajador con este RUT
        trabajador_existente = Trabajador.query.filter_by(rut=rut_formateado).first()
        if trabajador_existente:
            flash('Ya existe un trabajador con este RUT', 'error')
            return redirect(url_for('controllers.admin_trabajadores'))
        
        # Validar que el recinto pertenezca al sector seleccionado
        if recinto_id:
            recinto = Recinto.query.get(recinto_id)
            if not recinto:
                flash('El recinto seleccionado no existe', 'error')
                return redirect(url_for('controllers.admin_trabajadores'))
            
            if recinto.sector_id != int(sector_id):
                flash('El recinto seleccionado no pertenece al sector elegido', 'error')
                return redirect(url_for('controllers.admin_trabajadores'))
        
        # Crear el trabajador
        nuevo_trabajador = Trabajador(
            nombre=nombre,
            rut=rut_formateado,
            profesion=profesion,
            email=email,
            telefono=telefono,
            nombrecorto=nombrecorto,
            password=password,
            sector_id=int(sector_id),
            recinto_id=int(recinto_id) if recinto_id else None
        )
        db.session.add(nuevo_trabajador)
        db.session.commit()
        
        mensaje = f'Trabajador agregado exitosamente en sector'
        if nuevo_trabajador.recinto:
            mensaje += f' y recinto {nuevo_trabajador.recinto.nombre}'
        flash(mensaje, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar trabajador: {e}', 'error')
    return redirect(url_for('controllers.admin_trabajadores'))

@controllers_bp.route('/admin/update_trabajador/<int:id>', methods=['POST'], endpoint='update_trabajador_admin')
@login_required
def update_trabajador_admin(id):
    """Actualizar trabajador desde panel de administración - Solo SUPERADMIN"""
    
    # Verificar que sea SUPERADMIN
    user_role = current_user.rol.value if hasattr(current_user.rol, 'value') else str(current_user.rol)
    if user_role != 'superadmin':
        flash('Acceso denegado. Solo SUPERADMIN puede realizar esta acción.', 'error')
        return redirect(url_for('controllers.ruta_trabajadores'))
    
    try:
        trabajador = Trabajador.query.get_or_404(id)
        
        # Actualizar campos básicos
        trabajador.nombre = request.form['name']
        trabajador.profesion = request.form.get('profesion', '')
        trabajador.email = request.form.get('email', '')
        trabajador.telefono = request.form.get('telefono', '')
        trabajador.nombrecorto = request.form.get('nombrecorto', '')
        
        # Actualizar RUT si se proporciona
        rut = request.form.get('rut', '').strip()
        if rut and rut != trabajador.rut:
            rut_clean = rut.replace('.', '').replace('-', '').replace(' ', '').upper()
            if len(rut_clean) < 8 or not rut_clean[:-1].isdigit():
                flash('El RUT ingresado no tiene un formato válido', 'error')
                return redirect(url_for('controllers.admin_trabajadores'))
            
            rut_formateado = Trabajador.format_rut(rut)
            # Verificar que no exista otro trabajador con el mismo RUT
            trabajador_existente = Trabajador.query.filter(
                Trabajador.rut == rut_formateado,
                Trabajador.id != id
            ).first()
            if trabajador_existente:
                flash('Ya existe otro trabajador con este RUT', 'error')
                return redirect(url_for('controllers.admin_trabajadores'))
            
            trabajador.rut = rut_formateado
        
        # Actualizar contraseña si se proporciona
        password = request.form.get('password', '').strip()
        if password:
            trabajador.password = password
        
        # Actualizar sector y recinto
        sector_id = request.form.get('sector_id')
        recinto_id = request.form.get('recinto_id')
        
        if not sector_id:
            flash('Debe seleccionar un sector', 'error')
            return redirect(url_for('controllers.admin_trabajadores'))
        
        # Validar que el recinto pertenezca al sector seleccionado
        if recinto_id:
            recinto = Recinto.query.get(recinto_id)
            if not recinto:
                flash('El recinto seleccionado no existe', 'error')
                return redirect(url_for('controllers.admin_trabajadores'))
            
            if recinto.sector_id != int(sector_id):
                flash('El recinto seleccionado no pertenece al sector elegido', 'error')
                return redirect(url_for('controllers.admin_trabajadores'))
        
        trabajador.sector_id = int(sector_id)
        trabajador.recinto_id = int(recinto_id) if recinto_id else None
        
        db.session.commit()
        
        mensaje = 'Trabajador actualizado exitosamente'
        if trabajador.recinto:
            mensaje += f' en sector {trabajador.sector.nombre} y recinto {trabajador.recinto.nombre}'
        else:
            mensaje += f' en sector {trabajador.sector.nombre}'
        flash(mensaje, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar trabajador: {e}', 'error')
    
    return redirect(url_for('controllers.admin_trabajadores'))

@controllers_bp.route('/update_trabajador/<int:id>', methods=['POST'], endpoint='update_trabajador')
@login_required
def update_trabajador(id):
    from app.utils.area_permissions import puede_editar_trabajador, puede_crear_trabajador_en_area, puede_editar_trabajador_recinto, puede_crear_trabajador_en_recinto
    
    try:
        trabajador = Trabajador.query.get_or_404(id)
        
        # Verificar permisos para editar este trabajador
        from app.models import CustomRole, AdministradorRecinto
        
        # Verificar si es SUPERADMIN
        is_superadmin = False
        if hasattr(current_user, 'rol') and current_user.rol:
            rol_name = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol)
            if rol_name and rol_name == 'SUPERADMIN':
                is_superadmin = True
        
        # Verificar si es administrador con recintos asignados
        is_administrador = False
        if hasattr(current_user, 'custom_role_id') and current_user.custom_role_id:
            custom_role = CustomRole.query.get(current_user.custom_role_id)
            if custom_role and custom_role.name.upper() == 'ADMINISTRADOR':
                is_administrador = True
        
        # Validar permisos según tipo de usuario
        puede_editar = False
        
        if is_superadmin:
            # SUPERADMIN puede editar cualquier trabajador
            puede_editar = True
        elif is_administrador:
            # ADMINISTRADOR solo puede editar trabajadores de sus recintos asignados
            if trabajador.recinto_id:
                puede_editar = AdministradorRecinto.tiene_acceso_recinto(current_user.id, trabajador.recinto_id)
            else:
                # Si el trabajador no tiene recinto asignado, no puede editarlo
                puede_editar = False
        else:
            # Otros usuarios usan la lógica existente
            puede_editar = puede_editar_trabajador_recinto(current_user, trabajador)
        
        if not puede_editar:
            flash('No tienes permisos para editar este trabajador', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Verificar nueva área si se está cambiando
        new_area_principal_id = request.form.get('area_id')
        if new_area_principal_id and new_area_principal_id.strip():
            new_area_principal_id = int(new_area_principal_id)
            if new_area_principal_id != trabajador.area_id and not puede_crear_trabajador_en_area(current_user, new_area_principal_id):  # TEMPORAL: usando area_id
                flash('No tienes permisos para asignar trabajadores a esta área', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        else:
            new_area_principal_id = None
        
        # Validar que trabajador tenga área asignada (excepto usuarios con permisos administrativos)
        if not new_area_principal_id and not user_has_admin_permissions(current_user):
            flash('Todo trabajador debe tener un área asignada', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        trabajador.nombre = request.form['name']
        trabajador.profesion = request.form['profesion']
        trabajador.nombrecorto = request.form['nombrecorto']
        
        # Actualizar email si se proporciona
        new_email = request.form.get('email', '').strip()
        if new_email:
            # Verificar que no exista otro trabajador con este email
            trabajador_existente_email = Trabajador.query.filter(
                Trabajador.email == new_email,
                Trabajador.id != trabajador.id
            ).first()
            if trabajador_existente_email:
                flash('Ya existe otro trabajador con este email', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            trabajador.email = new_email
        elif not new_email:
            # Si se vacía el email, remover credenciales
            trabajador.email = None
            trabajador.password_hash = None
            trabajador.rol = None
            trabajador.custom_role_id = None
        
        # Solo actualizar password si se proporciona uno nuevo
        new_password = request.form.get('password')
        if new_password and new_password.strip():
            if len(new_password) < 8:
                flash('La contraseña debe tener al menos 8 caracteres', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            trabajador.password = new_password
        
        # Actualizar sector y recinto
        sector_id = request.form.get('sector_id')
        recinto_id = request.form.get('recinto_id')
        
        if not sector_id:
            flash('Debe seleccionar un sector', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar permisos para cambio de recinto
        if recinto_id and int(recinto_id) != trabajador.recinto_id:
            nuevo_recinto_id = int(recinto_id)
            puede_mover = False
            
            if is_superadmin:
                # SUPERADMIN puede mover a cualquier recinto
                puede_mover = True
            elif is_administrador:
                # ADMINISTRADOR solo puede mover a recintos donde tiene asignación
                puede_mover = AdministradorRecinto.tiene_acceso_recinto(current_user.id, nuevo_recinto_id)
            else:
                # Otros usuarios usan la lógica existente
                puede_mover = puede_crear_trabajador_en_recinto(current_user, nuevo_recinto_id)
            
            if not puede_mover:
                flash('No tienes permisos para mover trabajadores al recinto seleccionado', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Validar que el recinto pertenezca al sector seleccionado
        if recinto_id:
            recinto = Recinto.query.get(recinto_id)
            if not recinto:
                flash('El recinto seleccionado no existe', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
            
            if recinto.sector_id != int(sector_id):
                flash('El recinto seleccionado no pertenece al sector elegido', 'error')
                return redirect(url_for('controllers.ruta_trabajadores'))
        
        trabajador.sector_id = int(sector_id)
        trabajador.recinto_id = int(recinto_id) if recinto_id else None
        
        # Actualizar rol solo si tiene email (credenciales de acceso)
        if trabajador.email:
            rol_tipo = request.form.get('rol_tipo')
            custom_role_id = request.form.get('custom_role_id')
            
            # Limpiar roles anteriores
            trabajador.rol = None
            trabajador.custom_role_id = None
            
            # Asignar nuevo rol
            if rol_tipo == 'system' and user_has_admin_permissions(current_user):
                # Solo usuarios con permisos administrativos pueden asignar rol de sistema
                from app.models import UserRole
                trabajador.rol = UserRole.SUPERADMIN
            elif rol_tipo == 'custom' and custom_role_id:
                # Validar que el rol personalizado existe
                from app.models import CustomRole
                rol_personalizado = CustomRole.query.filter_by(id=custom_role_id, active=True).first()
                if rol_personalizado:
                    trabajador.custom_role_id = int(custom_role_id)
                else:
                    flash('El rol personalizado seleccionado no es válido', 'error')
                    return redirect(url_for('controllers.ruta_trabajadores'))
        
        # Actualizar área principal (mantener compatibilidad temporal)
        trabajador.area_id = new_area_principal_id  # TEMPORAL: usando area_id hasta migración
            
        db.session.commit()
        
        mensaje = 'Trabajador actualizado exitosamente'
        if trabajador.recinto:
            mensaje += f' en recinto {trabajador.recinto.nombre} (sector {trabajador.sector.nombre})'
        else:
            mensaje += f' en sector {trabajador.sector.nombre}'
        flash(mensaje, 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar trabajador: {e}', 'error')
    return redirect(url_for('controllers.ruta_trabajadores'))

@controllers_bp.route('/eliminar_trabajador/<int:id>', methods=['POST'], endpoint='eliminar_trabajador')
@login_required
def eliminar_trabajador(id):
    from app.utils.area_permissions import puede_editar_trabajador
    
    try:
        trabajador = Trabajador.query.get_or_404(id)
        
        # Verificar permisos para eliminar este trabajador
        from app.models import CustomRole, AdministradorRecinto
        
        # Verificar si es SUPERADMIN
        is_superadmin = False
        if hasattr(current_user, 'rol') and current_user.rol:
            rol_name = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol)
            if rol_name and rol_name == 'SUPERADMIN':
                is_superadmin = True
        
        # Verificar si es administrador con recintos asignados
        is_administrador = False
        if hasattr(current_user, 'custom_role_id') and current_user.custom_role_id:
            custom_role = CustomRole.query.get(current_user.custom_role_id)
            if custom_role and custom_role.name.upper() == 'ADMINISTRADOR':
                is_administrador = True
        
        # Validar permisos según tipo de usuario
        puede_eliminar = False
        
        if is_superadmin:
            # SUPERADMIN puede eliminar cualquier trabajador
            puede_eliminar = True
        elif is_administrador:
            # ADMINISTRADOR solo puede eliminar trabajadores de sus recintos asignados
            if trabajador.recinto_id:
                puede_eliminar = AdministradorRecinto.tiene_acceso_recinto(current_user.id, trabajador.recinto_id)
            else:
                # Si el trabajador no tiene recinto asignado, no puede eliminarlo
                puede_eliminar = False
        else:
            # Otros usuarios usan la lógica existente
            puede_eliminar = puede_editar_trabajador(current_user, trabajador)
        
        if not puede_eliminar:
            flash('No tienes permisos para eliminar este trabajador', 'error')
            return redirect(url_for('controllers.ruta_trabajadores'))
        
        db.session.delete(trabajador)
        db.session.commit()
        flash('Trabajador eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar trabajador: {e}', 'error')
    return redirect(url_for('controllers.ruta_trabajadores'))

@controllers_bp.route('/api/reset_password/<int:id>', methods=['POST'], endpoint='reset_password')
@login_required
def reset_password(id):
    """Resetear contraseña de un trabajador - Solo SUPERADMIN"""
    try:
        # Verificar que sea SUPERADMIN
        user_role = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol) if current_user.rol else None
        if not user_role or user_role.upper() != 'SUPERADMIN':
            return jsonify({'success': False, 'message': 'Acceso denegado. Solo SUPERADMIN puede resetear contraseñas.'}), 403
        
        trabajador = Trabajador.query.get_or_404(id)
        
        # Verificar que el trabajador tenga email (credenciales)
        if not trabajador.email:
            return jsonify({'success': False, 'message': 'Este trabajador no tiene credenciales de acceso.'}), 400
        
        # Generar nueva contraseña temporal
        import secrets
        import string
        
        # Generar contraseña segura: 8 caracteres con letras, números y símbolos
        alphabet = string.ascii_letters + string.digits + "!@#$%&*"
        nueva_password = ''.join(secrets.choice(alphabet) for _ in range(12))
        
        # Asignar nueva contraseña
        trabajador.password = nueva_password
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': 'Contraseña reseteada exitosamente',
            'new_password': nueva_password
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================================================================================
# Rutas CRUD para Financiamiento 
@controllers_bp.route('/financiamientos', endpoint='ruta_financiamientos')
@login_required
def financiamientos():
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/financiamientos')):
        flash('No tiene permisos para acceder a esta página', 'error')
        return redirect(url_for('main.dashboard'))
    
    financiamientos = Financiamiento.query.all()
    return render_template('financiamientos.html', financiamientos=financiamientos)

@controllers_bp.route('/add_financiamiento', methods=['POST'], endpoint='add_financiamiento')
@login_required
def add_financiamiento():
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/financiamientos')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    from flask import session
    
    # Verificar token para prevenir doble envío
    form_token = request.form.get('form_token')
    session_token = session.get('last_add_form_token')
    
    if form_token and form_token == session_token:
        flash('Operación ya procesada', 'warning')
        return redirect(url_for('controllers.ruta_financiamientos'))
    
    try:
        nombre = request.form['name']
        descripcion = request.form.get('descripcion', '')  # valor por defecto vacío
        
        nuevo_financiamiento = Financiamiento(
            nombre=nombre,
            descripcion=descripcion
        )
        db.session.add(nuevo_financiamiento)
        db.session.commit()
        
        # Guardar token para prevenir doble procesamiento
        if form_token:
            session['last_add_form_token'] = form_token
            
        flash('Financiamiento agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar financiamiento: {e}')
    return redirect(url_for('controllers.ruta_financiamientos'))

@controllers_bp.route('/update_financiamiento/<int:id>', methods=['POST'], endpoint='update_financiamiento')
@login_required
def update_financiamiento(id):
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/financiamientos')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    from flask import session
    
    # Verificar token para prevenir doble envío
    form_token = request.form.get('form_token')
    session_token = session.get('last_form_token')
    
    if form_token and form_token == session_token:
        flash('Operación ya procesada', 'warning')
        return redirect(url_for('controllers.ruta_financiamientos'))
    
    try:
        financiamiento = Financiamiento.query.get_or_404(id)
        financiamiento.nombre = request.form['name']
        financiamiento.descripcion = request.form['descripcion']
        db.session.commit()
        
        # Guardar token para prevenir doble procesamiento
        if form_token:
            session['last_form_token'] = form_token
            
        flash('Financiamiento actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar financiamiento: {e}')
    return redirect(url_for('controllers.ruta_financiamientos'))

@controllers_bp.route('/eliminar_financiamiento/<int:id>', methods=['POST'], endpoint='eliminar_financiamiento')
@login_required
def eliminar_financiamiento(id):
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/financiamientos')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        financiamiento = Financiamiento.query.get_or_404(id)
        db.session.delete(financiamiento)
        db.session.commit()
        flash('Financiamiento eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar financiamiento: {e}')
    return redirect(url_for('controllers.ruta_financiamientos'))

# ==================================================================================
# Rutas CRUD para Especialidad
@controllers_bp.route('/especialidades', endpoint='ruta_especialidades')
@login_required
def especialidades():
    especialidades = Especialidad.query.all()
    return render_template('especialidades.html', especialidades=especialidades)  # Cambiar de 'especialidad.html' a 'especialidades.html'

@controllers_bp.route('/add_especialidad', methods=['POST'], endpoint='add_especialidad')
@login_required
def add_especialidad():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_especialidad_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_especialidades'))
        
        nombre = request.form['name']
        nuevo_especialidad = Especialidad(nombre=nombre)
        db.session.add(nuevo_especialidad)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_especialidad_form_token'] = form_token
        
        flash('Especialidad agregada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar especialidad: {e}')
    return redirect(url_for('controllers.ruta_especialidades'))

@controllers_bp.route('/update_especialidad/<int:id>', methods=['POST'], endpoint='update_especialidad')
@login_required
def update_especialidad(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_especialidad_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_especialidades'))
        
        especialidad = Especialidad.query.get_or_404(id)
        especialidad.nombre = request.form['name']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_especialidad_form_token'] = form_token
        
        flash('Especialidad actualizada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar especialidad: {e}')
    return redirect(url_for('controllers.ruta_especialidades'))

@controllers_bp.route('/eliminar_especialidad/<int:id>', methods=['POST'], endpoint='eliminar_especialidad')
@login_required
def eliminar_especialidad(id):
    try:
        especialidad = Especialidad.query.get_or_404(id)
        db.session.delete(especialidad)
        db.session.commit()
        flash('Especialidad eliminada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar especialidad: {e}')
    return redirect(url_for('controllers.ruta_especialidades'))

# ==================================================================================
# Rutas CRUD para Áreas
@controllers_bp.route('/areas', endpoint='ruta_areas')
@login_required
def areas():
    areas = Area.query.all()
    return render_template('areas.html', areas=areas)

@controllers_bp.route('/add_area', methods=['POST'], endpoint='add_area')
@login_required
def add_area():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_area_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_areas'))
        
        nombre = request.form['name']
        descripcion = request.form.get('descripcion', '')
        nueva_area = Area(nombre=nombre, descripcion=descripcion)
        db.session.add(nueva_area)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_area_form_token'] = form_token
        
        flash('Área agregada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar área: {e}', 'error')
    return redirect(url_for('controllers.ruta_areas'))

@controllers_bp.route('/update_area/<int:id>', methods=['POST'], endpoint='update_area')
@login_required
def update_area(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_area_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_areas'))
        
        area = Area.query.get_or_404(id)
        area.nombre = request.form['name']
        area.descripcion = request.form.get('descripcion', '')
        # Manejar el campo activo
        area.activo = request.form.get('activo') == '1'
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_area_form_token'] = form_token
        
        flash('Área actualizada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar área: {e}', 'error')
    return redirect(url_for('controllers.ruta_areas'))

@controllers_bp.route('/eliminar_area/<int:id>', methods=['POST'], endpoint='eliminar_area')
@login_required
def eliminar_area(id):
    try:
        area = Area.query.get_or_404(id)
        # Verificar si hay trabajadores asignados a esta área
        if area.trabajadores.count() > 0:
            flash(f'No se puede eliminar el área "{area.nombre}" porque tiene trabajadores asignados', 'error')
        else:
            db.session.delete(area)
            db.session.commit()
            flash('Área eliminada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar área: {e}', 'error')
    return redirect(url_for('controllers.ruta_areas'))

# ==================================================================================
# Rutas CRUD para Equipos
@controllers_bp.route('/equipos', endpoint='ruta_equipos')
@login_required
def equipos():
    equipos = Equipo.query.all()
    return render_template('equipo.html', equipos=equipos)

@controllers_bp.route('/add_equipo', methods=['POST'], endpoint='add_equipo')
@login_required
def add_equipo():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_equipo_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_equipos'))
        
        nombre = request.form['name']
        nuevo_equipo = Equipo(nombre=nombre)
        db.session.add(nuevo_equipo)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_equipo_form_token'] = form_token
        
        flash('Equipo agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar equipo: {e}')
    return redirect(url_for('controllers.ruta_equipos'))

@controllers_bp.route('/update_equipo/<int:id>', methods=['POST'], endpoint='update_equipo')
@login_required
def update_equipo(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_equipo_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_equipos'))
        
        equipo = Equipo.query.get_or_404(id)
        equipo.nombre = request.form['name']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_equipo_form_token'] = form_token
        
        flash('Equipo actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar equipo: {e}')
    return redirect(url_for('controllers.ruta_equipos'))

@controllers_bp.route('/eliminar_equipo/<int:id>', methods=['POST'], endpoint='eliminar_equipo')
@login_required
def eliminar_equipo(id):
    try:
        equipo = Equipo.query.get_or_404(id)
        db.session.delete(equipo)
        db.session.commit()
        flash('Equipo eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar equipo: {e}')
    return redirect(url_for('controllers.ruta_equipos'))

# ==================================================================================
# Rutas CRUD para Tipologías
@controllers_bp.route('/tipologias', endpoint='ruta_tipologias')
@login_required
def tipologias():
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/tipologias')):
        flash('No tiene permisos para acceder a esta página', 'error')
        return redirect(url_for('main.dashboard'))
    
    tipologias = Tipologia.query.all()
    fases = Fase.query.filter_by(activo=True).all()  # Cargar fases activas
    return render_template('tipologias.html', tipologias=tipologias, fases=fases)

# Agregar endpoint para obtener fases en formato JSON
@controllers_bp.route('/get_tipologias', methods=['GET'])
@login_required
def get_tipologias():
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/tipologias')):
        return jsonify({'error': 'Sin permisos'}), 403
    
    tipologias = Tipologia.query.all()
    return jsonify([{'id': e.id, 'nombre': e.nombre} for e in tipologias])

@controllers_bp.route('/add_tipologia', methods=['POST'], endpoint='add_tipologia')
@login_required
def add_tipologia():
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/tipologias')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_tipologia_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_tipologias'))
        
        nombre = request.form['nombre']
        nombrecorto = request.form.get('nombrecorto', '')  # valor por defecto vacío
        id_fase = request.form.get('id_fase', '1')  # valor por defecto 1
        
        nueva_tipologia = Tipologia(
            nombre=nombre,
            nombrecorto=nombrecorto,
            id_fase=id_fase  
        )
        db.session.add(nueva_tipologia)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_tipologia_form_token'] = form_token
        
        flash('Tipologia agregada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar tipologia: {e}')
    return redirect(url_for('controllers.ruta_tipologias'))

@controllers_bp.route('/update_tipologia/<int:id>', methods=['POST'], endpoint='update_tipologia')
@login_required
def update_tipologia(id):
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/tipologias')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_tipologia_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_tipologias'))
        
        tipologia = Tipologia.query.get_or_404(id)
        tipologia.nombre = request.form['nombre']
        tipologia.nombrecorto = request.form['nombrecorto']
        tipologia.id_fase = request.form['id_fase']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_tipologia_form_token'] = form_token
        
        flash('Tipologia actualizada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar tipologia: {e}')
    return redirect(url_for('controllers.ruta_tipologias'))

@controllers_bp.route('/eliminar_tipologia/<int:id>', methods=['POST'], endpoint='eliminar_tipologia')
@login_required
def eliminar_tipologia(id):
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/tipologias')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        tipologia = Tipologia.query.get_or_404(id)
        db.session.delete(tipologia)
        db.session.commit()
        flash('Tipologia eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar tipologia: {e}')
    return redirect(url_for('controllers.ruta_tipologias'))



# ==================================================================================
# Rutas CRUD para Prioridades
@controllers_bp.route('/prioridades', endpoint='ruta_prioridades')
def prioridades():
    prioridades = Prioridad.query.order_by(Prioridad.orden).all()
    return render_template('prioridades.html', prioridades=prioridades)

@controllers_bp.route('/add_prioridad', methods=['POST'], endpoint='add_prioridad')
def add_prioridad():
    try:
        # Verificar token de formulario para prevenir envíos duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_prioridad_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_prioridades'))
        
        nombre = request.form['name']
        descripcion = request.form.get('descripcion', '')
        urgencia = 'urgencia' in request.form
        importancia = 'importancia' in request.form
        color = request.form.get('color', '#007bff')
        
        nueva_prioridad = Prioridad(
            nombre=nombre,
            descripcion=descripcion,
            urgencia=urgencia,
            importancia=importancia,
            color=color
        )
        db.session.add(nueva_prioridad)
        db.session.commit()
        
        # Guardar token para prevenir reenvíos
        if form_token:
            session['last_add_prioridad_form_token'] = form_token
        
        flash('Prioridad agregada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar prioridad: {e}')
    return redirect(url_for('controllers.ruta_prioridades'))

@controllers_bp.route('/update_prioridad/<int:id>', methods=['POST'], endpoint='update_prioridad')
def update_prioridad(id):
    try:
        # Verificar token de formulario para prevenir envíos duplicados
        form_token = request.form.get('form_token')
        session_token = session.get(f'last_edit_prioridad_{id}_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_prioridades'))
        
        prioridad = Prioridad.query.get_or_404(id)
        prioridad.nombre = request.form['name']
        prioridad.descripcion = request.form.get('descripcion', '')
        prioridad.urgencia = 'urgencia' in request.form
        prioridad.importancia = 'importancia' in request.form
        prioridad.color = request.form.get('color', '#007bff')
        
        db.session.commit()
        
        # Guardar token para prevenir reenvíos
        if form_token:
            session[f'last_edit_prioridad_{id}_form_token'] = form_token
        
        flash('Prioridad actualizada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar prioridad: {e}')
    return redirect(url_for('controllers.ruta_prioridades'))

@controllers_bp.route('/eliminar_prioridad/<int:id>', methods=['POST'], endpoint='eliminar_prioridad')
def eliminar_prioridad(id):
    try:
        prioridad = Prioridad.query.get_or_404(id)
        db.session.delete(prioridad)
        db.session.commit()
        flash('Prioridad eliminada exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar prioridad: {e}')
    return redirect(url_for('controllers.ruta_prioridades'))

# ==================================================================================
# Rutas CRUD para Tipo de Proyecto 
@controllers_bp.route('/tipoproyectos', endpoint='ruta_tipoproyectos')
@login_required
def tipoproyectos():
    tipoproyectos = TipoProyecto.query.all()
    return render_template('tipoproyectos.html', tipoproyectos=tipoproyectos)

@controllers_bp.route('/add_tipoproyecto', methods=['POST'], endpoint='add_tipoproyecto')
@login_required
def add_tipoproyecto():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_tipoproyecto_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_tipoproyectos'))
        
        nombre = request.form['name']
        nombrecorto = request.form.get('namecorto', '')  # valor por defecto vacío
        descripcion = request.form.get('descripcion', '')  # valor por defecto vacío
        
        nuevo_tipoproyecto = TipoProyecto(
            nombre=nombre,
            nombrecorto=nombrecorto,
            descripcion=descripcion
        )
        db.session.add(nuevo_tipoproyecto)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_tipoproyecto_form_token'] = form_token
        
        flash('Tipo de Proyecto agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar tipo de proyecto: {e}')
    return redirect(url_for('controllers.ruta_tipoproyectos'))

@controllers_bp.route('/update_tipoproyecto/<int:id>', methods=['POST'], endpoint='update_tipoproyecto')
@login_required
def update_tipoproyecto(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_tipoproyecto_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_tipoproyectos'))
        
        tipoproyecto = TipoProyecto.query.get_or_404(id)
        tipoproyecto.nombre = request.form['name']
        tipoproyecto.nombrecorto = request.form['namecorto']
        tipoproyecto.descripcion = request.form['descripcion']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_tipoproyecto_form_token'] = form_token
        
        flash('Tipo de Proyecto actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar tipo de proyecto: {e}')
    return redirect(url_for('controllers.ruta_tipoproyectos'))

@controllers_bp.route('/eliminar_tipoproyecto/<int:id>', methods=['POST'], endpoint='eliminar_tipoproyecto')
@login_required
def eliminar_tipoproyecto(id):
    try:
        tipoproyecto = TipoProyecto.query.get_or_404(id)
        db.session.delete(tipoproyecto)
        db.session.commit()
        flash('Tipo de Proyecto eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar tipo de proyecto: {e}')
    return redirect(url_for('controllers.ruta_tipoproyectos'))

# ==================================================================================
# Rutas CRUD para Estados
@controllers_bp.route('/estados', endpoint='ruta_estados')
@login_required
def estados():
    estados = Estado.query.all()
    return render_template('estados.html', estados=estados)

@controllers_bp.route('/add_estado', methods=['POST'], endpoint='add_estado')
@login_required
def add_estado():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_estado_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_estados'))
        
        nombre = request.form['name']
        nuevo_estado = Estado(nombre=nombre)
        db.session.add(nuevo_estado)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_estado_form_token'] = form_token
        
        flash('Estado agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar estado: {e}')
    return redirect(url_for('controllers.ruta_estados'))

@controllers_bp.route('/update_estado/<int:id>', methods=['POST'], endpoint='update_estado')
@login_required
def update_estado(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_estado_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_estados'))
        
        estado = Estado.query.get_or_404(id)
        estado.nombre = request.form['name']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_estado_form_token'] = form_token
        
        flash('Estado actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar estado: {e}')
    return redirect(url_for('controllers.ruta_estados'))

@controllers_bp.route('/eliminar_estado/<int:id>', methods=['POST'], endpoint='eliminar_estado')
@login_required
def eliminar_estado(id):
    try:
        estado = Estado.query.get_or_404(id)
        db.session.delete(estado)
        db.session.commit()
        flash('Estado eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar estado: {e}')
    return redirect(url_for('controllers.ruta_estados'))

# ==================================================================================
# Ruta para la página "Acerca de mí"
@controllers_bp.route('/acerca-de-mi', endpoint='ruta_acerca_de_mi')
def about_me():
    estados = Estado.query.all()
    return render_template('about_me.html', estados=estados)

@controllers_bp.route('/add_aboutme', methods=['POST'], endpoint='add_aboutme')
def add_aboutme():
    try:
        nombre = request.form['name']
        nuevo_estado = Estado(nombre=nombre)
        db.session.add(nuevo_estado)
        db.session.commit()
        flash('Estado agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar estado: {e}')
    return redirect(url_for('controllers.ruta_acerca_de_mi'))

@controllers_bp.route('/update_aboutme/<int:id>', methods=['POST'], endpoint='update_aboutme')
def update_aboutme(id):
    try:
        estado = Estado.query.get_or_404(id)
        estado.nombre = request.form['name']
        db.session.commit()
        flash('Estado actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar estado: {e}')
    return redirect(url_for('controllers.ruta_acerca_de_mi'))

@controllers_bp.route('/eliminar_aboutme/<int:id>', methods=['POST'], endpoint='eliminar_aboutme')
def eliminar_aboutme(id):
    try:
        estado = Estado.query.get_or_404(id)
        db.session.delete(estado)
        db.session.commit()
        flash('Estado eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar estado: {e}')
    return redirect(url_for('controllers.ruta_acerca_de_mi'))

# ==================================================================================
# Ruta para la página "Contacto"
@controllers_bp.route('/contacto', endpoint='ruta_contacto')
def contact():
    estados = Estado.query.all()
    return render_template('contact.html', estados=estados)






@controllers_bp.route('/update_requerimiento_rechazar/<int:id>', methods=['POST'])
@login_required
def update_requerimiento_rechazar(id):
    """
    Endpoint para rechazar un requerimiento con verificación de permisos moderna
    """
    print(f"🔴 INICIO ENDPOINT: update_requerimiento_rechazar ID={id}")
    print(f"🔍 Usuario autenticado: {current_user.is_authenticated}")
    print(f"🔍 Usuario: {current_user.email if current_user.is_authenticated else 'No autenticado'}")
    print(f"🔍 Método: {request.method}")
    print(f"🔍 Form data: {dict(request.form) if request.form else 'Sin form data'}")
    
    # Verificar JSON data de forma segura
    try:
        json_data = request.get_json()
        print(f"🔍 JSON data: {json_data}")
    except Exception as json_error:
        print(f"🔍 JSON data error: {json_error}")
    
    try:
        # Verificar permisos usando el sistema unificado
        print(f"🔐 Verificando permisos...")
        print(f"🔐 is_superadmin: {current_user.is_superadmin() if hasattr(current_user, 'is_superadmin') else 'No method'}")
        print(f"🔐 has_page_permission: {current_user.has_page_permission('/requerimientos_aceptar') if hasattr(current_user, 'has_page_permission') else 'No method'}")
        
        if not (current_user.is_superadmin() or current_user.has_page_permission('/requerimientos_aceptar')):
            print("❌ ERROR: Usuario sin permisos")
            return jsonify({'error': 'No tiene permisos para rechazar requerimientos'}), 403
        
        print("✅ Permisos verificados correctamente")
        
        # Verificar que los datos del formulario estén presentes
        print(f"🔍 Content-Type: {request.content_type}")
        print(f"🔍 Request headers: {dict(request.headers)}")
        
        print(f"🔍 Buscando requerimiento ID={id}")
        requerimiento = Requerimiento.query.get_or_404(id)
        print(f"✅ Requerimiento encontrado: {requerimiento.nombre}")
        
        observacion_texto = request.form.get('observacion', '').strip()
        print(f"📝 Observación recibida: '{observacion_texto}'")
        
        # Actualizar estado a rechazado
        requerimiento.id_estado = 9  # Estado: Rechazado
        # No establecer fecha_aceptacion para rechazados
        
        # Solo actualizar observación si hay texto (mantener compatibilidad con sistema anterior)
        if observacion_texto:
            requerimiento.observacion = observacion_texto
        
        # Guardar observación en historial si hay texto
        if observacion_texto:
            try:
                from app.models import ObservacionRequerimiento
                nueva_observacion = ObservacionRequerimiento(
                    id_requerimiento=id,
                    observacion=observacion_texto,
                    fecha_registro=datetime.now(),
                    id_usuario=current_user.id,
                    pagina_origen='requerimientos_aceptar',
                    tipo_evento='rechazado'
                )
                db.session.add(nueva_observacion)
            except Exception as obs_error:
                print(f"⚠️ Error al guardar observación: {obs_error}")
                # Continuar sin fallar el proceso principal
        
        db.session.commit()
        print(f"✅ Requerimiento ID {id} rechazado exitosamente por usuario {current_user.email}")
        flash('Requerimiento rechazado y actualizado correctamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error al rechazar requerimiento ID {id}: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al rechazar requerimiento: {str(e)}', 'error')
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500
        
    return redirect(url_for('controllers.ruta_requerimientos_aceptar'))

# Endpoint de prueba para diagnosticar problemas
@controllers_bp.route('/test_rechazar/<int:id>', methods=['POST', 'GET'])
def test_rechazar(id):
    """Endpoint de prueba para diagnosticar el problema de rechazo"""
    print(f"🧪 TEST ENDPOINT - ID: {id}, Método: {request.method}")
    print(f"🧪 Usuario autenticado: {current_user.is_authenticated}")
    print(f"🧪 Usuario: {current_user.email if current_user.is_authenticated else 'Anónimo'}")
    print(f"🧪 Form: {dict(request.form)}")
    
    # Verificar JSON de forma segura
    json_data = None
    try:
        json_data = request.get_json()
        print(f"🧪 JSON: {json_data}")
    except Exception as json_error:
        print(f"🧪 JSON error: {json_error}")
    
    return jsonify({
        'success': True,
        'message': 'Endpoint de prueba funcionando',
        'id': id,
        'method': request.method,
        'authenticated': current_user.is_authenticated,
        'user': current_user.email if current_user.is_authenticated else 'anonymous',
        'form': dict(request.form),
        'json_data': json_data
    })

# ==================================================================================
# Rutas CRUD para Requerimientos-completar
# NOTA: Esta funcionalidad se movió al blueprint de requerimientos en:
# app/controllers/requerimientos_controller.py
# 
# @controllers_bp.route('/requerimientos_completar', endpoint='ruta_requerimientos_completar')
# def requerimientos_completar():
#     # Obtener solo los requerimientos en estado "En Desarrollo - Preparación" (id_estado = 2)
#     requerimientos = Requerimiento.query.filter_by(id_estado=2).\
#         order_by(Requerimiento.fecha_aceptacion.desc()).all()
#         
#     tipologias = Tipologia.query.order_by(Tipologia.id).all()
#     financiamientos = Financiamiento.query.all()
#     tipoproyectos = TipoProyecto.query.all()
#     trabajadores = Trabajador.query.all()
#     especialidades = Especialidad.query.all()
#     prioridades = Prioridad.query.order_by(Prioridad.orden).all()
#     grupos = Grupo.query.filter_by(activo=True).order_by(Grupo.nombre).all()
#     
#     return render_template('requerimiento-completar.html',
#                          requerimientos=requerimientos,
#                          tipologias=tipologias,
#                          financiamientos=financiamientos,
#                          tipoproyectos=tipoproyectos,
#                          trabajadores=trabajadores,
#                          especialidades=especialidades,
#                          prioridades=prioridades,
#                          grupos=grupos)

# Agregar nueva ruta para manejar la adición de trabajadores
@controllers_bp.route('/agregar_trabajador_requerimiento', methods=['POST'])
def agregar_trabajador_requerimiento():
    try:
        data = request.get_json()
        print("Datos recibidos:", data)  # Para debug
        requerimiento = Requerimiento.query.get_or_404(data['id_requerimiento'])
        
        if data.get('es_nuevo'):
            # Crear nuevo trabajador con campos opcionales
            trabajador = Trabajador(
                nombre=data['nombre'],
                profesion=data.get('profesion') or '',  # Si es None o vacío, usar string vacío
                nombrecorto=data.get('nombre_corto') or ''  # Si es None o vacío, usar string vacío
            )
            db.session.add(trabajador)
            db.session.flush()  # Obtener el ID antes del commit
        else:
            trabajador = Trabajador.query.get_or_404(data['trabajador_id'])
        
        # Crear nuevo equipo de trabajo
        equipo = EquipoTrabajo(
            id_requerimiento=data['id_requerimiento'],
            id_trabajador=trabajador.id,
            id_especialidad=data['especialidad_id']
        )
        
        db.session.add(equipo)
        db.session.commit()
        
        # Obtener la especialidad para la respuesta
        especialidad = Especialidad.query.get_or_404(data['especialidad_id'])
        
        return jsonify({
            'success': True,
            'message': 'Trabajador agregado exitosamente',
            'trabajador': {
                'id': trabajador.id,
                'nombre': trabajador.nombre,
                'profesion': trabajador.profesion or '',
                'especialidad': especialidad.nombre
            }
        })
    except Exception as e:
        print("Error:", str(e))  # Para debug
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@controllers_bp.route('/quitar_trabajador_requerimiento/<int:id_req>/<int:id_trab>', methods=['POST'])
def quitar_trabajador_requerimiento(id_req, id_trab):
    try:
        requerimiento = Requerimiento.query.get_or_404(id_req)
        trabajador = Trabajador.query.get_or_404(id_trab)
        
        if trabajador in requerimiento.trabajadores:
            requerimiento.trabajadores.remove(trabajador)
            db.session.commit()
            
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@controllers_bp.route('/agregar_miembro_equipo', methods=['POST'])
def agregar_miembro_equipo():
    try:
        data = request.get_json()
        trabajador_nuevo = False
        
        if data.get('es_nuevo'):
            # Validar que el RUT no exista ya
            rut_existente = Trabajador.query.filter_by(rut=data['rut']).first()
            if rut_existente:
                return jsonify({'success': False, 'error': 'Ya existe un trabajador con este RUT'})
            
            trabajador = Trabajador(
                nombre=data['nombre'],
                rut=data['rut'],
                profesion=data.get('profesion', ''),
                nombrecorto=data.get('nombre_corto', '')
            )
            db.session.add(trabajador)
            db.session.flush()
            trabajador_nuevo = True
        else:
            trabajador = Trabajador.query.get_or_404(data['id_trabajador'])
        
        requerimiento = Requerimiento.query.get_or_404(data['id_requerimiento'])
        
        # Usar una especialidad por defecto o crear una genérica
        # Buscar si existe una especialidad "General" o "Responsable"
        especialidad = Especialidad.query.filter_by(nombre='General').first()
        if not especialidad:
            # Si no existe, buscar cualquier especialidad o crear una por defecto
            especialidad = Especialidad.query.first()
            if not especialidad:
                # Crear especialidad por defecto si no hay ninguna
                especialidad = Especialidad(nombre='Responsable General')
                db.session.add(especialidad)
                db.session.flush()
        
        # Verificar si ya existe este trabajador en el equipo
        equipo_existente = EquipoTrabajo.query.filter_by(
            id_requerimiento=requerimiento.id,
            id_trabajador=trabajador.id
        ).first()
        
        if equipo_existente:
            return jsonify({'success': False, 'error': 'Este trabajador ya es miembro del equipo'})
        
        # Crear el equipo de trabajo
        equipo = EquipoTrabajo(
            id_requerimiento=requerimiento.id,
            id_trabajador=trabajador.id,
            id_especialidad=especialidad.id
        )
        
        db.session.add(equipo)
        db.session.commit()
        
        # **NUEVO: Recalcular progresos de actividades del proyecto**
        print(f"🔄 Miembro agregado al equipo, recalculando progresos del proyecto {requerimiento.id}")
        recalcular_progresos_proyecto(requerimiento.id)
        
        return jsonify({
            'success': True,
            'trabajador_nuevo': trabajador_nuevo,
            'miembro': {
                'id': equipo.id,
                'trabajador_id': trabajador.id,
                'trabajador_nombre': trabajador.nombre,
                'profesion': trabajador.profesion,
                'especialidad_nombre': especialidad.nombre
            }
        })
        
    except Exception as e:
        print("Error:", str(e))
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@controllers_bp.route('/quitar_miembro_equipo/<int:id_equipo>', methods=['POST'])
def quitar_miembro_equipo(id_equipo):
    try:
        equipo = EquipoTrabajo.query.get(id_equipo)
        
        if equipo is None:
            return jsonify({
                'success': True,
                'message': 'El miembro ya no existe en el equipo'
            })

        # Guardar el id del requerimiento antes de eliminar
        id_requerimiento = equipo.id_requerimiento
        
        # Eliminar el equipo
        db.session.delete(equipo)
        db.session.commit()
        
        # **NUEVO: Recalcular progresos de actividades del proyecto**
        print(f"🔄 Miembro eliminado del equipo, recalculando progresos del proyecto {id_requerimiento}")
        recalcular_progresos_proyecto(id_requerimiento)
        
        # Contar miembros restantes
        miembros_restantes = EquipoTrabajo.query.filter_by(id_requerimiento=id_requerimiento).count()
        
        return jsonify({
            'success': True,
            'message': 'Miembro eliminado correctamente',
            'miembros_restantes': miembros_restantes
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al quitar miembro: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Error al eliminar el miembro del equipo'
        }), 500

# NOTA: Esta funcionalidad se movió al blueprint de requerimientos en:
# app/controllers/requerimientos_controller.py
# 
# @controllers_bp.route('/update_requerimiento_completar/<int:id>', methods=['POST'])
# @login_required
# def update_requerimiento_completar(id):
#     try:
#         requerimiento = Requerimiento.query.get_or_404(id)
#         
#         # Validar y convertir campos requeridos, asignando None si están vacíos
#         id_tipologia = request.form.get('id_tipologia')
#         if id_tipologia and id_tipologia.strip():
#             requerimiento.id_tipologia = int(id_tipologia)
#         
#         id_financiamiento = request.form.get('id_financiamiento')  
#         if id_financiamiento and id_financiamiento.strip():
#             requerimiento.id_financiamiento = int(id_financiamiento)
#             
#         id_tipoproyecto = request.form.get('id_tipoproyecto')
#         if id_tipoproyecto and id_tipoproyecto.strip():
#             requerimiento.id_tipoproyecto = int(id_tipoproyecto)
#             
#         id_prioridad = request.form.get('id_prioridad')
#         if id_prioridad and id_prioridad.strip():
#             requerimiento.id_prioridad = int(id_prioridad)
#         
#         # Agregar el campo grupo
#         id_grupo = request.form.get('id_grupo')
#         if id_grupo and id_grupo.strip():
#             requerimiento.id_grupo = int(id_grupo)
#         
#         # Manejar observación de completado como historial
#         observacion_nueva = request.form.get('observacion')
#         if observacion_nueva and observacion_nueva.strip():
#             # Crear nueva entrada en el historial de observaciones
#             nueva_observacion = ObservacionRequerimiento(
#                 id_requerimiento=id,
#                 observacion=observacion_nueva.strip(),
#                 id_usuario=current_user.id,
#                 pagina_origen='requerimiento_completar',
#                 tipo_evento='completado'
#             )
#             db.session.add(nueva_observacion)
# 
#         # Contar miembros del equipo de trabajo usando .count() para relaciones lazy
#         equipos_count = requerimiento.equipos_trabajo.count()
# 
#         # Validar completitud: todos los campos requeridos (observación ya no es obligatoria aquí)
#         # Todos los campos deben estar completos para cambiar de estado (100%)
#         campos_llenos = all([
#             requerimiento.id_tipologia,
#             requerimiento.id_financiamiento,
#             requerimiento.id_tipoproyecto,
#             requerimiento.id_prioridad,
#             requerimiento.id_grupo,
#             equipos_count > 0
#         ])
# 
#         # Procesar las relaciones trabajador-especialidad
#         equipos_trabajo = requerimiento.equipos_trabajo.all()
#         for equipo in equipos_trabajo:
#             exists = db.session.query(requerimiento_trabajador_especialidad).filter_by(
#                 requerimiento_id=id,
#                 trabajador_id=equipo.id_trabajador,
#                 especialidad_id=equipo.id_especialidad
#             ).first()
#             if not exists:
#                 stmt = requerimiento_trabajador_especialidad.insert().values(
#                     requerimiento_id=id,
#                     trabajador_id=equipo.id_trabajador,
#                     especialidad_id=equipo.id_especialidad
#                 )
#                 db.session.execute(stmt)
# 
#         # Si está completo, cambiar estado a 3. Si no, solo guardar datos.
#         if campos_llenos:
#             requerimiento.id_estado = 3
#             db.session.commit()
#             flash('Requerimiento completado y enviado a ejecución exitosamente', 'success')
#         else:
#             db.session.commit()
#             flash('Información guardada. Complete todos los campos y agregue al menos un miembro para avanzar a ejecución.', 'warning')
# 
#         return redirect(url_for('controllers.ruta_requerimientos_completar'))
# 
#     except Exception as e:
#         db.session.rollback()
#         print(f"Error en update_requerimiento_completar: {str(e)}")
#         flash(f'Error al actualizar requerimiento: {str(e)}', 'error')
#         return redirect(url_for('controllers.ruta_requerimientos_completar'))


@controllers_bp.route('/obtener_observaciones_requerimiento/<int:id>', methods=['GET'])
@login_required
def obtener_observaciones_requerimiento(id):
    """Obtener todas las observaciones de un requerimiento específico"""
    try:
        observaciones = ObservacionRequerimiento.query.filter_by(id_requerimiento=id)\
            .order_by(ObservacionRequerimiento.fecha_registro.desc())\
            .all()
        
        observaciones_list = []
        for obs in observaciones:
            observaciones_list.append({
                'id': obs.id,
                'observacion': obs.observacion,
                'fecha_registro': obs.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                'usuario_nombre': obs.usuario.nombre if obs.usuario else 'Usuario eliminado',
                'pagina_origen': obs.pagina_origen,
                'tipo_evento': obs.tipo_evento
            })
        
        return jsonify({
            'success': True,
            'observaciones': observaciones_list
        })
    
    except Exception as e:
        print(f"Error al obtener observaciones: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        })
    

# ==================================================================================
# Rutas CRUD para Proyectos-aceptar
# NOTA: Esta funcionalidad se movió al blueprint de requerimientos en:
# app/controllers/requerimientos_controller.py
# 
# @controllers_bp.route('/proyectos_aceptar', endpoint='ruta_proyectos_aceptar')
# @login_required
# def proyectos_aceptar():
    from app.models import CustomRole, AdministradorRecinto
    
    # Verificar el tipo de usuario y permisos
    is_superadmin = False
    is_administrador = False
    
    # Verificar si es SUPERADMIN del sistema
    if hasattr(current_user, 'rol') and current_user.rol:
        rol_name = current_user.rol.name if hasattr(current_user.rol, 'name') else str(current_user.rol)
        if rol_name and rol_name == 'SUPERADMIN':
            is_superadmin = True
    
    # Verificar si tiene rol personalizado ADMINISTRADOR
    if hasattr(current_user, 'custom_role_id') and current_user.custom_role_id:
        custom_role = CustomRole.query.get(current_user.custom_role_id)
        if custom_role and custom_role.name.upper() == 'ADMINISTRADOR':
            is_administrador = True
    
    # Filtrar requerimientos según permisos (solo estado "En Desarrollo - Ejecución" id_estado = 3)
    if is_superadmin:
        # SUPERADMIN ve todos los requerimientos
        requerimientos = Requerimiento.query.filter_by(id_estado=3).options(
            db.joinedload(Requerimiento.recinto),
            db.joinedload(Requerimiento.sector),
            db.joinedload(Requerimiento.estado),
            db.joinedload(Requerimiento.prioridad)
        ).order_by(Requerimiento.fecha_aceptacion.desc()).all()
    elif is_administrador:
        # ADMINISTRADOR solo ve requerimientos de sus recintos asignados
        recintos_asignados = AdministradorRecinto.obtener_recintos_administrador(current_user.id)
        
        if recintos_asignados:
            # Obtener IDs de recintos asignados
            recinto_ids = [asignacion.recinto_id for asignacion in recintos_asignados]
            
            # Filtrar requerimientos por recintos asignados y estado
            requerimientos = Requerimiento.query.filter_by(id_estado=3).options(
                db.joinedload(Requerimiento.recinto),
                db.joinedload(Requerimiento.sector),
                db.joinedload(Requerimiento.estado),
                db.joinedload(Requerimiento.prioridad)
            ).filter(Requerimiento.id_recinto.in_(recinto_ids)).order_by(Requerimiento.fecha_aceptacion.desc()).all()
        else:
            # Si no tiene recintos asignados, no ve ningún requerimiento
            requerimientos = []
    else:
        # Otros usuarios solo ven requerimientos de su recinto
        user_recinto_id = current_user.recinto_id
        if user_recinto_id:
            requerimientos = Requerimiento.query.filter_by(id_estado=3).options(
                db.joinedload(Requerimiento.recinto),
                db.joinedload(Requerimiento.sector),
                db.joinedload(Requerimiento.estado),
                db.joinedload(Requerimiento.prioridad)
            ).filter(Requerimiento.id_recinto == user_recinto_id).order_by(Requerimiento.fecha_aceptacion.desc()).all()
        else:
            # Si el usuario no tiene recinto asignado, no ve ningún requerimiento
            requerimientos = []
    
    # Obtener sectores disponibles según permisos (igual que en requerimientos)
    if is_superadmin:
        sectores = Sector.query.all()
    elif is_administrador:
        # ADMINISTRADOR solo ve sectores de sus recintos asignados
        recintos_asignados = AdministradorRecinto.obtener_recintos_administrador(current_user.id)
        
        if recintos_asignados:
            recinto_ids = [asignacion.recinto_id for asignacion in recintos_asignados]
            recintos = Recinto.query.filter(Recinto.id.in_(recinto_ids)).all()
            
            # Obtener sectores únicos de los recintos asignados
            sectores_ids = set()
            for recinto in recintos:
                if recinto.tiporecinto and recinto.tiporecinto.id_sector:
                    sectores_ids.add(recinto.tiporecinto.id_sector)
            
            if sectores_ids:
                sectores = Sector.query.filter(Sector.id.in_(sectores_ids)).all()
            else:
                sectores = []
        else:
            sectores = []
    else:
        # Otros usuarios solo ven sectores relacionados con su recinto
        if current_user.recinto_id and current_user.recinto:
            user_sector_id = current_user.recinto.tiporecinto.id_sector if current_user.recinto.tiporecinto else None
            if user_sector_id:
                sectores = Sector.query.filter_by(id=user_sector_id).all()
            else:
                sectores = []
        else:
            sectores = []
    
    # NOTA: Esta función está obsoleta - usar proyectos_controller.py
    return redirect(url_for('proyectos.ruta_proyectos_aceptar'))

# NOTA: Esta funcionalidad se movió al blueprint de requerimientos en:
# app/controllers/requerimientos_controller.py
# 
# @controllers_bp.route('/update_proyecto_aceptar/<int:id>', methods=['POST'])
# @login_required
# def update_proyecto_aceptar(id):
    try:
        requerimiento = Requerimiento.query.get_or_404(id)
        observacion_texto = request.form.get('observacion', '').strip()

        if not observacion_texto:
            flash('La observación es requerida', 'error')
            return redirect(url_for('controllers.ruta_proyectos_aceptar'))

        # Actualizar estado
        requerimiento.id_estado = 4  # Estado: Finalizado
        
        # Solo actualizar observación si hay texto (mantener compatibilidad con sistema anterior)
        if observacion_texto:
            requerimiento.observacion = observacion_texto
        
        # Guardar observación en historial
        if observacion_texto:
            nueva_observacion = ObservacionRequerimiento(
                id_requerimiento=id,
                observacion=observacion_texto,
                fecha_registro=datetime.now(),
                id_usuario=current_user.id if current_user.is_authenticated else None,
                pagina_origen='proyectos_aceptar',
                tipo_evento='Finalizado'
            )
            db.session.add(nueva_observacion)
        
        db.session.commit()
        flash('Proyecto finalizado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al finalizar proyecto: {e}', 'error')
    return redirect(url_for('controllers.ruta_proyectos_aceptar'))

# NOTA: Esta funcionalidad se movió al blueprint de requerimientos en:
# app/controllers/requerimientos_controller.py
# 
# @controllers_bp.route('/update_proyecto_rechazar/<int:id>', methods=['POST'])
# @login_required
# def update_proyecto_rechazar(id):
    try:
        requerimiento = Requerimiento.query.get_or_404(id)
        observacion_texto = request.form.get('observacion', '').strip()

        if not observacion_texto:
            flash('La observación es requerida', 'error')
            return redirect(url_for('controllers.ruta_proyectos_aceptar'))

        # Actualizar estado
        requerimiento.id_estado = 9  # Estado: Rechazado
        
        # Solo actualizar observación si hay texto (mantener compatibilidad con sistema anterior)
        if observacion_texto:
            requerimiento.observacion = observacion_texto
        
        # Guardar observación en historial
        if observacion_texto:
            nueva_observacion = ObservacionRequerimiento(
                id_requerimiento=id,
                observacion=observacion_texto,
                fecha_registro=datetime.now(),
                id_usuario=current_user.id if current_user.is_authenticated else None,
                pagina_origen='proyectos_aceptar',
                tipo_evento='Rechazado'
            )
            db.session.add(nueva_observacion)
        
        db.session.commit()
        flash('Proyecto rechazado y actualizado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al rechazar proyecto: {e}', 'error')
    return redirect(url_for('controllers.ruta_proyectos_aceptar'))

# ==================================================================================
# Rutas CRUD para Proyectos-completar
@controllers_bp.route('/proyectos_completar', endpoint='ruta_proyectos_completar')
def proyectos_completar():
    """Página para visualizar proyectos finalizados y sus cartas Gantt"""
    try:
        # Obtener solo los requerimientos en estado "Finalizado" (id_estado = 4)
        requerimientos = Requerimiento.query.filter_by(id_estado=4).\
            order_by(Requerimiento.fecha.desc()).all()
            
        sectores = Sector.query.all()
        
        return render_template('proyecto-completar.html',
                             requerimientos=requerimientos,
                             sectores=sectores)
    except Exception as e:
        flash(f'Error al cargar proyectos completados: {str(e)}', 'error')
        return redirect(url_for('index'))




@controllers_bp.route('/ver_gantt/<int:req_id>', methods=['GET'])
def ver_gantt(req_id):
    """
    Descarga o muestra el archivo XLSX de Carta Gantt asociado al requerimiento.
    Debes guardar los archivos con un nombre único, por ejemplo: gantt_<req_id>.xlsx
    """
    import os
    from flask import current_app, abort

    upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
    filename = f"gantt_{req_id}.xlsx"
    filepath = os.path.join(upload_folder, filename)

    if not os.path.isfile(filepath):
        flash('No se ha subido una Carta Gantt para este proyecto.', 'error')
        return redirect(request.referrer or url_for('controllers.ruta_proyectos_completar'))

    # Puedes cambiar as_attachment a False si quieres que se muestre in
    return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@controllers_bp.route('/subir_gantt_xlsx/<int:req_id>', methods=['POST'])
def subir_gantt_xlsx(req_id):
    """
    Sube y procesa un archivo XLSX de Gantt para un requerimiento específico.
    Guarda las actividades, procesa los recursos y crea registros de avance_actividad.
    """
    file_path = None
    try:
        # Verificar que el requerimiento existe
        requerimiento = Requerimiento.query.get_or_404(req_id)
        
        # Verificar que se subió un archivo
        if 'archivo_gantt' not in request.files:
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('controllers.ruta_proyectos_completar'))
        
        file = request.files['archivo_gantt']
        
        if file.filename == '':
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('controllers.ruta_proyectos_completar'))
        
        if not file.filename.lower().endswith('.xlsx'):
            flash('Por favor, seleccione un archivo XLSX válido.', 'error')
            return redirect(url_for('controllers.ruta_proyectos_completar'))
        
        # Crear directorio para archivos Gantt si no existe
        gantt_dir = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'gantt')
        os.makedirs(gantt_dir, exist_ok=True)
        
        # Generar nombre único para el archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"gantt_req_{req_id}_{timestamp}_{secure_filename(file.filename)}"
        file_path = os.path.join(gantt_dir, filename)
        
        # Guardar archivo temporalmente
        file.save(file_path)
        print(f"📁 Archivo guardado temporalmente en: {file_path}")
        
        # Leer el archivo para guardarlo en la base de datos
        with open(file_path, 'rb') as f:
            archivo_bytes = f.read()
        
        # Limpiar cualquier transacción pendiente
        try:
            db.session.rollback()
        except:
            pass
        
        # Verificar si ya existe un archivo para este requerimiento
        gantt_existente = GanttArchivo.query.filter_by(id_requerimiento=req_id).first()
        if gantt_existente:
            # Actualizar archivo existente
            gantt_existente.archivo = archivo_bytes
            gantt_existente.nombre_archivo = file.filename
            gantt_existente.tipo_archivo = 'xlsx'
            gantt_existente.tamano_archivo = len(archivo_bytes)
            gantt_existente.fecha_subida = datetime.now()
            print(f"📝 Actualizando archivo existente en GanttArchivo")
        else:
            # Crear nuevo registro
            gantt_archivo = GanttArchivo(
                id_requerimiento=req_id,
                nombre_archivo=file.filename,
                tipo_archivo='xlsx',
                archivo=archivo_bytes,
                tamano_archivo=len(archivo_bytes),
                fecha_subida=datetime.now()
            )
            db.session.add(gantt_archivo)
            print(f"📝 Creando nuevo registro en GanttArchivo")
        
        # Guardar primero el archivo en GanttArchivo
        try:
            db.session.commit()
            print(f"✅ Archivo guardado en GanttArchivo")
        except Exception as e:
            print(f"❌ Error al guardar archivo en GanttArchivo: {str(e)}")
            db.session.rollback()
            flash(f"❌ Error al guardar el archivo: {str(e)}", 'error')
            return redirect(url_for('controllers.ruta_proyectos_completar'))
        
        # Procesar archivo y guardar actividades con recursos
        resultado = procesar_gantt_con_recursos(file_path, req_id)
        
        if resultado['success']:
            print(f"✅ Archivo procesado exitosamente")
            
            mensaje = (f"✅ Archivo procesado exitosamente: "
                       f"{resultado['actividades_procesadas']} actividades, "
                       f"{resultado['recursos_procesados']} asignaciones de recursos y "
                       f"{resultado['avances_creados']} registros de avance creados.")
            if resultado['errores']:
                mensaje += f" Se encontraron {len(resultado['errores'])} filas con errores."
                # Agregar detalles de los errores más comunes
                print(f"📋 Detalles de errores encontrados:")
                for i, error in enumerate(resultado['errores'][:5]):  # Mostrar máximo 5 errores
                    print(f"   {i+1}. {error}")
                if len(resultado['errores']) > 5:
                    print(f"   ... y {len(resultado['errores']) - 5} errores más.")
            flash(mensaje, 'success')
        else:
            error_detalle = resultado.get('error', 'Error desconocido')
            print(f"❌ Error procesando archivo: {error_detalle}")
            flash(f"❌ Error al procesar las actividades: {error_detalle}.", 'error')

    except Exception as e:
        import traceback
        print(f"❌ Error fatal en subir_gantt_xlsx: {traceback.format_exc()}")
        flash(f'Ocurrió un error inesperado al procesar el archivo: {str(e)}', 'error')
        db.session.rollback()
    
    finally:
        # Eliminar archivo temporal después de procesarlo
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"🗑️ Archivo temporal eliminado: {file_path}")
            except Exception as e:
                print(f"⚠️ No se pudo eliminar el archivo temporal: {str(e)}")
    
    return redirect(url_for('controllers.ruta_proyectos_completar'))

@controllers_bp.route('/gantt_data/<int:req_id>', methods=['GET'])
def gantt_data(req_id):
    """
    Obtiene los datos de actividades del Gantt para un requerimiento específico
    desde la base de datos o desde el archivo XLSX almacenado.
    """
    try:
        print(f"📡 Solicitando datos de Gantt para requerimiento ID: {req_id}")
        
        # Verificar que el requerimiento existe
        requerimiento = Requerimiento.query.get_or_404(req_id)
        
        # Primero intentar obtener desde la base de datos (actividades procesadas)
        actividades_bd = ActividadProyecto.query.filter_by(
            requerimiento_id=req_id, 
            activo=True
        ).order_by(ActividadProyecto.edt).all()
        
        if actividades_bd:
            print(f"📊 Obteniendo {len(actividades_bd)} actividades desde la base de datos")
            
            # Convertir actividades de BD a formato compatible con el frontend
            actividades = []
            for act in actividades_bd:
                actividad_dict = {
                    'ID': act.id,
                    'EDT': act.edt,
                    'E.D.T.': act.edt,
                    'Edt': act.edt,
                    'WBS': act.edt,
                    'Nivel de esquema': act.nivel_esquema,
                    'Nombre de tarea': act.nombre_tarea,
                    'Actividad': act.nombre_tarea,
                    'Nombre': act.nombre_tarea,
                    'Task Name': act.nombre_tarea,
                    'Comienzo': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else '',
                    'Inicio': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else '',
                    'Start': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else '',
                    'Fecha Inicio': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else '',
                    'Fin': act.fecha_fin.strftime('%Y-%m-%d') if act.fecha_fin else '',
                    'End': act.fecha_fin.strftime('%Y-%m-%d') if act.fecha_fin else '',
                    'Fecha Fin': act.fecha_fin.strftime('%Y-%m-%d') if act.fecha_fin else '',
                    'Duración': act.duracion,
                    'Duration': act.duracion,
                    'Días': act.duracion,
                    'Recursos': act.recursos or '',
                    'Resource Names': act.recursos or '',
                    'Nombres de los recursos': act.recursos or '',
                    'Progreso': float(act.progreso) if act.progreso else 0.0,
                    'Progress': float(act.progreso) if act.progreso else 0.0,
                    '% Completado': float(act.progreso) if act.progreso else 0.0,
                    'Porcentaje completado': float(act.progreso) if act.progreso else 0.0
                }
                actividades.append(actividad_dict)
            
            return jsonify({
                'success': True,
                'actividades': actividades,
                'info': {
                    'total_actividades': len(actividades),
                    'fuente': 'base_de_datos',
                    'fecha_actualizacion': max([act.updated_at for act in actividades_bd]).strftime('%Y-%m-%d %H:%M:%S') if actividades_bd else None,
                    'proyecto_id': req_id,
                    'proyecto_nombre': requerimiento.nombre
                }
            })
        
        # Si no hay actividades en BD, intentar obtener desde archivo Gantt almacenado
        gantt_archivo = GanttArchivo.query.filter_by(id_requerimiento=req_id).first()
        
        if not gantt_archivo:
            print(f"❌ No se encontraron actividades ni archivo Gantt para requerimiento {req_id}")
            return jsonify({
                'success': False,
                'error': 'No hay actividades registradas para este proyecto. Sube un archivo XLSX primero.'
            }), 404
        
        # Leer actividades desde archivo Gantt almacenado
        print(f"📄 Leyendo actividades desde archivo XLSX almacenado: {gantt_archivo.nombre_archivo}")
        
        try:
            import pandas as pd
            from io import BytesIO
            
            # Leer el archivo desde la base de datos
            df = pd.read_excel(BytesIO(gantt_archivo.archivo), engine='openpyxl')
            
            if df.empty:
                return jsonify({
                    'success': False,
                    'error': 'El archivo XLSX está vacío'
                }), 400
            
            print(f"📋 Archivo leído: {len(df)} filas, columnas: {list(df.columns)}")
            
            # Convertir DataFrame a lista de diccionarios
            actividades = []
            for index, row in df.iterrows():
                actividad_dict = {}
                for col in df.columns:
                    value = row[col]
                    if pd.notna(value):
                        # Formatear fechas
                        if 'fecha' in col.lower() or 'inicio' in col.lower() or 'fin' in col.lower() or 'start' in col.lower() or 'end' in col.lower():
                            try:
                                if isinstance(value, str):
                                    fecha_obj = pd.to_datetime(value, errors='coerce')
                                else:
                                    fecha_obj = pd.to_datetime(value)
                                
                                if not pd.isna(fecha_obj):
                                    actividad_dict[col] = fecha_obj.strftime('%Y-%m-%d')
                                else:
                                    actividad_dict[col] = str(value)
                            except:
                                actividad_dict[col] = str(value)
                        else:
                            actividad_dict[col] = value
                    else:
                        actividad_dict[col] = ''
                
                # Solo agregar filas que tengan al menos un EDT/WBS y nombre
                edt_cols = ['EDT', 'WBS', 'E.D.T.', 'Edt']
                nombre_cols = ['Nombre de tarea', 'Actividad', 'Nombre', 'Task Name']
                
                tiene_edt = any(col in actividad_dict and actividad_dict[col] for col in edt_cols)
                tiene_nombre = any(col in actividad_dict and actividad_dict[col] for col in nombre_cols)
                
                if tiene_edt and tiene_nombre:
                    actividades.append(actividad_dict)
            
            print(f"✅ {len(actividades)} actividades válidas extraídas del archivo")
            
            return jsonify({
                'success': True,
                'actividades': actividades,
                'info': {
                    'total_actividades': len(actividades),
                    'fuente': 'archivo_xlsx',
                    'archivo': gantt_archivo.nombre_archivo,
                    'fecha_subida': gantt_archivo.fecha_subida.strftime('%Y-%m-%d %H:%M:%S'),
                    'proyecto_id': req_id,
                    'proyecto_nombre': requerimiento.nombre
                }
            })
            
        except Exception as archivo_error:
            print(f"❌ Error al leer archivo XLSX: {str(archivo_error)}")
            return jsonify({
                'success': False,
                'error': f'Error al procesar archivo XLSX: {str(archivo_error)}'
            }), 500
            
    except Exception as e:
        print(f"❌ Error general en gantt_data: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500

def convertir_fecha_segura(fecha_raw, nombre_campo="fecha"):
    """
    Convierte un valor raw de fecha a un objeto date válido.
    Retorna None si no se puede convertir.
    """
    import pandas as pd
    from datetime import datetime, date
    
    if pd.isna(fecha_raw) or fecha_raw is None:
        return None
    
    try:
        # Si ya es un objeto date, retornarlo
        if isinstance(fecha_raw, date):
            return fecha_raw
        
        # Si es datetime, extraer la fecha
        if isinstance(fecha_raw, datetime):
            return fecha_raw.date()
        
        # Si es string, intentar convertir
        if isinstance(fecha_raw, str):
            fecha_raw = fecha_raw.strip()
            if not fecha_raw or fecha_raw.lower() in ['nat', 'none', 'null', '']:
                return None
            
            # Intentar varios formatos de fecha
            formatos_fecha = [
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%m/%d/%Y',
                '%d-%m-%Y',
                '%Y/%m/%d',
                '%d.%m.%Y',
                '%Y.%m.%d'
            ]
            
            for formato in formatos_fecha:
                try:
                    return datetime.strptime(fecha_raw, formato).date()
                except ValueError:
                    continue
        
        # Intentar con pandas
        fecha_parsed = pd.to_datetime(fecha_raw, errors='coerce')
        if pd.isna(fecha_parsed):
            return None
        
        return fecha_parsed.date()
        
    except Exception as e:
        print(f"⚠️ Error convirtiendo {nombre_campo}: {fecha_raw} - {e}")
        return None

def procesar_gantt_con_recursos(file_path, req_id):
    """
    Procesa el archivo Gantt y crea registros en actividades_gantt y avance_actividad
    basándose en los recursos asignados (nombrecorto de trabajadores).
    """
    try:
        import pandas as pd
        from datetime import datetime
        import re
        
        # Leer el archivo Excel
        df = pd.read_excel(file_path, engine='openpyxl')
        
        if df.empty:
            return {'success': False, 'error': 'El archivo está vacío'}
        
        print(f"📊 Procesando {len(df)} filas del archivo Gantt")
        print(f"📋 Columnas disponibles en el archivo: {list(df.columns)}")
        
        # Mostrar muestra de datos para debug
        print(f"📄 Primeras 2 filas del archivo para referencia:")
        for i in range(min(2, len(df))):
            print(f"   Fila {i+1}: {dict(df.iloc[i])}")
            
        # Verificar si hay datos vacíos
        if df.empty or len(df) == 0:
            return {'success': False, 'error': 'El archivo está vacío o no contiene datos'}
            
        # Verificar que no todas las filas estén vacías
        filas_con_datos = df.dropna(how='all')
        if len(filas_con_datos) == 0:
            return {'success': False, 'error': 'Todas las filas del archivo están vacías'}
        
        # Limpiar registros anteriores para este requerimiento
        from app.models import ActividadProyecto, AvanceActividad
        
        # Eliminar actividades anteriores
        ActividadProyecto.query.filter_by(requerimiento_id=req_id).delete()
        
        # Eliminar avances anteriores
        AvanceActividad.query.filter_by(requerimiento_id=req_id).delete()
        
        actividades_procesadas = 0
        recursos_procesados = 0
        avances_creados = 0
        errores = []
        
        # Mapeo de columnas más flexible
        columnas_mapeo = {
            'ID': ['ID', 'id', 'Id', 'Número', 'No', '#', 'Índice'],
            'Nivel de esquema': ['Nivel de esquema', 'Nivel', 'Level', 'Nivel esquema', 'Outline Level'],
            'EDT': ['EDT', 'WBS', 'E.D.T.', 'Edt', 'Work Breakdown Structure'],
            'Nombre de tarea': ['Nombre de tarea', 'Actividad', 'Tarea', 'Task Name', 'Nombre', 'Task', 'Name'],
            'Duración': ['Duración', 'Duration', 'Dias', 'Días', 'Days', 'Duracion'],
            'Comienzo': ['Comienzo', 'Inicio', 'Start', 'Fecha Inicio', 'Start Date', 'Fecha de inicio'],
            'Fin': ['Fin', 'Final', 'End', 'Fecha Fin', 'Finish', 'End Date', 'Fecha de fin'],
            'Recursos': ['Nombres de los recursos', 'Recursos', 'Resources', 'Resource Names', 'Assigned Resources'],
            'Progreso': ['Progreso', 'Progress', '% Completado', 'Porcentaje completado', 'Complete', 'Percent Complete', '% Complete']
        }
        
        # Buscar columnas existentes con coincidencia más flexible
        columnas_encontradas = {}
        columnas_df_lower = {col.lower().strip(): col for col in df.columns}
        
        for clave, variantes in columnas_mapeo.items():
            encontrada = False
            for variante in variantes:
                # Buscar coincidencia exacta primero
                if variante.lower().strip() in columnas_df_lower:
                    columnas_encontradas[clave] = columnas_df_lower[variante.lower().strip()]
                    encontrada = True
                    break
                # Buscar coincidencia parcial
                for col_df in columnas_df_lower.keys():
                    if variante.lower().strip() in col_df or col_df in variante.lower().strip():
                        columnas_encontradas[clave] = columnas_df_lower[col_df]
                        encontrada = True
                        break
                if encontrada:
                    break
        
        print(f"📋 Columnas encontradas después del mapeo: {columnas_encontradas}")
        
        # Verificar que se encontraron las columnas mínimas requeridas
        columnas_requeridas = ['EDT', 'Nombre de tarea', 'Comienzo', 'Fin']
        columnas_faltantes = [col for col in columnas_requeridas if col not in columnas_encontradas]
        
        if columnas_faltantes:
            error_msg = f"No se encontraron las columnas requeridas: {', '.join(columnas_faltantes)}"
            print(f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # Mostrar las primeras filas para debug
        print(f"📄 Iniciando procesamiento de filas:")
        
        for index, row in df.iterrows():
            try:
                print(f"\n🔍 Procesando fila {index + 1}: {dict(row)}")
                
                # Extraer datos básicos de la actividad con manejo de errores mejorado
                edt = None
                nombre_tarea = None
                recursos_texto = None
                
                # Intentar obtener EDT
                if 'EDT' in columnas_encontradas:
                    edt_raw = row.get(columnas_encontradas['EDT'])
                    if pd.notna(edt_raw):
                        edt = str(edt_raw).strip()
                
                # Intentar obtener nombre de tarea
                if 'Nombre de tarea' in columnas_encontradas:
                    nombre_raw = row.get(columnas_encontradas['Nombre de tarea'])
                    if pd.notna(nombre_raw):
                        nombre_tarea = str(nombre_raw).strip()
                
                # Intentar obtener recursos
                if 'Recursos' in columnas_encontradas:
                    recursos_raw = row.get(columnas_encontradas['Recursos'])
                    if pd.notna(recursos_raw):
                        recursos_texto = str(recursos_raw).strip()
                
                print(f"   EDT: '{edt}', Nombre: '{nombre_tarea}', Recursos: '{recursos_texto}'")
                
                # Validar datos mínimos requeridos
                if not edt or edt == 'nan' or not nombre_tarea or nombre_tarea == 'nan':
                    error_msg = f"Fila {index + 1}: EDT o nombre de tarea vacío (EDT: '{edt}', Nombre: '{nombre_tarea}')"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue
                
                # Validar longitudes de campos
                if len(str(edt)) > 50:
                    error_msg = f"Fila {index + 1}: EDT muy largo (máximo 50 caracteres): '{edt}'"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue
                    
                if len(str(nombre_tarea)) > 500:
                    error_msg = f"Fila {index + 1}: Nombre de tarea muy largo (máximo 500 caracteres): '{nombre_tarea[:100]}...'"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue
                
                # Procesar fechas con manejo de errores mejorado
                fecha_inicio = None
                fecha_fin = None
                
                try:
                    if 'Comienzo' in columnas_encontradas:
                        inicio_raw = row.get(columnas_encontradas['Comienzo'])
                        print(f"   Fecha inicio raw: {inicio_raw} (tipo: {type(inicio_raw)})")
                        fecha_inicio = convertir_fecha_segura(inicio_raw, "fecha_inicio")
                        
                    if 'Fin' in columnas_encontradas:
                        fin_raw = row.get(columnas_encontradas['Fin'])
                        print(f"   Fecha fin raw: {fin_raw} (tipo: {type(fin_raw)})")
                        fecha_fin = convertir_fecha_segura(fin_raw, "fecha_fin")
                                
                except Exception as fecha_error:
                    print(f"⚠️ Error procesando fechas en fila {index + 1}: {fecha_error}")
                    fecha_inicio = None
                    fecha_fin = None
                
                print(f"   Fechas procesadas - Inicio: {fecha_inicio}, Fin: {fecha_fin}")
                
                # Validar que las fechas no sean None (requerido por el modelo)
                if not fecha_inicio or not fecha_fin:
                    error_msg = f"Fila {index + 1}: Fechas requeridas son nulas o inválidas (Inicio: {fecha_inicio}, Fin: {fecha_fin})"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue
                
                # Procesar duración
                duracion = 0
                try:
                    if 'Duración' in columnas_encontradas:
                        duracion_raw = row.get(columnas_encontradas['Duración'])
                        if pd.notna(duracion_raw):
                            # Extraer solo números de la duración
                            duracion_str = str(duracion_raw).replace('days', '').replace('días', '').replace('d', '').strip()
                            duracion = int(float(duracion_str)) if duracion_str.replace('.', '').isdigit() else 0
                except Exception as duracion_error:
                    print(f"⚠️ Error procesando duración en fila {index + 1}: {duracion_error}")
                    duracion = 0
                
                # Procesar progreso
                progreso = 0.0
                try:
                    if 'Progreso' in columnas_encontradas:
                        progreso_raw = row.get(columnas_encontradas['Progreso'])
                        if pd.notna(progreso_raw):
                            progreso_val = float(str(progreso_raw).replace('%', ''))
                            # Si el valor es mayor a 1, asumimos que está en porcentaje
                            progreso = progreso_val if progreso_val <= 1 else progreso_val / 100
                            progreso = max(0.0, min(1.0, progreso))  # Limitar entre 0 y 1
                except Exception as progreso_error:
                    print(f"⚠️ Error procesando progreso en fila {index + 1}: {progreso_error}")
                    progreso = 0.0
                
                # Procesar nivel de esquema
                nivel_esquema = 1
                try:
                    if 'Nivel de esquema' in columnas_encontradas:
                        nivel_raw = row.get(columnas_encontradas['Nivel de esquema'])
                        if pd.notna(nivel_raw):
                            nivel_esquema = int(float(str(nivel_raw)))
                except Exception as nivel_error:
                    print(f"⚠️ Error procesando nivel en fila {index + 1}: {nivel_error}")
                    nivel_esquema = 1
                
                print(f"   Datos finales - Duración: {duracion}, Progreso: {progreso}, Nivel: {nivel_esquema}")
                
                print(f"   Datos finales - Duración: {duracion}, Progreso: {progreso}, Nivel: {nivel_esquema}")
                
                # Validar que el progreso esté en el rango válido para Numeric(5,2)
                progreso_porcentaje = progreso * 100
                if progreso_porcentaje > 999.99 or progreso_porcentaje < 0:
                    error_msg = f"Fila {index + 1}: Progreso fuera de rango (0-999.99): {progreso_porcentaje}"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue
                
                # Crear actividad en actividad_proyecto
                try:
                    # Validación final de fechas antes de insertar
                    if fecha_inicio is None or fecha_fin is None:
                        error_msg = f"Fila {index + 1}: Fechas nulas no permitidas para inserción (Inicio: {fecha_inicio}, Fin: {fecha_fin})"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        continue
                    
                    # Validación adicional para asegurar que son objetos date válidos
                    try:
                        # Intentar formatear las fechas para asegurar que son válidas
                        fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
                        fecha_fin_str = fecha_fin.strftime('%Y-%m-%d')
                        print(f"   ✅ Fechas validadas - Inicio: {fecha_inicio_str}, Fin: {fecha_fin_str}")
                    except (AttributeError, ValueError) as fecha_val_error:
                        error_msg = f"Fila {index + 1}: Fechas no son objetos date válidos - {fecha_val_error}"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        continue
                    
                    actividad = ActividadProyecto(
                        requerimiento_id=req_id,
                        edt=edt,
                        nivel_esquema=nivel_esquema,
                        nombre_tarea=nombre_tarea,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        duracion=duracion,
                        recursos=recursos_texto or '',
                        progreso=progreso * 100,  # Guardar como porcentaje (0-100)
                        activo=True
                        # Los campos created_at y updated_at se manejan automáticamente por TimestampMixin
                    )
                    
                    db.session.add(actividad)
                    db.session.flush()  # Para obtener el ID
                    actividades_procesadas += 1
                    print(f"✅ Actividad creada con ID: {actividad.id} - EDT: {edt} - Nombre: {nombre_tarea}")
                    
                    # Procesar recursos y crear registros de avance_actividad
                    if recursos_texto and recursos_texto.lower() != 'nan':
                        trabajadores_encontrados = procesar_recursos_trabajadores(recursos_texto, req_id, actividad.id)
                        recursos_procesados += len(trabajadores_encontrados)
                        avances_creados += len(trabajadores_encontrados)
                        print(f"   👥 Recursos procesados: {len(trabajadores_encontrados)}")
                    
                except Exception as actividad_error:
                    error_msg = f"Error creando actividad en fila {index + 1}: {str(actividad_error)}"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    print(f"   Datos de la fila: EDT='{edt}', Nombre='{nombre_tarea}', Inicio={fecha_inicio}, Fin={fecha_fin}, Duración={duracion}")
                    import traceback
                    print(f"   Stack trace: {traceback.format_exc()}")
                    continue
                
            except Exception as e:
                error_msg = f"Error general en fila {index + 1}: {str(e)}"
                errores.append(error_msg)
                print(f"❌ {error_msg}")
                continue
        
        # Guardar todos los cambios
        try:
            print(f"💾 Intentando guardar {actividades_procesadas} actividades en la base de datos...")
            db.session.commit()
            print(f"✅ Cambios guardados en la base de datos exitosamente")
            
            # **NUEVO: Recalcular progresos después de procesar el Gantt**
            if actividades_procesadas > 0:
                print(f"🔄 Recalculando progresos de todas las actividades del proyecto {req_id}")
                recalcular_progresos_proyecto(req_id)
                
        except Exception as commit_error:
            db.session.rollback()
            print(f"❌ Error al guardar en la base de datos: {commit_error}")
            import traceback
            print(f"📋 Traceback del error de commit: {traceback.format_exc()}")
            return {'success': False, 'error': f'Error al guardar en la base de datos: {commit_error}'}
        
        resultado = {
            'success': True,
            'actividades_procesadas': actividades_procesadas,
            'recursos_procesados': recursos_procesados,
            'avances_creados': avances_creados,
            'errores': errores
        }
        
        print(f"📊 Resultado final del procesamiento:")
        print(f"   ✅ Actividades procesadas: {actividades_procesadas}")
        print(f"   👥 Recursos procesados: {recursos_procesados}")
        print(f"   📈 Avances creados: {avances_creados}")
        print(f"   ❌ Errores encontrados: {len(errores)}")
        print(f"   📋 Total de filas procesadas: {len(df)}")
        
        return resultado
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error general en procesar_gantt_con_recursos: {str(e)}")
        import traceback
        print(f"📋 Traceback completo: {traceback.format_exc()}")
        return {'success': False, 'error': str(e)}

def procesar_recursos_trabajadores(recursos_texto, req_id, actividad_id):
    """
    Procesa el texto de recursos y crea registros en avance_actividad
    para cada trabajador encontrado o creado automáticamente.
    
    Formatos esperados:
    - "PM1[100%]"
    - "PM1[100%];ARQ1[50%]"
    - "PM1, ARQ1" (sin porcentaje, se asigna 100% por defecto)
    """
    try:
        import re
        from app.models import Trabajador, AvanceActividad
        
        trabajadores_encontrados = []
        
        print(f"🔍 Procesando recursos: '{recursos_texto}'")
        
        # Limpiar el texto - mantener punto y coma como separador principal
        recursos_texto = recursos_texto.replace('\n', ';').strip()
        
        # Dividir por punto y coma primero para manejar múltiples asignaciones
        recursos_individuales = [recurso.strip() for recurso in recursos_texto.split(';') if recurso.strip()]
        
        print(f"📋 Recursos individuales encontrados: {recursos_individuales}")
        
        for recurso_individual in recursos_individuales:
            try:
                # Patrón para extraer código y porcentaje: "CODIGO[PORCENTAJE%]"
                patron_con_porcentaje = r'^([A-Za-z0-9]+)\[(\d+)%\]$'
                match_porcentaje = re.match(patron_con_porcentaje, recurso_individual.strip())
                
                if match_porcentaje:
                    # Formato: "PM1[100%]"
                    codigo = match_porcentaje.group(1).upper().strip()
                    porcentaje = int(match_porcentaje.group(2))
                    
                    print(f"   📌 Procesando: {codigo} con {porcentaje}% de asignación")
                    
                    # Buscar o crear trabajador
                    trabajador = buscar_o_crear_trabajador_por_codigo(codigo)
                    
                    if trabajador:
                        # Crear registro de avance en avance_actividad
                        exito = crear_registro_avance(req_id, trabajador.id, actividad_id, porcentaje)
                        if exito:
                            trabajadores_encontrados.append({
                                'trabajador_id': trabajador.id,
                                'codigo': codigo,
                                'porcentaje': porcentaje,
                                'nombre': trabajador.nombre,
                                'creado_automaticamente': trabajador.password == "123456"
                            })
                            print(f"   ✅ Asignación creada: {codigo} -> {porcentaje}%")
                        else:
                            print(f"   ❌ Error creando asignación para: {codigo}")
                    else:
                        print(f"   ❌ No se pudo obtener/crear trabajador para: {codigo}")
                        
                else:
                    # Formato sin porcentaje: "PM1" o "PM1, ARQ2" (separado por comas)
                    # Dividir por comas si no tiene formato de porcentaje
                    codigos_sin_porcentaje = [codigo.strip() for codigo in recurso_individual.split(',') if codigo.strip()]
                    
                    for codigo_raw in codigos_sin_porcentaje:
                        # Limpiar código (solo alfanuméricos)
                        codigo = re.sub(r'[^A-Za-z0-9]', '', codigo_raw.upper())
                        
                        if codigo:
                            print(f"   📌 Procesando: {codigo} sin porcentaje (asignando 100% por defecto)")
                            
                            # Buscar o crear trabajador
                            trabajador = buscar_o_crear_trabajador_por_codigo(codigo)
                            
                            if trabajador:
                                # Crear registro de avance con 100% por defecto
                                exito = crear_registro_avance(req_id, trabajador.id, actividad_id, 100)
                                if exito:
                                    trabajadores_encontrados.append({
                                        'trabajador_id': trabajador.id,
                                        'codigo': codigo,
                                        'porcentaje': 100,
                                        'nombre': trabajador.nombre,
                                        'creado_automaticamente': trabajador.password == "123456"
                                    })
                                    print(f"   ✅ Asignación creada: {codigo} -> 100% (por defecto)")
                                else:
                                    print(f"   ❌ Error creando asignación para: {codigo}")
                            else:
                                print(f"   ❌ No se pudo obtener/crear trabajador para: {codigo}")
                                
            except Exception as recurso_error:
                print(f"   ❌ Error procesando recurso individual '{recurso_individual}': {str(recurso_error)}")
                continue
        
        if trabajadores_encontrados:
            print(f"👥 ✅ Total procesado: {len(trabajadores_encontrados)} trabajadores en recursos: '{recursos_texto}'")
            # Mostrar resumen
            for trabajador in trabajadores_encontrados:
                estado = "🆕 CREADO" if trabajador.get('creado_automaticamente', False) else "📍 EXISTENTE"
                print(f"     {estado}: {trabajador['codigo']} ({trabajador['nombre']}) - {trabajador['porcentaje']}%")
        else:
            print(f"⚠️ No se procesaron trabajadores para: '{recursos_texto}'")
        
        return trabajadores_encontrados
        
    except Exception as e:
        print(f"❌ Error general procesando recursos '{recursos_texto}': {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        return []

def buscar_o_crear_trabajador_por_codigo(codigo):
    """
    Busca un trabajador por su nombrecorto (código).
    Si no existe, lo crea automáticamente con password "123456".
    """
    try:
        from app.models import Trabajador
        
        # Buscar exacto primero
        trabajador = Trabajador.query.filter_by(nombrecorto=codigo).first()
        
        if not trabajador:
            # Buscar case-insensitive
            trabajador = Trabajador.query.filter(
                Trabajador.nombrecorto.ilike(codigo)
            ).first()
        
        if trabajador:
            print(f"✅ Trabajador encontrado: {codigo} -> {trabajador.nombre}")
            return trabajador
        else:
            # Crear trabajador automáticamente
            print(f"🆕 Creando trabajador automáticamente para código: {codigo}")
            nuevo_trabajador = Trabajador(
                nombre=f"Trabajador {codigo}",  # Nombre temporal basado en el código
                nombrecorto=codigo,
                email=f"{codigo.lower()}@empresa.com",  # Email temporal
                password="123456",  # Password por defecto según requerimiento
                profesion="Por definir",  # Profesión temporal
                activo=True
            )
            
            db.session.add(nuevo_trabajador)
            db.session.flush()  # Para obtener el ID sin hacer commit completo
            
            print(f"✅ Trabajador creado automáticamente: {codigo} -> ID: {nuevo_trabajador.id}")
            return nuevo_trabajador
            
    except Exception as e:
        print(f"❌ Error buscando/creando trabajador {codigo}: {str(e)}")
        return None

def buscar_trabajador_por_codigo(codigo):
    """
    Busca un trabajador por su nombrecorto (código).
    """
    try:
        # Buscar exacto primero
        trabajador = Trabajador.query.filter_by(nombrecorto=codigo).first()
        
        if not trabajador:
            # Buscar case-insensitive
            trabajador = Trabajador.query.filter(
                Trabajador.nombrecorto.ilike(codigo)
            ).first()
        
        if trabajador:
            print(f"✅ Trabajador encontrado: {codigo} -> {trabajador.nombre}")
        else:
            print(f"⚠️ Trabajador no encontrado para código: {codigo}")
            
        return trabajador
        
    except Exception as e:
        print(f"❌ Error buscando trabajador {codigo}: {str(e)}")
        return None

def crear_registro_avance(req_id, trabajador_id, actividad_id, porcentaje_asignacion):
    """
    Crea un registro inicial en avance_actividad con el porcentaje de asignación.
    Retorna True si se creó/actualizó exitosamente, False en caso contrario.
    """
    try:
        from app.models import AvanceActividad
        from datetime import datetime
        
        # Verificar si ya existe un registro
        avance_existente = AvanceActividad.query.filter_by(
            requerimiento_id=req_id,
            trabajador_id=trabajador_id,
            actividad_id=actividad_id
        ).first()
        
        if not avance_existente:
            # Crear nuevo registro
            avance = AvanceActividad(
                requerimiento_id=req_id,
                trabajador_id=trabajador_id,
                actividad_id=actividad_id,
                porcentaje_asignacion=porcentaje_asignacion,
                progreso_actual=0.0,  # Inicialmente 0%
                fecha_registro=datetime.now().date(),
                fecha_creacion=datetime.now(),
                observaciones=f"Asignación automática desde Gantt ({porcentaje_asignacion}% asignado)"
            )
            
            db.session.add(avance)
            print(f"   📝 ✅ Registro de avance CREADO: Trabajador {trabajador_id}, Actividad {actividad_id}, Asignación {porcentaje_asignacion}%")
            return True
            
        else:
            # Actualizar el existente
            avance_existente.porcentaje_asignacion = porcentaje_asignacion
            avance_existente.fecha_actualizacion = datetime.now()
            avance_existente.observaciones = f"Asignación actualizada desde Gantt ({porcentaje_asignacion}% asignado)"
            
            print(f"   📝 🔄 Registro de avance ACTUALIZADO: Trabajador {trabajador_id}, Actividad {actividad_id}, Asignación {porcentaje_asignacion}%")
            return True
            
    except Exception as e:
        print(f"   ❌ Error creando/actualizando registro de avance: {str(e)}")
        import traceback
        print(f"   📋 Traceback: {traceback.format_exc()}")
        return False
    
    """
    Obtiene todas las actividades del proyecto desde la base de datos
    """
    try:
        # Verificar que el requerimiento existe
        requerimiento = Requerimiento.query.get_or_404(req_id)
        
        # Obtener actividades de la base de datos
        actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=req_id, 
            activo=True
        ).order_by(ActividadProyecto.edt).all()
        
        if not actividades:
            return jsonify({
                'success': False,
                'error': 'No hay actividades guardadas en la base de datos para este proyecto.'
            }), 404
        
        # Convertir a formato JSON
        actividades_data = [act.to_dict() for act in actividades]
        
        return jsonify({
            'success': True,
            'actividades': actividades_data,
            'info': {
                'total_actividades': len(actividades_data),
                'proyecto_id': req_id,
                'proyecto_nombre': requerimiento.nombre,
                'fuente': 'base_de_datos'
            }
        })
        
    except Exception as e:
        print(f"Error al obtener actividades del proyecto: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al obtener actividades: {str(e)}'
        }), 500


@controllers_bp.route('/actualizar_actividad/<int:actividad_id>', methods=['PUT'])
def actualizar_actividad(actividad_id):
    """
    Actualiza una actividad específica en la base de datos
    """
    try:
        actividad = ActividadProyecto.query.get_or_404(actividad_id)
        data = request.get_json()
        
        # Actualizar campos permitidos
        if 'progreso' in data:
            progreso = float(data['progreso'])
            actividad.progreso = max(0.0, min(100.0, progreso))
        
        if 'nombre_tarea' in data:
            actividad.nombre_tarea = str(data['nombre_tarea']).strip()
        
        if 'recursos' in data:
            actividad.recursos = str(data['recursos']).strip()
        
        if 'observaciones' in data:
            # Actualizar datos adicionales con observaciones
            if not actividad.datos_adicionales:
                actividad.datos_adicionales = {}
            actividad.datos_adicionales['observaciones'] = str(data['observaciones']).strip()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Actividad actualizada exitosamente',
            'actividad': actividad.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al actualizar actividad: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al actualizar actividad: {str(e)}'
        }), 500


# ==================================================================================
# Rutas para avance de actividades
@controllers_bp.route('/avance-actividades')
@login_required
def avance_actividades():
    """Página para registrar avance de actividades - Auto-selección del usuario logueado"""
    try:
        # Verificar que el usuario esté autenticado y sea un trabajador válido
        if not current_user.is_authenticated:
            flash('Debe iniciar sesión para acceder a esta página', 'error')
            return redirect(url_for('auth.login'))
        
        # El current_user ya es una instancia de Trabajador gracias al user_loader
        trabajador_actual = current_user
        
        # Verificar que el trabajador tenga proyectos asignados
        tiene_proyectos = db.session.query(AvanceActividad).filter_by(
            trabajador_id=trabajador_actual.id
        ).first() is not None
        
        if not tiene_proyectos:
            # Verificar también en equipo_trabajo como fallback
            tiene_proyectos = db.session.query(EquipoTrabajo).filter_by(
                id_trabajador=trabajador_actual.id
            ).first() is not None
        
        # Fecha actual para el formulario
        from datetime import date
        fecha_actual = date.today().isoformat()
        
        return render_template('avance-actividades.html', 
                             trabajador_actual=trabajador_actual,
                             tiene_proyectos=tiene_proyectos,
                             fecha_actual=fecha_actual)
                             
    except Exception as e:
        print(f"Error al cargar página de avance de actividades: {e}")
        flash(f'Error al cargar la página: {str(e)}', 'error')
        return redirect(url_for('auth.login'))

@controllers_bp.route('/avance-actividades-all')
@login_required
def avance_actividades_all():
    """Página para registrar avance de actividades con trabajadores - Todos los proyectos"""
    try:
        # Obtener todos los trabajadores de la base de datos
        trabajadores = Trabajador.query.order_by(Trabajador.nombre).all()
        
        # Fecha actual para el formulario
        from datetime import date
        fecha_actual = date.today().isoformat()
        
        return render_template('avance-actividades-all.html', 
                             trabajadores=trabajadores,
                             fecha_actual=fecha_actual)
    except Exception as e:
        print(f"Error al cargar trabajadores: {e}")
        flash(f'Error al cargar la página: {str(e)}', 'error')
        return render_template('avance-actividades-all.html', 
                             trabajadores=[],
                             fecha_actual=date.today().isoformat())

@controllers_bp.route('/proyectos_por_trabajador/<int:trabajador_id>')
def proyectos_por_trabajador(trabajador_id):
    """Obtener proyectos asignados a un trabajador específico según tabla avance_actividad"""
    try:
        # Verificar que el trabajador existe
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        
        # Obtener proyectos donde el trabajador tiene asignaciones en avance_actividad
        # trabajador.id → avance_actividad.trabajador_id → avance_actividad.requerimiento_id → requerimiento
        proyectos_query = db.session.query(Requerimiento).join(
            AvanceActividad, Requerimiento.id == AvanceActividad.requerimiento_id
        ).filter(
            AvanceActividad.trabajador_id == trabajador_id
        ).distinct()
        
        proyectos = proyectos_query.all()
        
        # Si no hay proyectos en avance_actividad, buscar también en equipo_trabajo como respaldo
        if not proyectos:
            proyectos_query_equipo = db.session.query(Requerimiento).join(
                EquipoTrabajo, Requerimiento.id == EquipoTrabajo.id_requerimiento
            ).filter(
                EquipoTrabajo.id_trabajador == trabajador_id,
                Requerimiento.id_estado == 4  # Solo proyectos en Estado 4 (seguimiento activo)
            ).distinct()
            
            proyectos = proyectos_query_equipo.all()
        
        # Convertir a formato JSON
        proyectos_data = []
        for proyecto in proyectos:
            # Contar cuántas actividades tiene asignadas este trabajador en este proyecto
            actividades_asignadas = db.session.query(AvanceActividad).filter_by(
                trabajador_id=trabajador_id,
                requerimiento_id=proyecto.id
            ).count()
            
            proyectos_data.append({
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'sector': proyecto.sector.nombre if proyecto.sector else 'Sin sector',
                'estado': proyecto.estado.nombre if proyecto.estado else 'Sin estado',
                'descripcion': proyecto.descripcion,
                'fecha': proyecto.fecha.strftime('%d-%m-%Y') if proyecto.fecha else '',
                'actividades_asignadas': actividades_asignadas
            })
        
        return jsonify({
            'success': True,
            'proyectos': proyectos_data,
            'trabajador': {
                'id': trabajador.id,
                'nombre': trabajador.nombre,
                'nombrecorto': trabajador.nombrecorto or '',
                'email': trabajador.email or '',
                'cargo': trabajador.profesion or ''
            }
        })
        
    except Exception as e:
        print(f"Error al obtener proyectos del trabajador {trabajador_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al cargar proyectos: {str(e)}'
        }), 500

@controllers_bp.route('/proyectos_por_trabajador_all/<int:trabajador_id>')
def proyectos_por_trabajador_all(trabajador_id):
    """Obtener proyectos donde el trabajador tiene tareas asignadas en la carta Gantt"""
    try:
        # Verificar que el trabajador existe
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        
        print(f"\n🔍 Buscando proyectos para trabajador ID {trabajador_id}:")
        print(f"   - Nombre: {trabajador.nombre}")
        print(f"   - Nombre corto: {trabajador.nombrecorto}")
        
        # Obtener TODOS los proyectos activos (sin filtrar por estado)
        # Cambio: antes solo buscaba en desarrollo (estados 2,3)
        # Ahora busca en todos los estados para mostrar todos los proyectos donde tiene tareas
        proyectos_query = db.session.query(Requerimiento).order_by(Requerimiento.nombre)
        
        proyectos = proyectos_query.all()
        print(f"   📊 Total proyectos en el sistema: {len(proyectos)}")
        
        # Mostrar IDs de proyectos
        if proyectos:
            ids_proyectos = [p.id for p in proyectos]
            print(f"   📌 IDs de proyectos: {ids_proyectos}")
        
        # Verificar qué actividades existen y cómo se ven sus recursos
        total_actividades = db.session.query(ActividadProyecto).filter(
            ActividadProyecto.activo == True
        ).count()
        print(f"   📋 Total actividades activas en el sistema: {total_actividades}")
        
        if total_actividades == 0:
            print(f"   ⚠️ NO HAY ACTIVIDADES DE PROYECTO EN EL SISTEMA")
            print(f"   💡 Necesitas llenar la carta Gantt en /proyecto-llenar")
            return jsonify({
                'success': True,
                'proyectos': [],
                'trabajador': {
                    'id': trabajador.id,
                    'nombre': trabajador.nombre,
                    'nombrecorto': trabajador.nombrecorto or '',
                    'email': trabajador.email or '',
                    'cargo': trabajador.profesion or ''
                },
                'mensaje': 'No hay actividades en el sistema. Llena la carta Gantt desde /proyecto-llenar'
            })
        
        # Mostrar muestra de recursos
        muestra_actividades = db.session.query(
            ActividadProyecto.id, 
            ActividadProyecto.nombre_tarea, 
            ActividadProyecto.recursos,
            ActividadProyecto.requerimiento_id
        ).filter(
            ActividadProyecto.activo == True,
            ActividadProyecto.recursos.isnot(None),
            ActividadProyecto.recursos != ''
        ).limit(5).all()
        
        print(f"   🔎 Muestra de actividades con recursos:")
        for act in muestra_actividades:
            print(f"      - ID {act.id} (Req {act.requerimiento_id}): '{act.nombre_tarea}' → Recursos: '{act.recursos}'")
        
        # Para cada proyecto, verificar si el trabajador tiene actividades asignadas
        proyectos_con_actividades = []
        
        # Buscar por nombrecorto (ARQ01, ARQ02, etc.)
        nombre_buscar = trabajador.nombrecorto if trabajador.nombrecorto else trabajador.nombre
        print(f"   🔍 Buscando por nombre: '{nombre_buscar}'")
        
        for proyecto in proyectos:
            # Contar actividades donde el trabajador está en recursos
            actividades_asignadas = db.session.query(ActividadProyecto).filter(
                ActividadProyecto.requerimiento_id == proyecto.id,
                ActividadProyecto.activo == True,
                ActividadProyecto.recursos.like(f'%{nombre_buscar}%')
            ).count()
            
            if actividades_asignadas > 0:
                proyectos_con_actividades.append(proyecto)
                print(f"   ✅ Proyecto '{proyecto.nombre}': {actividades_asignadas} actividades con {nombre_buscar}")
        
        print(f"   📈 Total proyectos con actividades del trabajador: {len(proyectos_con_actividades)}")
        
        # Convertir a formato JSON
        proyectos_data = []
        for proyecto in proyectos_con_actividades:
            # Contar actividades para este proyecto y trabajador
            actividades_count = db.session.query(ActividadProyecto).filter(
                ActividadProyecto.requerimiento_id == proyecto.id,
                ActividadProyecto.activo == True,
                ActividadProyecto.recursos.like(f'%{nombre_buscar}%')
            ).count()
            
            proyectos_data.append({
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'sector': proyecto.sector.nombre if proyecto.sector else 'Sin sector',
                'estado': proyecto.estado.nombre if proyecto.estado else 'Sin estado',
                'descripcion': proyecto.descripcion,
                'fecha': proyecto.fecha.strftime('%d-%m-%Y') if proyecto.fecha else '',
                'actividades_asignadas': actividades_count
            })
        
        return jsonify({
            'success': True,
            'proyectos': proyectos_data,
            'trabajador': {
                'id': trabajador.id,
                'nombre': trabajador.nombre,
                'nombrecorto': trabajador.nombrecorto or '',
                'email': trabajador.email or '',
                'cargo': trabajador.profesion or ''
            }
        })
        
    except Exception as e:
        print(f"Error al obtener todos los proyectos para trabajador {trabajador_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al cargar proyectos: {str(e)}'
        }), 500

@controllers_bp.route('/actividades_trabajador_proyecto/<int:trabajador_id>/<int:proyecto_id>')
def actividades_trabajador_proyecto(trabajador_id, proyecto_id):
    """Obtener actividades de un proyecto específico para un trabajador"""
    try:
        # Verificar que el trabajador y proyecto existen
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        proyecto = Requerimiento.query.get_or_404(proyecto_id)
        
        # Verificar que el trabajador está asignado al proyecto usando avance_actividad
        asignacion_proyecto = AvanceActividad.query.filter_by(
            trabajador_id=trabajador_id,
            requerimiento_id=proyecto_id
        ).first()
        
        # Si no se encuentra en avance_actividad, verificar en EquipoTrabajo como fallback
        if not asignacion_proyecto:
            equipo = EquipoTrabajo.query.filter_by(
                id_trabajador=trabajador_id,
                id_requerimiento=proyecto_id
            ).first()
            
            if not equipo:
                return jsonify({
                    'success': False,
                    'error': 'El trabajador no está asignado a este proyecto'
                }), 403
        
        # Primero intentar obtener actividades desde la tabla actividad_proyecto
        actividades_bd = ActividadProyecto.query.filter_by(
            requerimiento_id=proyecto_id,
            activo=True
        ).order_by(ActividadProyecto.edt).all()
        
        if actividades_bd:
            # Convertir actividades de BD a formato compatible
            # Solo mostrar actividades donde el trabajador esté asignado (tiene registro en avance_actividad)
            actividades = []
            for act in actividades_bd:
                # Verificar si el trabajador está asignado a esta actividad específica
                asignacion = AvanceActividad.query.filter_by(
                    trabajador_id=trabajador_id,
                    actividad_id=act.id,
                    requerimiento_id=proyecto_id
                ).first()
                
                # Solo mostrar actividades donde el trabajador esté específicamente asignado
                if asignacion:
                    actividad_dict = {
                        'ID': act.id,
                        'EDT': act.edt,
                        'Nivel de esquema': act.nivel_esquema,
                        'Nombre de tarea': act.nombre_tarea,
                        'Comienzo': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else '',
                        'Inicio': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else '',
                        'Fin': act.fecha_fin.strftime('%Y-%m-%d') if act.fecha_fin else '',
                        'Duración': act.duracion,
                        'Recursos': act.recursos or '',
                        'Progreso': float(asignacion.progreso_actual) if asignacion else 0.0,  # CORRECCIÓN: Progreso personal del trabajador
                        '% Completado': float(asignacion.progreso_actual) if asignacion else 0.0,  # CORRECCIÓN: Progreso personal del trabajador
                        'Asignacion': asignacion.porcentaje_asignacion if asignacion else 100,
                        'Progreso_Trabajador': asignacion.progreso_actual if asignacion else 0.0,
                        'Progreso_Total_Actividad': float(act.progreso) if act.progreso else 0.0  # NUEVO: Para referencia del progreso total
                    }
                    actividades.append(actividad_dict)
            
            return jsonify({
                'success': True,
                'actividades': actividades,
                'info': {
                    'total_actividades': len(actividades),
                    'fuente': 'base_de_datos',
                    'trabajador': trabajador.nombre,
                    'trabajador_codigo': trabajador.nombrecorto,
                    'proyecto': proyecto.nombre
                }
            })
        
        # Si no hay actividades en BD, intentar obtener desde archivo Gantt
        gantt = GanttArchivo.query.filter_by(id_requerimiento=proyecto_id).first()
        
        if not gantt:
            return jsonify({
                'success': False,
                'error': 'No hay actividades registradas ni archivo Gantt para este proyecto'
            }), 404
        
        # Leer actividades desde archivo Gantt y filtrar por trabajador asignado
        try:
            import pandas as pd
            from io import BytesIO
            
            df = pd.read_excel(BytesIO(gantt.archivo), engine='openpyxl')
            
            # Obtener todas las actividades guardadas en BD para verificar asignaciones
            actividades_proyecto = ActividadProyecto.query.filter_by(proyecto_id=proyecto_id).all()
            actividades_asignadas = set()
            
            # Verificar qué actividades tiene asignadas este trabajador en la BD
            for actividad_bd in actividades_proyecto:
                avances = AvanceActividad.query.filter_by(
                    actividad_proyecto_id=actividad_bd.id,
                    trabajador_id=trabajador_id
                ).all()
                if avances:
                    actividades_asignadas.add(actividad_bd.id)
            
            # Si no hay asignaciones en BD, buscar en el archivo Gantt por nombrecorto
            if not actividades_asignadas:
                actividades_filtradas = []
                import re
                
                for _, row in df.iterrows():
                    recursos = str(row.get('Nombres de los recursos', '') or row.get('Recursos', ''))
                    
                    # Buscar el nombrecorto del trabajador en los recursos usando regex
                    patron_trabajador = r'\b' + re.escape(trabajador.nombrecorto) + r'\b'
                    if re.search(patron_trabajador, recursos, re.IGNORECASE):
                        # **CORRECCIÓN**: Para actividades del Gantt sin registro en BD, progreso inicial es 0
                        actividad = {
                            'EDT': row.get('EDT', ''),
                            'Nombre de tarea': row.get('Nombre de tarea', ''),
                            'Comienzo': row.get('Comienzo', ''),
                            'Fin': row.get('Fin', ''),
                            'Duración': row.get('Duración', 0),
                            'Recursos': recursos,
                            'Progreso': 0.0,  # CORRECCIÓN: Progreso personal inicial es 0% para trabajador
                            '% Completado': 0.0,  # CORRECCIÓN: Progreso personal inicial es 0%
                            'Progreso_Total_Actividad': row.get('Progreso', 0) or row.get('% Completado', 0) or 0,  # Progreso total del Gantt
                            'asignado_desde': 'archivo_gantt'
                        }
                        actividades_filtradas.append(actividad)
                
                return jsonify({
                    'success': True,
                    'actividades': actividades_filtradas,
                    'info': {
                        'total_actividades': len(actividades_filtradas),
                        'fuente': 'archivo_gantt',
                        'trabajador': f"{trabajador.nombrecorto} - {trabajador.nombre}",
                        'proyecto': proyecto.nombre,
                        'mensaje': 'Actividades encontradas en archivo Gantt'
                    }
                })
            
            # Si hay asignaciones en BD, mostrar solo las actividades asignadas del archivo
            actividades_filtradas = []
            for _, row in df.iterrows():
                edt = str(row.get('EDT', ''))
                # Buscar si esta actividad está en las asignadas
                for actividad_bd in actividades_proyecto:
                    if (actividad_bd.id in actividades_asignadas and 
                        (edt == str(actividad_bd.edt) or 
                         row.get('Nombre de tarea', '') == actividad_bd.nombre)):
                        
                        # **CORRECCIÓN**: Obtener el progreso personal del trabajador desde BD
                        avance_trabajador = AvanceActividad.query.filter_by(
                            trabajador_id=trabajador_id,
                            actividad_id=actividad_bd.id,
                            requerimiento_id=proyecto_id
                        ).first()
                        
                        progreso_personal = avance_trabajador.progreso_actual if avance_trabajador else 0.0
                        
                        actividad = {
                            'EDT': edt,
                            'Nombre de tarea': row.get('Nombre de tarea', ''),
                            'Comienzo': row.get('Comienzo', ''),
                            'Fin': row.get('Fin', ''),
                            'Duración': row.get('Duración', 0),
                            'Recursos': str(row.get('Nombres de los recursos', '') or row.get('Recursos', '')),
                            'Progreso': progreso_personal,  # CORRECCIÓN: Progreso personal del trabajador
                            '% Completado': progreso_personal,  # CORRECCIÓN: Progreso personal del trabajador
                            'Progreso_Total_Actividad': row.get('Progreso', 0) or row.get('% Completado', 0) or 0,  # Progreso total del Gantt
                            'asignado_desde': 'base_datos'
                        }
                        actividades_filtradas.append(actividad)
                        break
            
            return jsonify({
                'success': True,
                'actividades': actividades_filtradas,
                'info': {
                    'total_actividades': len(actividades_filtradas),
                    'fuente': 'archivo_gantt_con_asignaciones_bd',
                    'trabajador': f"{trabajador.nombrecorto} - {trabajador.nombre}",
                    'proyecto': proyecto.nombre,
                    'mensaje': 'Actividades asignadas en base de datos'
                }
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Error al procesar archivo Gantt: {str(e)}'
            }), 500
        
    except Exception as e:
        print(f"Error al obtener actividades: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}'
        }), 500

def recalcular_progresos_proyecto(requerimiento_id):
    """
    Recalcula el progreso de todas las actividades de un proyecto.
    Se ejecuta cuando se agregan o quitan trabajadores del equipo.
    
    Args:
        requerimiento_id: ID del proyecto/requerimiento
    """
    try:
        from app.models import ActividadProyecto
        
        print(f"🔄 Recalculando progresos de todas las actividades del proyecto {requerimiento_id}")
        
        # Obtener todas las actividades del proyecto
        actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=requerimiento_id,
            activo=True
        ).all()
        
        actividades_actualizadas = 0
        
        for actividad in actividades:
            try:
                # Calcular nuevo progreso usando la función existente
                progreso_anterior = actividad.progreso
                progreso_nuevo = calcular_progreso_actividad(actividad.id)
                
                # Solo actualizar si hay cambio
                if abs(progreso_anterior - progreso_nuevo) > 0.01:  # Tolerancia de 0.01%
                    actividad.progreso = progreso_nuevo
                    actividad.fecha_actualizacion = datetime.now()
                    actividades_actualizadas += 1
                    
                    print(f"   📊 Actividad {actividad.edt}: {progreso_anterior:.1f}% → {progreso_nuevo:.1f}%")
                
            except Exception as e:
                print(f"   ❌ Error recalculando actividad {actividad.edt}: {str(e)}")
                continue
        
        if actividades_actualizadas > 0:
            db.session.commit()
            print(f"✅ Recálculo completado: {actividades_actualizadas} actividades actualizadas")
        else:
            print(f"✅ Recálculo completado: Sin cambios necesarios")
            
    except Exception as e:
        print(f"❌ Error en recalcular_progresos_proyecto: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")

def calcular_progreso_actividad(actividad_id):
    """
    Calcula el progreso total de una actividad basado en el avance de todos los trabajadores asignados.
    
    Fórmula:
    progreso_total = (suma de horas_completadas_por_trabajador) / (suma de horas_totales_por_trabajador)
    
    Donde:
    - horas_totales_trabajador = (porcentaje_asignado * 8 horas * duracion) / 100
    - horas_completadas_trabajador = (horas_totales_trabajador * progreso_actual_trabajador) / 100
    
    Args:
        actividad_id: ID de la actividad en la tabla actividad_proyecto
        
    Returns:
        float: Progreso total de la actividad como porcentaje (0-100)
    """
    try:
        from app.models import ActividadProyecto, AvanceActividad
        
        # Obtener la actividad
        actividad = ActividadProyecto.query.get(actividad_id)
        if not actividad:
            print(f"❌ Actividad {actividad_id} no encontrada")
            return 0.0
        
        # Obtener todos los avances de trabajadores asignados a esta actividad
        avances = AvanceActividad.query.filter_by(actividad_id=actividad_id).all()
        
        if not avances:
            print(f"⚠️ No hay trabajadores asignados a la actividad {actividad_id}")
            # Verificar si es una actividad padre (sin asignaciones porque tiene sub-actividades)
            print(f"   📋 Actividad {actividad.edt} sin trabajadores - probablemente es padre de otras actividades")
            return 0.0  # Las actividades padre no llevan progreso individual
        
        total_horas_completadas = 0.0
        total_horas_asignadas = 0.0
        
        print(f"🧮 Calculando progreso para actividad ID {actividad_id} (EDT: {actividad.edt})")
        print(f"   📊 Duración: {actividad.duracion} días")
        print(f"   👥 Trabajadores asignados: {len(avances)}")
        
        for avance in avances:
            # Calcular horas totales para este trabajador
            # horas_totales = (porcentaje_asignado * 8 horas * duracion) / 100
            horas_por_dia = (avance.porcentaje_asignacion * 8) / 100
            horas_totales = horas_por_dia * actividad.duracion
            
            # Calcular horas completadas por este trabajador
            # horas_completadas = (horas_totales * progreso_actual) / 100
            horas_completadas = (horas_totales * avance.progreso_actual) / 100
            
            total_horas_asignadas += horas_totales
            total_horas_completadas += horas_completadas
            
            print(f"   👤 Trabajador {avance.trabajador_id}:")
            print(f"      📈 Asignación: {avance.porcentaje_asignacion}%")
            print(f"      ⏰ Horas/día: {horas_por_dia:.1f}")
            print(f"      📝 Horas totales: {horas_totales:.1f}")
            print(f"      ✅ Progreso personal: {avance.progreso_actual}%")
            print(f"      🎯 Horas completadas: {horas_completadas:.1f}")
        
        # Calcular progreso total de la actividad
        if total_horas_asignadas > 0:
            progreso_total = (total_horas_completadas / total_horas_asignadas) * 100
            progreso_total = max(0.0, min(100.0, progreso_total))  # Limitar entre 0 y 100
            
            print(f"   📊 RESULTADO:")
            print(f"      🎯 Total horas completadas: {total_horas_completadas:.1f}")
            print(f"      📝 Total horas asignadas: {total_horas_asignadas:.1f}")
            print(f"      ✅ Progreso total actividad: {progreso_total:.1f}%")
            
            return progreso_total
        else:
            print(f"   ⚠️ No hay horas asignadas para la actividad {actividad_id}")
            return 0.0
            
    except Exception as e:
        print(f"❌ Error calculando progreso de actividad {actividad_id}: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        return 0.0

def calcular_progreso_jerarquico(actividad_id):
    """
    Calcula el progreso de una actividad considerando la jerarquía EDT.
    
    - Si la actividad tiene hijas (es un nodo padre), calcula el promedio ponderado 
      por duración de sus hijas directas.
    - Si la actividad NO tiene hijas (es una hoja), calcula el progreso basado en 
      trabajadores asignados usando calcular_progreso_actividad().
    
    Estructura EDT: 1 → 1.1 → 1.1.1 → 1.1.4.1
    Ejemplo: EDT '1.1' tiene hijas directas '1.1.1', '1.1.2', '1.1.3', '1.1.4'
             (NO incluye '1.1.4.1' que es hija de '1.1.4')
    
    Args:
        actividad_id: ID de la actividad a calcular
        
    Returns:
        float: Progreso calculado (0-100)
    """
    try:
        from app.models import ActividadProyecto
        
        actividad = ActividadProyecto.query.get(actividad_id)
        if not actividad:
            print(f"⚠️ Actividad {actividad_id} no encontrada")
            return 0.0
        
        print(f"\n🔍 Calculando progreso jerárquico para actividad {actividad.edt} (ID: {actividad_id})")
        
        # Obtener hijas directas usando patrón EDT
        # Ejemplo: Si EDT es '1.1', buscar '1.1.%' pero excluir '1.1.%.%'
        edt_pattern = f"{actividad.edt}.%"
        edt_pattern_excluir = f"{actividad.edt}.%.%"
        
        hijas = ActividadProyecto.query.filter(
            ActividadProyecto.edt.like(edt_pattern),
            ~ActividadProyecto.edt.like(edt_pattern_excluir),
            ActividadProyecto.requerimiento_id == actividad.requerimiento_id
        ).all()
        
        if not hijas:
            # Es una hoja (sin hijas), calcular por trabajadores asignados
            print(f"   📄 Es hoja (sin hijas) - calculando por trabajadores")
            progreso = calcular_progreso_actividad(actividad_id)
            print(f"   ✅ Progreso calculado: {progreso:.2f}%")
            return progreso
        
        # Es un nodo padre, calcular promedio ponderado de hijas
        print(f"   🌳 Es nodo padre - calculando promedio ponderado de {len(hijas)} hijas")
        
        total_peso = sum(float(h.duracion or 0) for h in hijas)
        if total_peso == 0:
            print(f"   ⚠️ Peso total es 0, retornando 0%")
            return 0.0
        
        progreso_ponderado = 0.0
        for hija in hijas:
            peso_hija = float(hija.duracion or 0)
            progreso_hija = float(hija.progreso or 0)
            contribucion = progreso_hija * peso_hija / total_peso
            progreso_ponderado += contribucion
            print(f"      - {hija.edt}: {progreso_hija:.1f}% × {peso_hija}h = {contribucion:.2f}% contribución")
        
        print(f"   ✅ Progreso ponderado calculado: {progreso_ponderado:.2f}%")
        return progreso_ponderado
        
    except Exception as e:
        print(f"❌ Error calculando progreso jerárquico de actividad {actividad_id}: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        return 0.0

def recalcular_padres_recursivo(edt_hijo, requerimiento_id):
    """
    Recalcula recursivamente el progreso de todos los padres en la jerarquía EDT.
    Propaga los cambios desde una tarea hija hacia arriba hasta la raíz.
    
    Ejemplo: Si se actualiza '1.1.1', recalcula '1.1' y luego '1'
    
    Args:
        edt_hijo: EDT de la tarea que cambió (ej: '1.1.1')
        requerimiento_id: ID del requerimiento al que pertenece
        
    Returns:
        None
    """
    try:
        from app.models import ActividadProyecto
        
        print(f"\n🔄 Recalculando padres desde {edt_hijo}")
        
        # Obtener partes del EDT para navegar hacia arriba
        partes = edt_hijo.split('.')
        
        # Recorrer desde el hijo hacia arriba (1.1.1 → 1.1 → 1)
        while len(partes) > 1:
            # Eliminar último nivel para obtener EDT padre
            partes.pop()
            edt_padre = '.'.join(partes)
            
            # Buscar actividad padre
            padre = ActividadProyecto.query.filter_by(
                edt=edt_padre,
                requerimiento_id=requerimiento_id
            ).first()
            
            if not padre:
                print(f"   ⚠️ Padre {edt_padre} no encontrado")
                continue
            
            # Recalcular progreso del padre usando lógica jerárquica
            nuevo_progreso = calcular_progreso_jerarquico(padre.id)
            progreso_anterior = float(padre.progreso or 0)
            
            # Actualizar solo si cambió
            if abs(nuevo_progreso - progreso_anterior) > 0.01:  # Tolerancia 0.01%
                padre.progreso = nuevo_progreso
                db.session.commit()
                print(f"   ✅ Actualizado {edt_padre}: {progreso_anterior:.2f}% → {nuevo_progreso:.2f}%")
            else:
                print(f"   ℹ️ {edt_padre} sin cambios ({progreso_anterior:.2f}%)")
        
        print(f"✅ Recalculación completada")
        
    except Exception as e:
        print(f"❌ Error recalculando padres desde {edt_hijo}: {str(e)}")
        import traceback
        print(f"📋 Traceback: {traceback.format_exc()}")
        db.session.rollback()

@controllers_bp.route('/guardar_avances_trabajador', methods=['POST'])
def guardar_avances_trabajador():
    """Guardar los avances de actividades registrados por un trabajador"""
    try:
        from datetime import datetime
        import uuid
        
        data = request.get_json()
        print(f"🔍 DEBUG: Datos recibidos completos: {data}")
        
        trabajador_id = data.get('trabajador_id')
        proyecto_id = data.get('proyecto_id')
        fecha_registro = data.get('fecha_registro')
        avances = data.get('avances', [])
        comentarios_generales = data.get('comentarios', '')  # Comentarios generales del guardado
        
        # Generar ID único para esta sesión de guardado
        sesion_guardado = str(uuid.uuid4())[:8] + '_' + datetime.now().strftime('%Y%m%d_%H%M%S')
        
        print(f"🔍 DEBUG: trabajador_id={trabajador_id}, proyecto_id={proyecto_id}, fecha_registro={fecha_registro}")
        print(f"🔍 DEBUG: Número de avances recibidos: {len(avances)}")
        print(f"🔍 DEBUG: Sesión de guardado: {sesion_guardado}")
        print(f"🔍 DEBUG: Primer avance: {avances[0] if avances else 'Sin avances'}")
        
        if not all([trabajador_id, proyecto_id, fecha_registro, avances]):
            print(f"❌ ERROR: Datos incompletos")
            return jsonify({
                'success': False,
                'error': 'Datos incompletos'
            })
        
        print(f"✅ Validación inicial pasada")
        
        # Verificar que el trabajador y proyecto existen
        print(f"🔍 Buscando trabajador con ID: {trabajador_id}")
        trabajador = Trabajador.query.get(trabajador_id)
        print(f"🔍 Trabajador encontrado: {trabajador.nombre if trabajador else 'NO ENCONTRADO'}")
        
        print(f"🔍 Buscando proyecto con ID: {proyecto_id}")
        proyecto = Requerimiento.query.get(proyecto_id)
        print(f"🔍 Proyecto encontrado: {proyecto.nombre if proyecto else 'NO ENCONTRADO'}")
        
        if not trabajador or not proyecto:
            print(f"❌ ERROR: Trabajador o proyecto no encontrado")
            print(f"❌ Trabajador existe: {trabajador is not None}")
            print(f"❌ Proyecto existe: {proyecto is not None}")
            return jsonify({
                'success': False,
                'error': 'Trabajador o proyecto no encontrado'
            })
        
        print(f"✅ Trabajador y proyecto encontrados")
        
        # Verificar que el trabajador está asignado al proyecto
        # Primero buscar en AvanceActividad (método principal)
        print(f"🔍 Buscando asignación en AvanceActividad para trabajador {trabajador_id} y proyecto {proyecto_id}")
        asignacion_avance = AvanceActividad.query.filter_by(
            trabajador_id=trabajador_id,
            requerimiento_id=proyecto_id
        ).first()
        print(f"🔍 Asignación en AvanceActividad encontrada: {asignacion_avance is not None}")
        
        # Si no está en AvanceActividad, buscar en EquipoTrabajo como respaldo
        asignacion_equipo = None
        if not asignacion_avance:
            print(f"🔍 Buscando en EquipoTrabajo para trabajador {trabajador_id} y proyecto {proyecto_id}")
            asignacion_equipo = EquipoTrabajo.query.filter_by(
                id_trabajador=trabajador_id,
                id_requerimiento=proyecto_id,
                activo=True
            ).first()
            print(f"🔍 Asignación en EquipoTrabajo encontrada: {asignacion_equipo is not None}")
        
        # El trabajador debe estar asignado en al menos una de las dos tablas
        if not asignacion_avance and not asignacion_equipo:
            print(f"❌ ERROR: El trabajador no está autorizado para este proyecto")
            print(f"❌ No encontrado en AvanceActividad ni en EquipoTrabajo")
            return jsonify({
                'success': False,
                'error': 'El trabajador no está autorizado para este proyecto'
            })
        
        print(f"✅ Trabajador autorizado para el proyecto")
        print(f"✅ Método de autorización: {'AvanceActividad' if asignacion_avance else 'EquipoTrabajo'}")
        
        # Guardar cada avance
        avances_guardados = 0
        print(f"🔄 Iniciando procesamiento de {len(avances)} avances...")
        
        for i, avance in enumerate(avances):
            print(f"🔄 Procesando avance {i+1}/{len(avances)}")
            edt = avance.get('edt')
            progreso_anterior = avance.get('progreso_anterior', 0)
            progreso_nuevo = avance.get('progreso_nuevo', 0)
            
            print(f"📊 EDT: {edt}, Progreso: {progreso_anterior}% → {progreso_nuevo}%")
            
            # Solo procesar si hubo cambio en el progreso
            if progreso_anterior != progreso_nuevo:
                # Buscar la actividad en la base de datos
                actividad = ActividadProyecto.query.filter_by(
                    requerimiento_id=proyecto_id,
                    edt=edt
                ).first()
                
                if actividad:
                    # Buscar el registro de avance en la tabla avance_actividad
                    avance_actividad = AvanceActividad.query.filter_by(
                        requerimiento_id=proyecto_id,
                        trabajador_id=trabajador_id,
                        actividad_id=actividad.id
                    ).first()
                    
                    if avance_actividad:
                        # Mover el progreso actual a progreso anterior y guardar el nuevo progreso
                        progreso_real_anterior = avance_actividad.progreso_actual
                        avance_actividad.progreso_anterior = avance_actividad.progreso_actual
                        avance_actividad.progreso_actual = progreso_nuevo
                        avance_actividad.fecha_actualizacion = datetime.now()
                        avance_actividad.observaciones = f"Progreso actualizado: {progreso_real_anterior}% → {progreso_nuevo}%"
                        print(f"✅ Registro de avance actualizado para actividad {edt}: {progreso_real_anterior}% → {progreso_nuevo}%")
                    else:
                        # Crear nuevo registro si no existe
                        progreso_real_anterior = 0.0
                        avance_actividad = AvanceActividad(
                            requerimiento_id=proyecto_id,
                            trabajador_id=trabajador_id,
                            actividad_id=actividad.id,
                            porcentaje_asignacion=100.0,  # Por defecto 100%
                            progreso_anterior=0.0,
                            progreso_actual=progreso_nuevo,
                            fecha_registro=datetime.now().date(),
                            fecha_creacion=datetime.now(),
                            observaciones=f"Primer registro de progreso: {progreso_nuevo}%"
                        )
                        db.session.add(avance_actividad)
                        print(f"✅ Nuevo registro de avance creado para actividad {edt}: 0% → {progreso_nuevo}%")
                    
                    # **NUEVO: Guardar en historial solo si hubo cambio**
                    if progreso_real_anterior != progreso_nuevo:
                        historial_entry = HistorialAvanceActividad(
                            requerimiento_id=proyecto_id,
                            trabajador_id=trabajador_id,
                            actividad_id=actividad.id,
                            progreso_anterior=progreso_real_anterior,
                            progreso_nuevo=progreso_nuevo,
                            diferencia=progreso_nuevo - progreso_real_anterior,
                            comentarios=comentarios_generales,
                            fecha_cambio=datetime.now(),
                            sesion_guardado=sesion_guardado
                        )
                        db.session.add(historial_entry)
                        print(f"📝 Historial guardado para {edt}: {progreso_real_anterior}% → {progreso_nuevo}% (Δ {progreso_nuevo - progreso_real_anterior:+.1f}%)")
                    
                    # **Calcular y actualizar el progreso total de la actividad**
                    print(f"🔄 Recalculando progreso total para actividad {edt}...")
                    progreso_actividad_calculado = calcular_progreso_actividad(actividad.id)
                    
                    # Actualizar el progreso en la tabla de actividades
                    actividad.progreso = progreso_actividad_calculado
                    actividad.fecha_actualizacion = datetime.now()
                    
                    print(f"✅ Progreso de actividad {edt} actualizado: {progreso_actividad_calculado:.1f}%")
                    print(f"ℹ️  Esperando validación de supervisor para recalcular jerarquía")
                    
                    avances_guardados += 1
                    print(f"✅ Avance procesado exitosamente para EDT {edt}")
                else:
                    print(f"⚠️ Actividad con EDT {edt} no encontrada en el proyecto {proyecto_id}")
            else:
                print(f"ℹ️ Sin cambios para EDT {edt} - saltando")
        
        print(f"🔄 Iniciando commit a la base de datos...")
        db.session.commit()
        print(f"✅ Commit completado exitosamente. Total avances guardados: {avances_guardados}")
        
        response_data = {
            'success': True,
            'message': f'Se guardaron {avances_guardados} avances correctamente',
            'avances_guardados': avances_guardados
        }
        print(f"📤 Enviando respuesta: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"❌ ERROR en guardar_avances_trabajador: {str(e)}")
        print(f"❌ Tipo de error: {type(e).__name__}")
        import traceback
        print(f"❌ Traceback completo: {traceback.format_exc()}")
        db.session.rollback()
        print(f"🔄 Rollback ejecutado")
        return jsonify({
            'success': False,
            'error': f'Error al guardar: {str(e)}'
        }), 500

@controllers_bp.route('/exportar_avances_trabajador/<int:trabajador_id>/<int:proyecto_id>')
def exportar_avances_trabajador(trabajador_id, proyecto_id):
    """Exportar avances de actividades de un trabajador a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        proyecto = Requerimiento.query.get_or_404(proyecto_id)
        
        # Obtener actividades del proyecto
        actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=proyecto_id,
            activo=True
        ).order_by(ActividadProyecto.edt).all()
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Avances de Actividades"
        
        # Información del encabezado
        ws.merge_cells('A1:G1')
        header_cell = ws['A1']
        header_cell.value = f"Avances de Actividades - {proyecto.nombre}"
        header_cell.font = Font(size=16, bold=True)
        header_cell.alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:G2')
        info_cell = ws['A2']
        info_cell.value = f"Trabajador: {trabajador.nombre} - Fecha: {datetime.now().strftime('%d-%m-%Y')}"
        info_cell.font = Font(size=12)
        info_cell.alignment = Alignment(horizontal="center")
        
        # Encabezados de tabla
        encabezados = [
            'EDT', 'Actividad', 'Fecha Inicio', 'Fecha Fin', 
            'Duración', 'Progreso (%)', 'Recursos'
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=4, column=col, value=encabezado)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos de actividades
        for row, actividad in enumerate(actividades, 5):
            ws.cell(row=row, column=1, value=actividad.edt)
            ws.cell(row=row, column=2, value=actividad.nombre_tarea)
            ws.cell(row=row, column=3, value=actividad.fecha_inicio.strftime('%d-%m-%Y') if actividad.fecha_inicio else '')
            ws.cell(row=row, column=4, value=actividad.fecha_fin.strftime('%d-%m-%Y') if actividad.fecha_fin else '')
            ws.cell(row=row, column=5, value=actividad.duracion)
            ws.cell(row=row, column=6, value=actividad.progreso)
            ws.cell(row=row, column=7, value=actividad.recursos or '')
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'avances_{trabajador.nombre}_{proyecto.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'error')
        return redirect(url_for('controllers.avance_actividades'))

@controllers_bp.route('/proyectos_estado_4')
@login_required
def proyectos_estado_4():
    """Obtener proyectos en estado 4 (Desarrollo Aceptado) y 5 (Desarrollo Completado)
    para control y seguimiento de actividades.
    
    Filtra por recinto del usuario (excepto superadmin que ve todos).
    """
    try:
        grupo_id = request.args.get('grupo_id', type=int)
        
        # Construir la consulta base - Estados 4 y 5
        query = db.session.query(Requerimiento)\
            .join(Sector)\
            .join(Estado)\
            .filter(Requerimiento.id_estado.in_([4, 5]))\
            .filter(Requerimiento.activo == True)
        
        # Filtrar por recinto del usuario (excepto superadmin)
        if not current_user.is_superadmin():
            query = query.filter(Requerimiento.id_recinto == current_user.recinto_id)
        
        # Aplicar filtro por grupo si se proporciona
        if grupo_id:
            query = query.filter(Requerimiento.id_grupo == grupo_id)
        
        # Ejecutar la consulta con las relaciones cargadas
        proyectos = query.options(
                db.joinedload(Requerimiento.sector),
                db.joinedload(Requerimiento.estado),
                db.joinedload(Requerimiento.grupo)
            ).all()
        
        proyectos_data = []
        
        for proyecto in proyectos:
            # Para relaciones lazy='dynamic', usar query directamente
            actividades_proyecto = proyecto.actividades_proyecto.all()
            equipos_trabajo = proyecto.equipos_trabajo.all()
            
            # Calcular estadísticas del proyecto
            total_actividades = len(actividades_proyecto)
            actividades_completadas = len([act for act in actividades_proyecto if act.progreso and act.progreso >= 100])
            
            # Obtener progreso del proyecto desde la actividad raíz (EDT nivel 1)
            # En lugar de calcular el promedio, usar el valor ya calculado y propagado
            actividad_raiz = next(
                (act for act in actividades_proyecto if act.nivel_esquema == 1),
                None
            )
            
            if actividad_raiz and actividad_raiz.progreso is not None:
                progreso_promedio = round(float(actividad_raiz.progreso), 1)
            else:
                # Fallback: si no hay actividad raíz, calcular promedio (casos legacy)
                if total_actividades > 0:
                    progreso_total = sum([float(act.progreso) for act in actividades_proyecto if act.progreso is not None])
                    progreso_promedio = round(progreso_total / total_actividades, 1)
                else:
                    progreso_promedio = 0
            
            # Calcular progreso esperado y estado del cronograma
            from datetime import date
            fecha_actual = date.today()
            progreso_esperado = 0
            porcentaje_atraso_adelanto = 0
            estado_cronograma = "En tiempo"
            
            if total_actividades > 0:
                # Encontrar fecha de inicio más temprana y fecha de fin más tardía
                fechas_inicio = [act.fecha_inicio for act in actividades_proyecto if act.fecha_inicio]
                fechas_fin = [act.fecha_fin for act in actividades_proyecto if act.fecha_fin]
                
                if fechas_inicio and fechas_fin:
                    fecha_inicio_proyecto = min(fechas_inicio)
                    fecha_fin_proyecto = max(fechas_fin)
                    
                    # Calcular progreso esperado basado en la fecha actual
                    duracion_total_proyecto = (fecha_fin_proyecto - fecha_inicio_proyecto).days
                    
                    if duracion_total_proyecto > 0:
                        if fecha_actual <= fecha_inicio_proyecto:
                            # El proyecto aún no ha comenzado
                            progreso_esperado = 0
                        elif fecha_actual >= fecha_fin_proyecto:
                            # El proyecto debería estar completado
                            progreso_esperado = 100
                        else:
                            # El proyecto está en curso
                            dias_transcurridos = (fecha_actual - fecha_inicio_proyecto).days
                            progreso_esperado = round((dias_transcurridos / duracion_total_proyecto) * 100, 1)
                        
                        # Calcular diferencia entre progreso real y esperado
                        diferencia_progreso = progreso_promedio - progreso_esperado
                        
                        if abs(diferencia_progreso) >= 5:  # Considerar significativo si es >= 5%
                            if diferencia_progreso > 0:
                                estado_cronograma = "Adelantado"
                                porcentaje_atraso_adelanto = round(diferencia_progreso, 1)
                            else:
                                estado_cronograma = "Atrasado"
                                porcentaje_atraso_adelanto = round(abs(diferencia_progreso), 1)
            
            # Contar trabajadores únicos que han registrado avances (desde avance_actividad)
            avances_trabajadores = db.session.query(AvanceActividad.trabajador_id)\
                .filter(AvanceActividad.requerimiento_id == proyecto.id)\
                .distinct()\
                .all()
            total_trabajadores = len(avances_trabajadores)
            
            # Obtener responsable (primer trabajador del equipo)
            responsable = None
            if equipos_trabajo:
                primer_miembro = equipos_trabajo[0]
                if primer_miembro and primer_miembro.trabajador:
                    responsable = primer_miembro.trabajador.nombre
            
            proyecto_info = {
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'descripcion': proyecto.descripcion,
                'fecha': proyecto.fecha.strftime('%Y-%m-%d') if proyecto.fecha else None,
                'fecha_inicio': proyecto.fecha.strftime('%Y-%m-%d') if proyecto.fecha else None,
                'sector': proyecto.sector.nombre if proyecto.sector else None,
                'estado': proyecto.estado.nombre if proyecto.estado else None,
                'grupo': proyecto.grupo.nombre if proyecto.grupo else None,
                'grupo_id': proyecto.id_grupo,
                'responsable': responsable,
                'progreso': progreso_promedio,
                'progreso_esperado': progreso_esperado,
                'estado_cronograma': estado_cronograma,
                'porcentaje_atraso_adelanto': porcentaje_atraso_adelanto,
                'total_actividades': total_actividades,
                'actividades_completadas': actividades_completadas,
                'total_trabajadores': total_trabajadores
            }
            
            proyectos_data.append(proyecto_info)
        
        return jsonify({
            'success': True,
            'proyectos': proyectos_data,
            'total': len(proyectos_data),
            'grupo_filtrado': grupo_id
        })
        
    except Exception as e:
        print(f"❌ Error al obtener proyectos estado 4: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@controllers_bp.route('/grupos_disponibles')
def grupos_disponibles():
    """Obtener todos los grupos disponibles para filtrar proyectos"""
    try:
        # Obtener grupos que tienen proyectos asignados en estado 4
        grupos_con_proyectos = db.session.query(Grupo)\
            .join(Requerimiento, Grupo.id == Requerimiento.id_grupo)\
            .filter(Requerimiento.id_estado == 4)\
            .filter(Requerimiento.activo == True)\
            .filter(Grupo.activo == True)\
            .distinct()\
            .all()
        
        # También obtener todos los grupos activos para el selector
        todos_grupos = Grupo.query.filter_by(activo=True).all()
        
        grupos_data = []
        for grupo in todos_grupos:
            # Contar proyectos en este grupo
            count_proyectos = db.session.query(Requerimiento)\
                .filter(Requerimiento.id_grupo == grupo.id)\
                .filter(Requerimiento.id_estado == 4)\
                .filter(Requerimiento.activo == True)\
                .count()
            
            grupos_data.append({
                'id': grupo.id,
                'nombre': grupo.nombre,
                'total_proyectos': count_proyectos,
                'tiene_proyectos': count_proyectos > 0
            })
        
        return jsonify({
            'success': True,
            'grupos': grupos_data,
            'total_grupos': len(grupos_data)
        })
        
    except Exception as e:
        print(f"❌ Error al obtener grupos: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@controllers_bp.route('/proyecto_detalle/<int:proyecto_id>')
def proyecto_detalle(proyecto_id):
    """Obtener detalles completos de un proyecto específico"""
    try:
        print(f"🔍 Debug: Buscando proyecto con ID {proyecto_id}")
        
        # Primero verificar si existe el proyecto de una manera más simple
        proyecto_simple = db.session.query(Requerimiento).filter(Requerimiento.id == proyecto_id).first()
        print(f"🔍 Debug: Proyecto simple encontrado: {proyecto_simple is not None}")
        
        # Verificar cuántos proyectos hay en total
        total_proyectos = db.session.query(Requerimiento).count()
        print(f"🔍 Debug: Total de proyectos en BD: {total_proyectos}")
        
        # Obtener lista de IDs de proyectos existentes
        proyectos_ids = [p.id for p in db.session.query(Requerimiento.id).all()]
        print(f"🔍 Debug: IDs de proyectos existentes: {proyectos_ids}")
        
        # Obtener el proyecto con sus relaciones básicas (sin equipos_trabajo ya que es lazy='dynamic')
        proyecto = db.session.query(Requerimiento)\
            .options(
                db.joinedload(Requerimiento.sector),
                db.joinedload(Requerimiento.estado),
                db.joinedload(Requerimiento.tipologia),
                db.joinedload(Requerimiento.financiamiento),
                db.joinedload(Requerimiento.tipoproyecto)
            )\
            .filter(Requerimiento.id == proyecto_id)\
            .first()
        
        if not proyecto:
            return jsonify({
                'success': False,
                'error': f'Proyecto con ID {proyecto_id} no encontrado. IDs disponibles: {proyectos_ids}'
            }), 404
        
        # Obtener actividades del proyecto desde la tabla ActividadProyecto (solo activas)
        actividades = db.session.query(ActividadProyecto)\
            .filter(ActividadProyecto.requerimiento_id == proyecto_id)\
            .filter(ActividadProyecto.activo == True)\
            .all()
        
        # Preparar lista detallada de actividades
        actividades_detalle = []
        for actividad in actividades:
            estado_actividad = "Pendiente"
            color_estado = "secondary"
            if actividad.progreso:
                if actividad.progreso >= 100:
                    estado_actividad = "Completada"
                    color_estado = "success"
                elif actividad.progreso > 0:
                    estado_actividad = "En Progreso"
                    color_estado = "warning"
            
            actividades_detalle.append({
                'id': actividad.id,
                'nombre': actividad.nombre_tarea,  # Usar nombre_tarea en lugar de nombre
                'descripcion': actividad.edt or 'Sin EDT',  # Usar edt como descripción
                'progreso': float(actividad.progreso) if actividad.progreso else 0,  # Convertir Decimal a float
                'fecha_inicio': actividad.fecha_inicio.strftime('%Y-%m-%d') if actividad.fecha_inicio else None,
                'fecha_fin': actividad.fecha_fin.strftime('%Y-%m-%d') if actividad.fecha_fin else None,
                'estado': estado_actividad,
                'color_estado': color_estado
            })
        
        # Calcular estadísticas detalladas
        total_actividades = len(actividades)
        actividades_completadas = len([act for act in actividades if act.progreso and float(act.progreso) >= 100])
        actividades_en_progreso = len([act for act in actividades if act.progreso and 0 < float(act.progreso) < 100])
        actividades_pendientes = len([act for act in actividades if not act.progreso or float(act.progreso) == 0])
        
        # Obtener progreso del proyecto desde la actividad raíz (EDT nivel 1)
        # Este valor ya está calculado y propagado por la jerarquía
        actividad_raiz = next(
            (act for act in actividades if act.nivel_esquema == 1),
            None
        )
        
        if actividad_raiz and actividad_raiz.progreso is not None:
            progreso_promedio = round(float(actividad_raiz.progreso), 1)
            print(f"✅ Progreso obtenido de actividad raíz (EDT: {actividad_raiz.edt}): {progreso_promedio}%")
        else:
            # Fallback: si no hay actividad raíz, calcular promedio (casos legacy)
            print(f"⚠️ No se encontró actividad raíz para proyecto {proyecto_id}, calculando promedio")
            if total_actividades > 0:
                progreso_total = sum([float(act.progreso) for act in actividades if act.progreso is not None])
                progreso_promedio = round(progreso_total / total_actividades, 1)
            else:
                progreso_promedio = 0
        
        # Calcular progreso esperado basado en fechas
        from datetime import date
        fecha_actual = date.today()
        progreso_esperado = 0
        porcentaje_atraso_adelanto = 0
        estado_cronograma = "En tiempo"
        
        if total_actividades > 0:
            # Encontrar fecha de inicio más temprana y fecha de fin más tardía
            fechas_inicio = [act.fecha_inicio for act in actividades if act.fecha_inicio]
            fechas_fin = [act.fecha_fin for act in actividades if act.fecha_fin]
            
            if fechas_inicio and fechas_fin:
                fecha_inicio_proyecto = min(fechas_inicio)
                fecha_fin_proyecto = max(fechas_fin)
                
                # Calcular progreso esperado basado en la fecha actual
                duracion_total_proyecto = (fecha_fin_proyecto - fecha_inicio_proyecto).days
                
                if duracion_total_proyecto > 0:
                    if fecha_actual <= fecha_inicio_proyecto:
                        # El proyecto aún no ha comenzado
                        progreso_esperado = 0
                    elif fecha_actual >= fecha_fin_proyecto:
                        # El proyecto debería estar completado
                        progreso_esperado = 100
                    else:
                        # El proyecto está en curso
                        dias_transcurridos = (fecha_actual - fecha_inicio_proyecto).days
                        progreso_esperado = round((dias_transcurridos / duracion_total_proyecto) * 100, 1)
                    
                    # Calcular diferencia entre progreso real y esperado
                    diferencia_progreso = progreso_promedio - progreso_esperado
                    
                    if abs(diferencia_progreso) >= 5:  # Considerar significativo si es >= 5%
                        if diferencia_progreso > 0:
                            estado_cronograma = "Adelantado"
                            porcentaje_atraso_adelanto = round(diferencia_progreso, 1)
                        else:
                            estado_cronograma = "Atrasado"
                            porcentaje_atraso_adelanto = round(abs(diferencia_progreso), 1)
        
        # Obtener información del equipo (relación lazy='dynamic')
        equipo_info = []
        equipos_query = proyecto.equipos_trabajo  # Esto es un query object
        for miembro in equipos_query.all():  # Ejecutar la consulta con .all()
            if miembro.trabajador:
                equipo_info.append({
                    'id': miembro.trabajador.id,
                    'nombre': miembro.trabajador.nombre,
                    'especialidad': miembro.especialidad.nombre if miembro.especialidad else None,
                    'porcentaje_asignacion': miembro.porcentaje_asignacion if hasattr(miembro, 'porcentaje_asignacion') else None
                })
        
        proyecto_detalle = {
            'id': proyecto.id,
            'nombre': proyecto.nombre,
            'descripcion': proyecto.descripcion,
            'observacion': proyecto.observacion,
            'fecha': proyecto.fecha.strftime('%Y-%m-%d') if proyecto.fecha else None,
            'fecha_inicio': proyecto.fecha.strftime('%Y-%m-%d') if proyecto.fecha else None,
            'fecha_aceptacion': proyecto.fecha_aceptacion.strftime('%Y-%m-%d') if proyecto.fecha_aceptacion else None,
            'sector': proyecto.sector.nombre if proyecto.sector else None,
            'estado': proyecto.estado.nombre if proyecto.estado else None,
            'tipologia': proyecto.tipologia.nombre if proyecto.tipologia else None,
            'financiamiento': proyecto.financiamiento.nombre if proyecto.financiamiento else None,
            'tipo_proyecto': proyecto.tipoproyecto.nombre if proyecto.tipoproyecto else None,
            'progreso': progreso_promedio,
            'progreso_esperado': progreso_esperado,
            'estado_cronograma': estado_cronograma,
            'porcentaje_atraso_adelanto': porcentaje_atraso_adelanto,
            'total_actividades': total_actividades,
            'actividades_completadas': actividades_completadas,
            'actividades_en_progreso': actividades_en_progreso,
            'actividades_pendientes': actividades_pendientes,
            'total_trabajadores': len(equipo_info),
            'equipo': equipo_info,
            'responsable': equipo_info[0]['nombre'] if equipo_info else None,
            'actividades': actividades_detalle
        }
        
        return jsonify({
            'success': True,
            'proyecto': proyecto_detalle
        })
        
    except Exception as e:
        print(f"❌ Error al obtener detalle del proyecto {proyecto_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@controllers_bp.route('/control_actividades')
def control_actividades():
    """Renderizar la página de control de actividades"""
    return render_template('control-actividades.html')

@controllers_bp.route('/resumen')
@login_required
def resumen():
    """Página de resumen ejecutivo de proyectos con estado de actividades"""
    return render_template('resumen.html')

@controllers_bp.route('/api/resumen_proyectos')
@login_required
def api_resumen_proyectos():
    """API para obtener resumen de proyectos con sus actividades y estado"""
    try:
        from datetime import date
        
        # Obtener proyectos en estado 5 (Desarrollo Completado) con actividades
        proyectos = db.session.query(Requerimiento).join(
            ActividadProyecto, Requerimiento.id == ActividadProyecto.requerimiento_id
        ).filter(
            Requerimiento.id_estado == 5,
            Requerimiento.activo == True,
            ActividadProyecto.activo == True
        ).distinct().all()
        
        proyectos_data = []
        fecha_actual = date.today()
        
        for proyecto in proyectos:
            # Obtener todas las actividades del proyecto
            actividades = ActividadProyecto.query.filter_by(
                requerimiento_id=proyecto.id,
                activo=True
            ).order_by(ActividadProyecto.edt).all()
            
            actividades_data = []
            for actividad in actividades:
                # Calcular progreso esperado basado en fechas
                duracion_total = (actividad.fecha_fin - actividad.fecha_inicio).days
                progreso_esperado = 0
                estado_cronograma = "En Fecha"
                
                if duracion_total > 0:
                    if fecha_actual < actividad.fecha_inicio:
                        progreso_esperado = 0
                    elif fecha_actual > actividad.fecha_fin:
                        progreso_esperado = 100
                        if float(actividad.progreso or 0) < 100:
                            estado_cronograma = "Atrasado"
                    else:
                        dias_transcurridos = (fecha_actual - actividad.fecha_inicio).days
                        progreso_esperado = round((dias_transcurridos / duracion_total) * 100, 1)
                        
                        # Comparar progreso real vs esperado
                        progreso_real = float(actividad.progreso or 0)
                        if progreso_real < (progreso_esperado - 5):  # 5% de tolerancia
                            estado_cronograma = "Atrasado"
                
                actividades_data.append({
                    'id': actividad.id,
                    'edt': actividad.edt,
                    'nombre_tarea': actividad.nombre_tarea,
                    'nivel_esquema': actividad.nivel_esquema,
                    'fecha_inicio': actividad.fecha_inicio.strftime('%d-%m-%Y'),
                    'fecha_fin': actividad.fecha_fin.strftime('%d-%m-%Y'),
                    'duracion': actividad.duracion,
                    'progreso': float(actividad.progreso or 0),
                    'progreso_esperado': progreso_esperado,
                    'recursos': actividad.recursos or 'Sin asignar',
                    'estado_cronograma': estado_cronograma
                })
            
            proyectos_data.append({
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'descripcion': proyecto.descripcion or '',
                'sector': proyecto.sector.nombre if proyecto.sector else 'Sin sector',
                'estado': proyecto.estado.nombre if proyecto.estado else 'Sin estado',
                'actividades': actividades_data,
                'total_actividades': len(actividades_data),
                'actividades_atrasadas': len([a for a in actividades_data if a['estado_cronograma'] == 'Atrasado'])
            })
        
        return jsonify({
            'success': True,
            'proyectos': proyectos_data,
            'total_proyectos': len(proyectos_data)
        })
        
    except Exception as e:
        print(f"❌ Error en api_resumen_proyectos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@controllers_bp.route('/control_actividades_tabla')
def control_actividades_tabla():
    """Renderizar la página de control de actividades solo como tabla
    
    Muestra únicamente proyectos en estado 5 (Desarrollo Completado) 
    que tienen carta Gantt lista para control y seguimiento.
    """
    # Obtener solo los requerimientos en estado 5 (Desarrollo Completado) que tienen actividades
    requerimientos_con_actividades = db.session.query(Requerimiento.id).join(
        ActividadProyecto, Requerimiento.id == ActividadProyecto.requerimiento_id
    ).filter(
        ActividadProyecto.activo == True,
        Requerimiento.activo == True,
        Requerimiento.id_estado == 5  # Solo estado "Desarrollo Completado"
    ).distinct().all()
    
    requerimientos_ids = [r[0] for r in requerimientos_con_actividades]
    
    if not requerimientos_ids:
        # Si no hay requerimientos con actividades, retornar lista vacía
        return render_template('control-actividades-tabla.html', proyectos_data=[])
    
    # Obtener los requerimientos completos
    requerimientos = Requerimiento.query.filter(
        Requerimiento.id.in_(requerimientos_ids),
        Requerimiento.activo == True
    ).all()
    
    # Obtener información de actividades y avances para cada requerimiento
    proyectos_data = []
    for req in requerimientos:
        # Contar actividades
        total_actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=req.id, activo=True
        ).count()
        
        # Obtener avances
        avances = AvanceActividad.query.filter_by(requerimiento_id=req.id).all()
        total_avances = len(avances)
        
        # Calcular progreso promedio
        if avances:
            progreso_promedio = sum(avance.progreso_actual for avance in avances) / len(avances)
        else:
            progreso_promedio = 0.0
            
        # Obtener trabajadores únicos asignados
        trabajadores_ids = list(set(avance.trabajador_id for avance in avances))
        trabajadores = Trabajador.query.filter(Trabajador.id.in_(trabajadores_ids)).all() if trabajadores_ids else []
        
        proyectos_data.append({
            'requerimiento': req,
            'total_actividades': total_actividades,
            'total_avances': total_avances,
            'progreso_promedio': progreso_promedio,
            'trabajadores': trabajadores
        })
    
    return render_template('control-actividades-tabla.html', proyectos_data=proyectos_data)

# ==================================================================================
# Rutas CRUD para Grupos
@controllers_bp.route('/grupos', endpoint='ruta_grupos')
@login_required
def grupos():
    grupos = Grupo.query.all()
    return render_template('grupo.html', grupos=grupos)

@controllers_bp.route('/add_grupo', methods=['POST'], endpoint='add_grupo')
@login_required
def add_grupo():
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_add_grupo_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_grupos'))
        
        nombre = request.form['name']
        nuevo_grupo = Grupo(nombre=nombre)
        db.session.add(nuevo_grupo)
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_add_grupo_form_token'] = form_token
        
        flash('Grupo agregado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar grupo: {e}')
    return redirect(url_for('controllers.ruta_grupos'))

@controllers_bp.route('/update_grupo/<int:id>', methods=['POST'], endpoint='update_grupo')
@login_required
def update_grupo(id):
    try:
        # Verificar form_token para evitar duplicados
        form_token = request.form.get('form_token')
        session_token = session.get('last_edit_grupo_form_token')
        
        if form_token and form_token == session_token:
            flash('Operación ya procesada', 'warning')
            return redirect(url_for('controllers.ruta_grupos'))
        
        grupo = Grupo.query.get_or_404(id)
        grupo.nombre = request.form['name']
        db.session.commit()
        
        # Guardar token en sesión
        if form_token:
            session['last_edit_grupo_form_token'] = form_token
        
        flash('Grupo actualizado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar grupo: {e}')
    return redirect(url_for('controllers.ruta_grupos'))

@controllers_bp.route('/eliminar_grupo/<int:id>', methods=['POST'], endpoint='eliminar_grupo')
@login_required
def eliminar_grupo(id):
    try:
        grupo = Grupo.query.get_or_404(id)
        db.session.delete(grupo)
        db.session.commit()
        flash('Grupo eliminado exitosamente')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar grupo: {e}')
    return redirect(url_for('controllers.ruta_grupos'))


# Rutas para gestión de proyectos con archivos XLSX
@controllers_bp.route('/descargar-plantilla-gantt', methods=['GET'], endpoint='descargar_plantilla_gantt')
def descargar_plantilla_gantt():
    """Descargar plantilla de ejemplo para archivos Gantt"""
    try:
        import tempfile
        import os
        from flask import send_file
        
        # Datos de ejemplo
        datos_ejemplo = {
            'Id': [1, 2, 3, 4, 5, 6, 7],
            'Nivel de esquema': [1, 2, 3, 3, 2, 3, 3],
            'EDT': ['1', '1.1', '1.1.1', '1.1.2', '1.2', '1.2.1', '1.2.2'],
            'Nombre de tarea': [
                'Proyecto Ejemplo',
                'Fase de Planificación',
                'Análisis de Requerimientos',
                'Diseño del Sistema',
                'Fase de Desarrollo',
                'Desarrollo Frontend',
                'Desarrollo Backend'
            ],
            'Duración': [100, 30, 10, 20, 70, 35, 35],
            'Comienzo': [
                '2025-01-01',
                '2025-01-01',
                '2025-01-01',
                '2025-01-11',
                '2025-01-31',
                '2025-01-31',
                '2025-02-15'
            ],
            'Fin': [
                '2025-04-30',
                '2025-01-30',
                '2025-01-10',
                '2025-01-30',
                '2025-04-30',
                '2025-02-14',
                '2025-04-30'
            ],
            'Predecesoras': ['', '', '', '3', '2', '4', '6'],
            'Nombres de los recursos': [
                'Equipo Completo',
                'Analista Senior',
                'Analista de Sistemas',
                'Arquitecto de Software',
                'Equipo Desarrollo',
                'Desarrollador Frontend',
                'Desarrollador Backend'
            ]
        }
        
        # Crear DataFrame
        df = pd.DataFrame(datos_ejemplo)
        
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            df.to_excel(tmp_file.name, index=False)
            
            return send_file(
                tmp_file.name,
                as_attachment=True,
                download_name='plantilla_proyecto_gantt_ejemplo.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    except Exception as e:
        flash(f'Error al generar plantilla: {str(e)}', 'error')
        return redirect(url_for('proyectos.ruta_proyecto_llenar'))

# NOTA: Esta función se movió al blueprint de proyectos (proyectos_controller.py)  
# @controllers_bp.route('/proyecto-llenar', endpoint='ruta_proyecto_llenar')
# @login_required
# def proyecto_llenar():
#     """Página para subir archivos XLSX y asignar proyectos a requerimientos"""
#     return redirect(url_for('proyectos.ruta_proyecto_llenar'))

# MIGRADO: Esta función se movió a proyectos_controller.py para mejor arquitectura
# @controllers_bp.route('/procesar-proyecto-xlsx', methods=['POST'], endpoint='procesar_proyecto_xyz')
def procesar_proyecto_xlsx_MIGRADO():
    """FUNCIÓN MIGRADA A proyectos_controller.py - NO USAR"""
    return jsonify({'success': False, 'message': 'Esta función se movió a proyectos_controller.py'})

# MIGRADO: Funciones auxiliares también movidas a proyectos_controller.py
def limpiar_trabajadores_huerfanos_MIGRADO():
    """FUNCIÓN MIGRADA - NO USAR"""
    pass

def extraer_y_crear_trabajadores_desde_recursos_MIGRADO():
    """FUNCIÓN MIGRADA - NO USAR"""  
    pass

def crear_avances_actividad_MIGRADO():
    """FUNCIÓN MIGRADA - NO USAR"""
    pass

def procesar_recursos_actividad_MIGRADO():
    """FUNCIÓN MIGRADA - NO USAR"""
    pass

@controllers_bp.route('/ver-recursos-proyecto/<int:requerimiento_id>', methods=['GET'])
def ver_recursos_proyecto(requerimiento_id):
    """Ver los recursos procesados de un proyecto"""
    try:
        requerimiento = Requerimiento.query.get_or_404(requerimiento_id)
        actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=requerimiento_id, activo=True
        ).all()
        
        recursos_summary = {
            'requerimiento': {
                'id': requerimiento.id,
                'nombre': requerimiento.nombre,
                'proyecto': requerimiento.proyecto
            },
            'actividades_con_recursos': [],
            'trabajadores_involucrados': [],
            'total_actividades': len(actividades),
            'actividades_con_recursos_count': 0
        }
        
        trabajadores_ids = set()
        
        for actividad in actividades:
            if actividad.datos_adicionales and 'recursos_procesados' in actividad.datos_adicionales:
                recursos_actividad = {
                    'edt': actividad.edt,
                    'nombre_tarea': actividad.nombre_tarea,
                    'recursos_originales': actividad.recursos,
                    'recursos_procesados': actividad.datos_adicionales['recursos_procesados']
                }
                recursos_summary['actividades_con_recursos'].append(recursos_actividad)
                recursos_summary['actividades_con_recursos_count'] += 1
                
                # Recopilar IDs de trabajadores
                for recurso in actividad.datos_adicionales['recursos_procesados']:
                    trabajadores_ids.add(recurso['trabajador_id'])
        
        # Obtener información de trabajadores
        if trabajadores_ids:
            trabajadores = Trabajador.query.filter(Trabajador.id.in_(trabajadores_ids)).all()
            recursos_summary['trabajadores_involucrados'] = [
                {
                    'id': t.id,
                    'nombre': t.nombre,
                    'nombrecorto': t.nombrecorto,
                    'profesion': t.profesion,
                    'email': t.email
                }
                for t in trabajadores
            ]
        
        return jsonify(recursos_summary)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@controllers_bp.route('/ver-gantt-proyecto/<int:requerimiento_id>', endpoint='ver_gantt_proyecto')
def ver_gantt_proyecto(requerimiento_id):
    """Ver el diagrama de Gantt de un proyecto"""
    requerimiento = Requerimiento.query.get_or_404(requerimiento_id)
    actividades_obj = ActividadProyecto.query.filter_by(
        requerimiento_id=requerimiento_id, activo=True
    ).order_by(ActividadProyecto.edt).all()
    
    # Convertir objetos SQLAlchemy a diccionarios para serialización JSON
    actividades = []
    for act in actividades_obj:
        actividades.append({
            'id': act.id,
            'edt': act.edt,
            'nombre_tarea': act.nombre_tarea,
            'nivel_esquema': act.nivel_esquema,
            'duracion': act.duracion,
            'fecha_inicio': act.fecha_inicio.strftime('%Y-%m-%d') if act.fecha_inicio else None,
            'fecha_inicio_display': act.fecha_inicio.strftime('%d/%m/%Y') if act.fecha_inicio else 'No definida',
            'fecha_fin': act.fecha_fin.strftime('%Y-%m-%d') if act.fecha_fin else None,
            'fecha_fin_display': act.fecha_fin.strftime('%d/%m/%Y') if act.fecha_fin else 'No definida',
            'progreso': float(act.progreso) if act.progreso else 0.0,
            'predecesoras': act.predecesoras or '',
            'recursos': act.recursos or '',
            'activo': act.activo
        })
    
    return render_template('gantt-proyecto.html', 
                         requerimiento=requerimiento,
                         actividades=actividades)

@controllers_bp.route('/gantt-general', endpoint='gantt_general')
def gantt_general():
    """Vista general de diagramas Gantt de todos los proyectos"""
    # Obtener todos los requerimientos con proyectos asignados
    requerimientos = Requerimiento.query.filter(
        Requerimiento.proyecto.isnot(None),
        Requerimiento.activo == True
    ).all()
    
    proyectos_data = []
    for req in requerimientos:
        # Contar actividades
        total_actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=req.id, activo=True
        ).count()
        
        # Obtener fecha de inicio y fin del proyecto
        primera_actividad = ActividadProyecto.query.filter_by(
            requerimiento_id=req.id, activo=True
        ).order_by(ActividadProyecto.fecha_inicio).first()
        
        ultima_actividad = ActividadProyecto.query.filter_by(
            requerimiento_id=req.id, activo=True
        ).order_by(ActividadProyecto.fecha_fin.desc()).first()
        
        # Calcular progreso promedio
        avances = AvanceActividad.query.filter_by(requerimiento_id=req.id).all()
        if avances:
            progreso_promedio = sum(avance.progreso_actual for avance in avances) / len(avances)
        else:
            progreso_promedio = 0.0
        
        proyectos_data.append({
            'requerimiento': req,
            'total_actividades': total_actividades,
            'fecha_inicio': primera_actividad.fecha_inicio if primera_actividad else None,
            'fecha_fin': ultima_actividad.fecha_fin if ultima_actividad else None,
            'progreso_promedio': progreso_promedio
        })
    
    return render_template('gantt-general.html', proyectos_data=proyectos_data)

@controllers_bp.route('/obtener-detalle-proyecto/<int:requerimiento_id>', endpoint='obtener_detalle_proyecto')
def obtener_detalle_proyecto(requerimiento_id):
    """Obtener detalles de un proyecto vía AJAX"""
    try:
        actividades = ActividadProyecto.query.filter_by(
            requerimiento_id=requerimiento_id, activo=True
        ).order_by(ActividadProyecto.edt).all()
        
        actividades_data = []
        for act in actividades:
            actividades_data.append({
                'edt': act.edt,
                'nombre_tarea': act.nombre_tarea,
                'nivel_esquema': act.nivel_esquema,
                'duracion': act.duracion,
                'fecha_inicio': act.fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': act.fecha_fin.strftime('%d/%m/%Y'),
                'progreso': float(act.progreso) if act.progreso else 0.0,
                'predecesoras': act.predecesoras or '',
                'recursos': act.recursos or ''
            })
        
        return jsonify({'success': True, 'actividades': actividades_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@controllers_bp.route('/exportar_actividades_xlsx', methods=['GET'])
def exportar_actividades_xlsx():
    """Exportar todas las actividades de proyectos a un archivo Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import make_response
        import io
        
        # Crear un nuevo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Actividades Proyecto"
        
        # Definir las columnas según los requerimientos
        columnas = [
            'Id',
            'Nivel de esquema', 
            'EDT',
            'Nombre de tarea',
            'Duración',
            'Comienzo',
            'Fin',
            'Predecesoras',
            'Nombres de los recursos',
            'Progreso (%)'
        ]
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Agregar headers
        for col_num, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=columna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Obtener todas las actividades de proyectos activos ordenadas por fecha de registro del requerimiento
        actividades = ActividadProyecto.query.join(Requerimiento).filter(
            ActividadProyecto.activo == True
        ).order_by(
            Requerimiento.created_at.asc(),
            ActividadProyecto.edt
        ).all()
        
        print(f"📊 Exportando {len(actividades)} actividades a Excel")
        
        # Agregar datos
        for row_num, actividad in enumerate(actividades, 2):
            # Formatear fechas
            fecha_inicio = actividad.fecha_inicio.strftime('%d/%m/%Y') if actividad.fecha_inicio else ''
            fecha_fin = actividad.fecha_fin.strftime('%d/%m/%Y') if actividad.fecha_fin else ''
            
            # Obtener progreso directamente de la tabla actividad_proyecto
            progreso = float(actividad.progreso) if actividad.progreso else 0.0
            
            # Datos de la fila
            datos_fila = [
                actividad.id,
                actividad.nivel_esquema or 1,
                actividad.edt or '',
                actividad.nombre_tarea or '',
                actividad.duracion or 0,
                fecha_inicio,
                fecha_fin,
                actividad.predecesoras or '',
                actividad.recursos or '',
                progreso
            ]
            
            # Escribir datos en las celdas
            for col_num, valor in enumerate(datos_fila, 1):
                ws.cell(row=row_num, column=col_num, value=valor)
        
        # Ajustar ancho de columnas
        column_widths = [8, 15, 12, 40, 12, 12, 12, 15, 30, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Crear el archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        from datetime import datetime
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'actividades_proyecto_{fecha_actual}.xlsx'
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        print(f"✅ Archivo Excel generado: {nombre_archivo} con {len(actividades)} actividades")
        return response
        
    except ImportError:
        return jsonify({
            'success': False, 
            'message': 'Error: openpyxl no está instalado. Ejecute: pip install openpyxl'
        }), 500
        
    except Exception as e:
        print(f"❌ Error al generar Excel: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Error al generar archivo Excel: {str(e)}'
        }), 500


@controllers_bp.route('/consultar-trabajadores', methods=['GET'])
def consultar_trabajadores():
    """Endpoint para consultar trabajadores y sus asignaciones actuales"""
    try:
        # Obtener todos los trabajadores con sus asignaciones
        trabajadores_con_asignaciones = db.session.query(
            Trabajador.id,
            Trabajador.nombre,
            Trabajador.nombrecorto,
            db.func.count(AvanceActividad.id).label('num_asignaciones')
        ).outerjoin(
            AvanceActividad, Trabajador.id == AvanceActividad.trabajador_id
        ).group_by(
            Trabajador.id, Trabajador.nombre, Trabajador.nombrecorto
        ).all()
        
        # Formatear resultado
        trabajadores_data = []
        for trabajador in trabajadores_con_asignaciones:
            # Obtener detalles de asignaciones
            asignaciones = db.session.query(
                ActividadProyecto.edt,
                ActividadProyecto.nombre_tarea,
                AvanceActividad.porcentaje_asignacion,
                Requerimiento.nombre.label('proyecto_nombre')
            ).join(
                AvanceActividad, ActividadProyecto.id == AvanceActividad.actividad_id
            ).join(
                Requerimiento, ActividadProyecto.requerimiento_id == Requerimiento.id
            ).filter(
                AvanceActividad.trabajador_id == trabajador.id
            ).all()
            
            trabajadores_data.append({
                'id': trabajador.id,
                'nombre': trabajador.nombre,
                'nombre_corto': trabajador.nombrecorto,
                'num_asignaciones': trabajador.num_asignaciones,
                'asignaciones': [
                    {
                        'edt': asig.edt,
                        'nombre_tarea': asig.nombre_tarea,
                        'porcentaje_asignacion': asig.porcentaje_asignacion,
                        'proyecto': asig.proyecto_nombre
                    }
                    for asig in asignaciones
                ]
            })
        
        return jsonify({
            'success': True,
            'trabajadores': trabajadores_data,
            'total_trabajadores': len(trabajadores_data),
            'trabajadores_con_asignaciones': len([t for t in trabajadores_data if t['num_asignaciones'] > 0]),
            'trabajadores_sin_asignaciones': len([t for t in trabajadores_data if t['num_asignaciones'] == 0])
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al consultar trabajadores: {str(e)}'
        }), 500

@controllers_bp.route('/limpiar-trabajadores', methods=['POST'])
def limpiar_trabajadores():
    """Endpoint para limpiar trabajadores que no están asignados a ninguna actividad"""
    try:
        count_eliminados = limpiar_trabajadores_huerfanos()
        
        return jsonify({
            'success': True,
            'message': f'Se eliminaron {count_eliminados} trabajadores sin asignaciones',
            'count_eliminados': count_eliminados
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al limpiar trabajadores: {str(e)}'
        }), 500

@controllers_bp.route('/subir_control_actividades', methods=['POST'])
def subir_control_actividades():
    """Función para procesar archivo Excel de control de actividades"""
    try:
        print("🚀 Iniciando procesamiento de archivo de control...")
        
        # Verificar que se haya enviado un archivo
        if 'archivo' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            }), 400
        
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            }), 400
        
        # Verificar extensión del archivo
        if not archivo.filename.lower().endswith('.xlsx'):
            return jsonify({
                'success': False,
                'message': 'El archivo debe ser un Excel (.xlsx)'
            }), 400
        
        # Generar sesión única para esta subida
        sesion_subida = str(uuid.uuid4())[:8]
        nombre_archivo = secure_filename(archivo.filename)
        
        print(f"📄 Procesando archivo: {nombre_archivo}")
        print(f"🆔 Sesión de subida: {sesion_subida}")
        
        # Leer el archivo Excel
        try:
            workbook = openpyxl.load_workbook(archivo, data_only=True)
            sheet = workbook.active
            print(f"📊 Hoja de trabajo cargada: {sheet.title}")
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error al leer el archivo Excel: {str(e)}'
            }), 400
        
        # Obtener las filas de datos
        filas = list(sheet.iter_rows(values_only=True))
        
        if len(filas) < 2:  # Al menos debe tener encabezados y una fila de datos
            return jsonify({
                'success': False,
                'message': 'El archivo debe contener al menos los encabezados y una fila de datos'
            }), 400
        
        # Procesar encabezados
        encabezados = filas[0]
        print(f"📋 Encabezados encontrados: {encabezados}")
        
        # Mapear columnas esperadas
        columnas_esperadas = {
            'edt': None,
            'nombre_tarea': None,
            'fecha_inicio': None,
            'fecha_fin': None,
            'duracion': None,
            'progreso': None,
            'recursos': None,
            'predecesoras': None,
            'proyecto': None
        }
        
        # Mapear encabezados (buscar por nombres similares)
        for i, encabezado in enumerate(encabezados):
            if not encabezado:
                continue
                
            encabezado_lower = str(encabezado).lower().strip()
            
            # Mapeo de columnas con múltiples variaciones
            if any(term in encabezado_lower for term in ['edt', 'código', 'codigo', 'wbs', 'id']):
                columnas_esperadas['edt'] = i
            elif any(term in encabezado_lower for term in ['nombre', 'task name', 'actividad', 'tarea', 'activity']):
                columnas_esperadas['nombre_tarea'] = i
            elif any(term in encabezado_lower for term in ['inicio', 'start', 'comienzo', 'fecha inicio', 'fecha_inicio', 'begin']):
                columnas_esperadas['fecha_inicio'] = i
            elif any(term in encabezado_lower for term in ['fin', 'end', 'finish', 'final', 'fecha fin', 'fecha_fin', 'término', 'termino']):
                columnas_esperadas['fecha_fin'] = i
            elif any(term in encabezado_lower for term in ['duración', 'duracion', 'duration', 'días', 'dias']):
                columnas_esperadas['duracion'] = i
            elif any(term in encabezado_lower for term in ['progreso', 'progress', '% complete', 'complete', 'avance', 'porcentaje']):
                columnas_esperadas['progreso'] = i
            elif any(term in encabezado_lower for term in ['recursos', 'resource', 'assigned', 'asignado', 'responsable']):
                columnas_esperadas['recursos'] = i
            elif any(term in encabezado_lower for term in ['predecesoras', 'predecessors', 'dependencias', 'dependencies']):
                columnas_esperadas['predecesoras'] = i
            elif any(term in encabezado_lower for term in ['proyecto', 'project', 'requerimiento']):
                columnas_esperadas['proyecto'] = i
        
        print(f"🗂️ Mapeo de columnas realizado:")
        for columna, indice in columnas_esperadas.items():
            if indice is not None:
                print(f"   ✅ {columna}: Columna {indice} = '{encabezados[indice]}'")
            else:
                print(f"   ❌ {columna}: No encontrada")
        
        # Verificar columnas obligatorias
        columnas_obligatorias = ['edt', 'nombre_tarea', 'fecha_inicio', 'fecha_fin']
        columnas_faltantes = []
        
        for columna in columnas_obligatorias:
            if columnas_esperadas[columna] is None:
                columnas_faltantes.append(columna)
        
        if columnas_faltantes:
            mensaje_error = f'Faltan las siguientes columnas obligatorias: {", ".join(columnas_faltantes)}\n\n'
            mensaje_error += 'Columnas encontradas en el archivo:\n'
            for i, enc in enumerate(encabezados):
                if enc:
                    mensaje_error += f'  - Columna {i+1}: "{enc}"\n'
            
            mensaje_error += '\nAsegúrate de que tu archivo tenga columnas con nombres similares a:\n'
            mensaje_error += '  - EDT/Código/ID\n'
            mensaje_error += '  - Nombre/Tarea/Actividad\n'  
            mensaje_error += '  - Fecha Inicio/Start/Comienzo\n'
            mensaje_error += '  - Fecha Fin/End/Final'
            
            return jsonify({
                'success': False,
                'message': mensaje_error
            }), 400
        
        # Procesar filas de datos
        actividades_procesadas = 0
        actividades_actualizadas = 0
        actividades_nuevas = 0
        errores = []
        
        for num_fila, fila in enumerate(filas[1:], start=2):  # Empezar desde fila 2 (primera fila de datos)
            try:
                print(f"📝 Procesando fila {num_fila}...")
                
                # Extraer datos de la fila
                edt = fila[columnas_esperadas['edt']] if columnas_esperadas['edt'] is not None else None
                nombre_tarea = fila[columnas_esperadas['nombre_tarea']] if columnas_esperadas['nombre_tarea'] is not None else None
                
                # Saltar filas vacías
                if not edt or not nombre_tarea:
                    print(f"⚠️ Fila {num_fila} está vacía, saltando...")
                    continue
                
                # Procesar datos
                datos_fila = {
                    'edt': str(edt).strip(),
                    'nombre_tarea': str(nombre_tarea).strip(),
                    'fecha_inicio': fila[columnas_esperadas['fecha_inicio']] if columnas_esperadas['fecha_inicio'] is not None else None,
                    'fecha_fin': fila[columnas_esperadas['fecha_fin']] if columnas_esperadas['fecha_fin'] is not None else None,
                    'duracion': fila[columnas_esperadas['duracion']] if columnas_esperadas['duracion'] is not None else None,
                    'progreso': fila[columnas_esperadas['progreso']] if columnas_esperadas['progreso'] is not None else None,
                    'recursos': fila[columnas_esperadas['recursos']] if columnas_esperadas['recursos'] is not None else None,
                    'predecesoras': fila[columnas_esperadas['predecesoras']] if columnas_esperadas['predecesoras'] is not None else None,
                }
                
                # Procesar duración - extraer número de texto como "5 días", "10", etc.
                if datos_fila['duracion'] is not None:
                    duracion_str = str(datos_fila['duracion']).strip()
                    # Extraer números del texto usando regex
                    import re
                    numeros = re.findall(r'\d+', duracion_str)
                    if numeros:
                        datos_fila['duracion'] = int(numeros[0])  # Usar el primer número encontrado
                    else:
                        datos_fila['duracion'] = None
                
                # MÉTODO SIMPLIFICADO: Solo buscar y actualizar actividades existentes
                # NO crear nuevas actividades automáticamente
                actividad_existente = ActividadProyecto.query.filter_by(edt=datos_fila['edt']).first()
                
                if actividad_existente:
                    proyecto = actividad_existente.requerimiento
                    print(f"🔍 Actividad existente encontrada: {datos_fila['edt']} - Proyecto: {proyecto.nombre} (ID: {proyecto.id})")
                    
                    # Actualizar actividad existente - DATOS ANTERIORES (incluyendo recursos)
                    datos_anteriores = {
                        'edt': actividad_existente.edt,
                        'nombre_tarea': actividad_existente.nombre_tarea,
                        'fecha_inicio': actividad_existente.fecha_inicio.isoformat() if actividad_existente.fecha_inicio else None,
                        'fecha_fin': actividad_existente.fecha_fin.isoformat() if actividad_existente.fecha_fin else None,
                        'duracion': int(actividad_existente.duracion) if actividad_existente.duracion else None,
                        'progreso': float(actividad_existente.progreso) if actividad_existente.progreso else 0.0,
                        'recursos': actividad_existente.recursos,
                        'predecesoras': actividad_existente.predecesoras
                    }
                    
                    # Actualizar campos
                    if datos_fila['nombre_tarea']:
                        actividad_existente.nombre_tarea = datos_fila['nombre_tarea']
                    
                    if datos_fila['fecha_inicio']:
                        try:
                            if isinstance(datos_fila['fecha_inicio'], datetime):
                                actividad_existente.fecha_inicio = datos_fila['fecha_inicio'].date()
                            elif hasattr(datos_fila['fecha_inicio'], 'date'):  # Para objetos date
                                actividad_existente.fecha_inicio = datos_fila['fecha_inicio']
                            else:
                                # Usar el nuevo parser de fechas en español
                                fecha_parseada = parsear_fecha_espanol(datos_fila['fecha_inicio'])
                                if fecha_parseada:
                                    actividad_existente.fecha_inicio = fecha_parseada
                        except Exception as e:
                            print(f"⚠️ Error al procesar fecha_inicio en fila {num_fila}: {e}")
                    
                    if datos_fila['fecha_fin']:
                        try:
                            if isinstance(datos_fila['fecha_fin'], datetime):
                                actividad_existente.fecha_fin = datos_fila['fecha_fin'].date()
                            elif hasattr(datos_fila['fecha_fin'], 'date'):  # Para objetos date
                                actividad_existente.fecha_fin = datos_fila['fecha_fin']
                            else:
                                # Usar el nuevo parser de fechas en español
                                fecha_parseada = parsear_fecha_espanol(datos_fila['fecha_fin'])
                                if fecha_parseada:
                                    actividad_existente.fecha_fin = fecha_parseada
                        except Exception as e:
                            print(f"⚠️ Error al procesar fecha_fin en fila {num_fila}: {e}")
                    
                    if datos_fila['duracion'] is not None:
                        actividad_existente.duracion = int(datos_fila['duracion'])
                    
                    if datos_fila['progreso'] is not None:
                        progreso_valor = float(datos_fila['progreso'])
                        # Si el progreso está en formato decimal (0.5), convertir a porcentaje (50)
                        if progreso_valor <= 1.0:
                            progreso_valor = progreso_valor * 100
                        actividad_existente.progreso = progreso_valor
                    
                    # PROCESAMIENTO DE RECURSOS: Revisar si fueron modificados y actualizar
                    if datos_fila['recursos']:
                        recursos_nuevos = str(datos_fila['recursos']).strip()
                        recursos_actuales = actividad_existente.recursos.strip() if actividad_existente.recursos else ""
                        
                        print(f"🧑‍💼 Revisando recursos para actividad {datos_fila['edt']}")
                        print(f"   📋 Recursos actuales: '{recursos_actuales}'")
                        print(f"   📄 Recursos del Excel: '{recursos_nuevos}'")
                        
                        # Comparar si los recursos han cambiado
                        if recursos_nuevos != recursos_actuales:
                            print(f"   🔄 Los recursos han cambiado, actualizando...")
                            actividad_existente.recursos = recursos_nuevos
                        else:
                            print(f"   ✅ Los recursos no han cambiado, mantiendo actuales")
                        
                        # Extraer y crear trabajadores desde los recursos asignados
                        trabajadores_asignados = extraer_y_crear_trabajadores_desde_recursos(recursos_nuevos)
                        
                        # Crear avances de actividad para los trabajadores asignados
                        if trabajadores_asignados:
                            progreso_para_avance = actividad_existente.progreso if actividad_existente.progreso else 0.0
                            crear_avances_actividad(
                                proyecto.id, 
                                actividad_existente.id, 
                                recursos_nuevos, 
                                progreso_para_avance
                            )
                            print(f"📋 Creados/actualizados avances para {len(trabajadores_asignados)} trabajadores en actividad {actividad_existente.edt}")
                        else:
                            print(f"⚠️ No se pudieron extraer trabajadores de: {recursos_nuevos}")
                    elif actividad_existente.recursos:
                        # Si no hay recursos en el Excel pero sí en la actividad, mantener los actuales
                        print(f"📋 No hay recursos en Excel, manteniendo recursos actuales: {actividad_existente.recursos}")
                    else:
                        print(f"📋 No hay recursos ni en Excel ni en la actividad")
                    
                    if datos_fila['predecesoras']:
                        actividad_existente.predecesoras = str(datos_fila['predecesoras'])
                    
                    actividad_existente.updated_at = datetime.utcnow()
                    
                    # Crear registro de historial
                    historial = HistorialControl(
                        sesion_subida=sesion_subida,
                        nombre_archivo=nombre_archivo,
                        actividad_id=actividad_existente.id,
                        requerimiento_id=proyecto.id,
                        tipo_operacion='UPDATE',
                        datos_anteriores=datos_anteriores,
                        datos_nuevos={
                            'edt': actividad_existente.edt,
                            'nombre_tarea': actividad_existente.nombre_tarea,
                            'fecha_inicio': actividad_existente.fecha_inicio.isoformat() if actividad_existente.fecha_inicio else None,
                            'fecha_fin': actividad_existente.fecha_fin.isoformat() if actividad_existente.fecha_fin else None,
                            'duracion': int(actividad_existente.duracion) if actividad_existente.duracion else None,
                            'progreso': float(actividad_existente.progreso) if actividad_existente.progreso else 0.0,
                            'recursos': actividad_existente.recursos,
                            'predecesoras': actividad_existente.predecesoras
                        },
                        fila_excel=num_fila
                    )
                    
                    db.session.add(historial)
                    actividades_actualizadas += 1
                    
                else:
                    # Actividad NO encontrada en la base de datos
                    print(f"⚠️ Actividad con EDT '{datos_fila['edt']}' NO encontrada en la BD - IGNORANDO")
                    print(f"   📋 Archivo de control solo debe actualizar actividades existentes")
                    print(f"   💡 Para crear nuevas actividades, usar el proceso de 'Llenar Proyecto'")
                    continue  # Saltar esta actividad
                
                actividades_procesadas += 1
                
            except Exception as e:
                error_msg = f'Error en fila {num_fila}: {str(e)}'
                print(f"❌ {error_msg}")
                errores.append({
                    'fila': num_fila,
                    'mensaje': error_msg
                })
                continue
        
        # Guardar cambios
        try:
            db.session.commit()
            
            # Limpiar trabajadores huérfanos después del procesamiento exitoso
            trabajadores_eliminados = limpiar_trabajadores_huerfanos()
            
            print(f"✅ Procesamiento completado:")
            print(f"   📊 Total procesadas: {actividades_procesadas}")
            print(f"   🔄 Actualizadas: {actividades_actualizadas}")
            print(f"   ➕ Nuevas: {actividades_nuevas}")
            print(f"   🧹 Trabajadores eliminados (sin asignaciones): {trabajadores_eliminados}")
            print(f"   ❌ Errores: {len(errores)}")
            
            return jsonify({
                'success': True,
                'message': f'Archivo procesado exitosamente. {actividades_procesadas} actividades procesadas. {trabajadores_eliminados} trabajadores sin asignaciones eliminados.',
                'actualizadas': actividades_actualizadas,
                'nuevas': actividades_nuevas,
                'trabajadores_eliminados': trabajadores_eliminados,
                'errores': errores,
                'sesion_subida': sesion_subida
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'Error al guardar los cambios: {str(e)}'
            }), 500
            
    except Exception as e:
        print(f"❌ Error general en subida de control: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        }), 500


# Rutas para historial de avances
@controllers_bp.route('/historial-avances')
def historial_avances():
    """Página para ver el historial de cambios en avances de actividades"""
    try:
        # Obtener filtros de los parámetros
        proyecto_id = request.args.get('proyecto_id', type=int)
        trabajador_id = request.args.get('trabajador_id', type=int)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        # Obtener listas para los filtros
        proyectos = Requerimiento.query.filter_by(id_estado=4).order_by(Requerimiento.nombre).all()
        trabajadores = Trabajador.query.filter_by(activo=True).order_by(Trabajador.nombre).all()
        
        return render_template('historial-avances.html', 
                               proyectos=proyectos,
                               trabajadores=trabajadores,
                               proyecto_id=proyecto_id,
                               trabajador_id=trabajador_id,
                               fecha_desde=fecha_desde,
                               fecha_hasta=fecha_hasta)
    except Exception as e:
        flash(f'Error al cargar página de historial: {str(e)}', 'error')
        return redirect(url_for('controllers.index'))


@controllers_bp.route('/api/historial-avances')
def api_historial_avances():
    """API para obtener historial de avances con filtros"""
    try:
        from datetime import datetime, date
        
        # Obtener filtros
        proyecto_id = request.args.get('proyecto_id', type=int)
        trabajador_id = request.args.get('trabajador_id', type=int)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        limit = request.args.get('limit', 100, type=int)
        
        # Construir query base
        query = HistorialAvanceActividad.query
        
        # Aplicar filtros
        if proyecto_id:
            query = query.filter(HistorialAvanceActividad.requerimiento_id == proyecto_id)
            
        if trabajador_id:
            query = query.filter(HistorialAvanceActividad.trabajador_id == trabajador_id)
            
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(HistorialAvanceActividad.fecha_cambio >= fecha_desde_obj)
            except ValueError:
                pass
                
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                # Agregar un día para incluir todo el día hasta
                fecha_hasta_obj = datetime.combine(fecha_hasta_obj, datetime.max.time())
                query = query.filter(HistorialAvanceActividad.fecha_cambio <= fecha_hasta_obj)
            except ValueError:
                pass
        
        # Ordenar por fecha descendente (más recientes primero) y limitar
        historial = query.order_by(HistorialAvanceActividad.fecha_cambio.desc()).limit(limit).all()
        
        # Agrupar por sesión de guardado para mejor visualización
        sesiones = {}
        for entry in historial:
            if entry.sesion_guardado not in sesiones:
                sesiones[entry.sesion_guardado] = {
                    'sesion_id': entry.sesion_guardado,
                    'fecha_cambio': entry.fecha_cambio,
                    'trabajador': entry.trabajador.nombre if entry.trabajador else 'N/A',
                    'proyecto': entry.requerimiento.nombre if entry.requerimiento else 'N/A',
                    'cambios': []
                }
            
            sesiones[entry.sesion_guardado]['cambios'].append({
                'actividad_edt': entry.actividad.edt if entry.actividad else 'N/A',
                'actividad_nombre': entry.actividad.nombre_tarea if entry.actividad else 'N/A',
                'progreso_anterior': entry.progreso_anterior,
                'progreso_nuevo': entry.progreso_nuevo,
                'diferencia': entry.diferencia,
                'comentarios': entry.comentarios
            })
        
        # Convertir a lista y ordenar por fecha
        sesiones_list = list(sesiones.values())
        sesiones_list.sort(key=lambda x: x['fecha_cambio'], reverse=True)
        
        return jsonify({
            'success': True,
            'sesiones': sesiones_list,
            'total_sesiones': len(sesiones_list),
            'total_cambios': len(historial)
        })
        
    except Exception as e:
        print(f"❌ ERROR en api_historial_avances: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al obtener historial: {str(e)}'
        }), 500

@controllers_bp.route('/api/trabajador/<int:id>', methods=['GET'], endpoint='api_trabajador_detalle')
@login_required
def api_trabajador_detalle(id):
    """API para obtener datos completos de un trabajador"""
    try:
        # Verificar permisos de acceso
        user_role = current_user.rol.value if hasattr(current_user.rol, 'value') else str(current_user.rol)
        
        trabajador = Trabajador.query.options(
            joinedload(Trabajador.sector),
            joinedload(Trabajador.recinto).joinedload(Recinto.tiporecinto).joinedload(TipoRecinto.sector)
        ).get_or_404(id)
        
        # Verificar si el usuario tiene permisos para ver este trabajador
        if user_role.upper() != 'SUPERADMIN':
            # Solo puede ver trabajadores del mismo recinto o a sí mismo
            user_recinto_id = current_user.recinto_id
            if trabajador.id != current_user.id and trabajador.recinto_id != user_recinto_id:
                return jsonify({'error': 'No tiene permisos para acceder a este trabajador'}), 403
        
        # Construir respuesta con todos los datos
        response_data = {
            'id': trabajador.id,
            'nombre': trabajador.nombre,
            'rut': trabajador.rut,
            'profesion': trabajador.profesion,
            'nombrecorto': trabajador.nombrecorto,
            'email': trabajador.email,
            'telefono': trabajador.telefono,
            'sector': trabajador.sector.nombre if trabajador.sector else None,
            'sector_id': trabajador.sector_id,
            'recinto': trabajador.recinto.nombre if trabajador.recinto else None,
            'recinto_id': trabajador.recinto_id,
            'tiporecinto': trabajador.recinto.tiporecinto.nombre if trabajador.recinto and trabajador.recinto.tiporecinto else None,
            'tiporecinto_id': trabajador.recinto.tiporecinto.id if trabajador.recinto and trabajador.recinto.tiporecinto else None
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({
            'error': f'Error al obtener datos del trabajador: {str(e)}'
        }), 500


# ========================================
# GESTIÓN DE ADMINISTRADORES (SUPERADMIN)
# ========================================

@controllers_bp.route('/gestion-administradores', endpoint='gestion_administradores')
@login_required
def gestion_administradores():
    """
    Página para gestión de administradores según los patrones de InstruccionesPROMPT.md
    
    Funcionalidades:
    - Permite asignar/desasignar recintos específicos a cada administrador
    - Solo accesible por SUPERADMIN o usuarios con permisos específicos
    - Matriz interactiva con estadísticas en tiempo real
    
    Returns:
        Response: Template con datos de administradores y recintos
    """
    start_time = time.time()
    
    try:
        # 1. Verificación de permisos unificada
        if not (current_user.is_superadmin() or current_user.has_page_permission('/gestion-administradores')):
            print(f"❌ Acceso denegado en {request.endpoint} para usuario {current_user.id}")
            flash('No tiene permisos para acceder a esta página', 'error')
            return redirect(url_for('main.dashboard'))
        
        # 2. Logging para debugging
        print(f"🔍 Acceso a {request.endpoint} por usuario {current_user.id} ({current_user.email})")
        print(f"👤 Nivel de usuario: {'SUPERADMIN' if current_user.is_superadmin() else 'ADMINISTRADOR'}")
        
        # 3. Obtener datos con optimización de queries
        from app.models import AdministradorRecinto
        query_start = time.time()
        
        try:
            administradores, estructura, asignaciones = AdministradorRecinto.obtener_matriz_completa()
            query_duration = time.time() - query_start
            print(f"📊 Query matriz completa ejecutada en {query_duration:.3f}s")
        except Exception as db_error:
            print(f"❌ Error en consulta de base de datos: {str(db_error)}")
            raise db_error
        
        # 4. Calcular estadísticas con validación
        try:
            total_administradores = len(administradores) if administradores else 0
            total_recintos = sum(len(recintos) for sector_tipos in estructura.values() 
                               for recintos in sector_tipos.values()) if estructura else 0
            total_asignaciones = sum(len(asignaciones_admin) for asignaciones_admin in asignaciones.values()) if asignaciones else 0
            
            estadisticas = {
                'total_administradores': total_administradores,
                'total_recintos': total_recintos,
                'total_asignaciones': total_asignaciones,
                'promedio_asignaciones': round(total_asignaciones / total_administradores, 1) if total_administradores > 0 else 0,
                'porcentaje_cobertura': round((total_asignaciones / total_recintos * 100), 1) if total_recintos > 0 else 0
            }
            
            print(f"📈 Estadísticas calculadas: {estadisticas}")
            
        except Exception as stats_error:
            print(f"❌ Error calculando estadísticas: {str(stats_error)}")
            # Estadísticas por defecto en caso de error
            estadisticas = {
                'total_administradores': 0,
                'total_recintos': 0,
                'total_asignaciones': 0,
                'promedio_asignaciones': 0,
                'porcentaje_cobertura': 0
            }
        
        # 5. Render template con métricas de performance
        render_start = time.time()
        response = render_template('admin/gestion_administradores.html',
                                 administradores=administradores,
                                 estructura=estructura,
                                 asignaciones=asignaciones,
                                 estadisticas=estadisticas,
                                 css_file='gestion-administradores.css')
        
        render_duration = time.time() - render_start
        total_duration = time.time() - start_time
        
        # 6. Logging de éxito con métricas
        print(f"✅ Página {request.endpoint} cargada exitosamente en {total_duration:.3f}s")
        print(f"📊 Métricas: Query={query_duration:.3f}s, Render={render_duration:.3f}s")
        print(f"📈 Datos cargados: {total_administradores} admins, {total_recintos} recintos, {total_asignaciones} asignaciones")
        
        return response
    
    except Exception as e:
        # 7. Manejo de errores con logging completo
        duration = time.time() - start_time
        print(f"❌ Error en {request.endpoint} después de {duration:.3f}s: {str(e)}")
        import traceback
        traceback.print_exc()
        
        flash(f'Error interno del servidor: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))


@controllers_bp.route('/api/asignar-recinto', methods=['POST'], endpoint='asignar_recinto')
@login_required
def asignar_recinto():
    """
    API para asignar/desasignar un recinto a un administrador según InstruccionesPROMPT.md
    
    Args:
        JSON body: {
            administrador_id: int,
            recinto_id: int,
            asignar: bool (True para asignar, False para desasignar)
        }
        
    Returns:
        JSON: Resultado de la operación con detalles
    """
    start_time = time.time()
    
    try:
        # 1. Verificación de permisos unificada
        if not (current_user.is_superadmin() or current_user.has_page_permission('/gestion-administradores')):
            print(f"❌ Acceso denegado en {request.endpoint} para usuario {current_user.id}")
            return jsonify({'success': False, 'error': 'No tiene permisos para realizar esta acción'}), 403
        
        print(f"🔧 API {request.endpoint} llamada por usuario {current_user.id} ({current_user.email})")
        
        # 2. Validación de datos de entrada
        try:
            data = request.get_json()
            if not data:
                return jsonify({'success': False, 'error': 'Datos JSON requeridos'}), 400
                
            print(f"📄 Datos recibidos: {data}")
        except Exception as json_error:
            print(f"❌ Error parseando JSON: {str(json_error)}")
            return jsonify({'success': False, 'error': 'Formato JSON inválido'}), 400
        
        # 3. Extracción y validación de parámetros
        administrador_id = data.get('administrador_id')
        recinto_id = data.get('recinto_id')
        asignar = data.get('asignar', True)  # True para asignar, False para desasignar
        
        if not administrador_id or not recinto_id:
            print(f"❌ Datos incompletos: admin_id={administrador_id}, recinto_id={recinto_id}")
            return jsonify({'success': False, 'error': 'Se requieren administrador_id y recinto_id'}), 400
        
        # 4. Validación de existencia de entidades
        from app.models import Trabajador, AdministradorRecinto, Recinto
        
        # Verificar administrador
        administrador = Trabajador.query.filter_by(id=administrador_id, activo=True).first()
        if not administrador:
            print(f"❌ Administrador {administrador_id} no encontrado o inactivo")
            return jsonify({'success': False, 'error': 'Administrador no encontrado'}), 404
        
        # Verificar recinto
        recinto = Recinto.query.filter_by(id=recinto_id, activo=True).first()
        if not recinto:
            print(f"❌ Recinto {recinto_id} no encontrado o inactivo")
            return jsonify({'success': False, 'error': 'Recinto no encontrado'}), 404
        
        # 5. Logging de la operación
        operacion = 'asignar' if asignar else 'desasignar'
        print(f"🔄 Ejecutando {operacion}: {administrador.nombre} ↔ {recinto.nombre}")
        
        # 6. Ejecutar operación con manejo de errores específicos
        operation_start = time.time()
        
        try:
            if asignar:
                success, message = AdministradorRecinto.asignar_recinto(administrador_id, recinto_id)
            else:
                success, message = AdministradorRecinto.desasignar_recinto(administrador_id, recinto_id)
                
            operation_duration = time.time() - operation_start
            print(f"🎯 Operación completada en {operation_duration:.3f}s: success={success}")
            
        except Exception as op_error:
            print(f"❌ Error en operación de base de datos: {str(op_error)}")
            return jsonify({'success': False, 'error': f'Error en operación: {str(op_error)}'}), 500
        
        # 7. Preparar respuesta exitosa
        result = {
            'success': success,
            'message': message,
            'administrador': administrador.nombre,
            'recinto': recinto.nombre,
            'operacion': operacion,
            'timestamp': time.time()
        }
        
        # 8. Logging de éxito con métricas
        total_duration = time.time() - start_time
        print(f"✅ API {request.endpoint} completada exitosamente en {total_duration:.3f}s")
        print(f"� Resultado: {administrador.nombre} {operacion} {recinto.nombre} - {message}")
        
        return jsonify(result)
    
    except ValueError as ve:
        # 9. Manejo específico de errores de validación
        duration = time.time() - start_time
        print(f"❌ Error de validación en {request.endpoint} después de {duration:.3f}s: {str(ve)}")
        return jsonify({'success': False, 'error': f'Error de validación: {str(ve)}'}), 400
        
    except Exception as e:
        # 10. Manejo de errores generales con logging
        duration = time.time() - start_time
        print(f"❌ Error en {request.endpoint} después de {duration:.3f}s: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500


@controllers_bp.route('/api/matriz-administradores', methods=['GET'], endpoint='matriz_administradores')
@login_required
def matriz_administradores():
    """
    API para obtener la matriz actualizada de administradores vs recintos según InstruccionesPROMPT.md
    
    Returns:
        JSON: Matriz completa con administradores, estructura jerárquica y asignaciones
    """
    start_time = time.time()
    
    try:
        # 1. Verificación de permisos unificada
        if not (current_user.is_superadmin() or current_user.has_page_permission('/gestion-administradores')):
            print(f"❌ Acceso denegado en {request.endpoint} para usuario {current_user.id}")
            return jsonify({'success': False, 'error': 'No tiene permisos para realizar esta acción'}), 403
        
        print(f"📊 API {request.endpoint} llamada por usuario {current_user.id} ({current_user.email})")
        
        # 2. Obtener datos con optimización de queries
        from app.models import AdministradorRecinto
        query_start = time.time()
        
        try:
            administradores, estructura, asignaciones = AdministradorRecinto.obtener_matriz_completa()
            query_duration = time.time() - query_start
            print(f"⚡ Query matriz completa ejecutada en {query_duration:.3f}s")
        except Exception as db_error:
            print(f"❌ Error en consulta de base de datos: {str(db_error)}")
            return jsonify({'success': False, 'error': f'Error de base de datos: {str(db_error)}'}), 500
        
        # 3. Transformación de datos con validación
        transform_start = time.time()
        
        try:
            # Convertir administradores a formato JSON serializable
            administradores_data = []
            for admin in administradores:
                admin_recintos = list(asignaciones.get(admin.id, set()))
                administradores_data.append({
                    'id': admin.id,
                    'nombre': admin.nombre,
                    'email': admin.email,
                    'recintos_asignados': admin_recintos,
                    'total_recintos': len(admin_recintos)
                })
            
            # Convertir estructura jerárquica a formato JSON
            estructura_data = {}
            total_recintos_estructura = 0
            
            for sector, tipos in estructura.items():
                estructura_data[sector.id] = {
                    'nombre': sector.nombre,
                    'tipos': {}
                }
                
                for tipo, recintos in tipos.items():
                    recintos_data = []
                    for recinto in recintos:
                        recintos_data.append({
                            'id': recinto.id,
                            'nombre': recinto.nombre,
                            'activo': getattr(recinto, 'activo', True)
                        })
                        total_recintos_estructura += 1
                    
                    estructura_data[sector.id]['tipos'][tipo.id] = {
                        'nombre': tipo.nombre,
                        'recintos': recintos_data,
                        'total_recintos': len(recintos_data)
                    }
            
            transform_duration = time.time() - transform_start
            print(f"🔄 Transformación de datos completada en {transform_duration:.3f}s")
            
        except Exception as transform_error:
            print(f"❌ Error en transformación de datos: {str(transform_error)}")
            return jsonify({'success': False, 'error': f'Error procesando datos: {str(transform_error)}'}), 500
        
        # 4. Calcular estadísticas mejoradas
        total_asignaciones = sum(len(asignaciones_admin) for asignaciones_admin in asignaciones.values())
        total_administradores = len(administradores)
        
        estadisticas = {
            'total_administradores': total_administradores,
            'total_recintos': total_recintos_estructura,
            'total_asignaciones': total_asignaciones,
            'promedio_asignaciones': round(total_asignaciones / total_administradores, 2) if total_administradores > 0 else 0,
            'porcentaje_cobertura': round((total_asignaciones / total_recintos_estructura * 100), 2) if total_recintos_estructura > 0 else 0,
            'admins_sin_asignaciones': sum(1 for admin_id in asignaciones if len(asignaciones[admin_id]) == 0)
        }
        
        # 5. Preparar respuesta completa
        matriz_data = {
            'administradores': administradores_data,
            'estructura': estructura_data,
            'estadisticas': estadisticas,
            'timestamp': time.time()
        }
        
        # 6. Logging de éxito con métricas
        total_duration = time.time() - start_time
        print(f"✅ API {request.endpoint} completada exitosamente en {total_duration:.3f}s")
        print(f"📊 Métricas: Query={query_duration:.3f}s, Transform={transform_duration:.3f}s")
        print(f"📈 Datos: {total_administradores} admins, {total_recintos_estructura} recintos, {total_asignaciones} asignaciones")
        
        return jsonify({
            'success': True,
            'data': matriz_data
        })
    
    except ValueError as ve:
        # 7. Manejo específico de errores de validación
        duration = time.time() - start_time
        print(f"❌ Error de validación en {request.endpoint} después de {duration:.3f}s: {str(ve)}")
        return jsonify({'success': False, 'error': f'Error de validación: {str(ve)}'}), 400
        
    except Exception as e:
        # 8. Manejo de errores generales con logging
        duration = time.time() - start_time
        print(f"❌ Error en {request.endpoint} después de {duration:.3f}s: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500


@controllers_bp.route('/api/health-check', methods=['GET'], endpoint='health_check')
@login_required
def health_check():
    """
    API para verificar el estado del servidor y la conexión
    """
    try:
        # Verificar conexión a base de datos
        from sqlalchemy import text
        db.session.execute(text('SELECT 1'))
        db.session.commit()
        
        return jsonify({
            'success': True,
            'status': 'OK',
            'message': 'Servidor funcionando correctamente',
            'timestamp': datetime.now().isoformat(),
            'user': {
                'id': current_user.id,
                'email': current_user.email,
                'rol': current_user.rol.name if hasattr(current_user, 'rol') and current_user.rol else None
            }
        })
    
    except Exception as e:
        print(f"❌ ERROR en health-check: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'status': 'ERROR',
            'message': f'Error del servidor: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500


# ========================================
# GESTIÓN DE USUARIOS POR ADMINISTRADOR
# ========================================

@controllers_bp.route('/gestion-usuarios', endpoint='gestion_usuarios')
@login_required
def gestion_usuarios():
    """
    Página para gestión de usuarios - Solo ADMINISTRADOR
    Permite asignar/desasignar recintos adicionales a trabajadores de sus recintos
    """
    try:
        # Verificar permisos usando sistema dinámico
        if not (current_user.is_superadmin() or current_user.has_page_permission('/gestion-usuarios-recinto')):
            flash('No tiene permisos para acceder a esta página', 'error')
            return redirect(url_for('main.dashboard'))
        
        # Si es SUPERADMIN, obtener todos los trabajadores, sino filtrar por administrador
        if current_user.is_superadmin():
            # SUPERADMIN ve todos los trabajadores activos
            trabajadores = Trabajador.query.filter_by(activo=True).order_by(Trabajador.nombre).all()
            
            # Obtener estructura completa de recintos
            from app.models import Sector, TipoRecinto, Recinto
            sectores = Sector.query.filter_by(activo=True).order_by(Sector.nombre).all()
            estructura = {}
            for sector in sectores:
                estructura[sector] = {}
                tipos_recinto = TipoRecinto.query.filter_by(
                    id_sector=sector.id, 
                    activo=True
                ).order_by(TipoRecinto.nombre).all()
                
                for tipo in tipos_recinto:
                    estructura[sector][tipo] = Recinto.query.filter_by(
                        id_tiporecinto=tipo.id,
                        activo=True
                    ).order_by(Recinto.nombre).all()
            
            # Obtener todas las asignaciones activas
            asignaciones = {}
            for trabajador in trabajadores:
                asignaciones[trabajador.id] = []
                asignaciones_trabajador = TrabajadorRecinto.query.filter_by(
                    trabajador_id=trabajador.id,
                    activo=True
                ).all()
                asignaciones[trabajador.id] = [asig.recinto_id for asig in asignaciones_trabajador]
        else:
            # Para administradores normales, usar la función filtrada
            trabajadores, estructura, asignaciones = TrabajadorRecinto.obtener_matriz_por_administrador(current_user.id)
        
        # Contar totales para estadísticas
        total_trabajadores = len(trabajadores)
        total_recintos = sum(len(recintos) for sector_tipos in estructura.values() 
                           for recintos in sector_tipos.values())
        total_asignaciones = sum(len(asignaciones_trabajador) for asignaciones_trabajador in asignaciones.values())
        
        # Obtener recintos que gestiona (todos para SUPERADMIN, filtrados para admin normal)
        if current_user.is_superadmin():
            # SUPERADMIN gestiona todos los recintos
            recintos_gestionados = Recinto.query.filter_by(activo=True).order_by(Recinto.nombre).all()
        else:
            # Para administradores normales, obtener solo sus recintos asignados
            recintos_admin = AdministradorRecinto.obtener_recintos_administrador(current_user.id)
            recintos_gestionados = []
            for asignacion in recintos_admin:
                recinto = Recinto.query.get(asignacion.recinto_id)
                if recinto:
                    recintos_gestionados.append(recinto)
        
        estadisticas = {
            'total_trabajadores': total_trabajadores,
            'total_recintos': total_recintos,
            'total_asignaciones': total_asignaciones,
            'promedio_asignaciones': round(total_asignaciones / total_trabajadores, 1) if total_trabajadores > 0 else 0,
            'recintos_gestionados': len(recintos_gestionados)
        }
        
        return render_template('admin/gestion_usuarios.html',
                             trabajadores=trabajadores,
                             estructura=estructura,
                             asignaciones=asignaciones,
                             estadisticas=estadisticas,
                             recintos_gestionados=recintos_gestionados)
    
    except Exception as e:
        print(f'❌ ERROR EN CONTROLADOR gestion_usuarios: {str(e)}')
        import traceback
        traceback.print_exc()
        flash(f'Error interno del servidor: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))


@controllers_bp.route('/api/asignar-recinto-usuario', methods=['POST'], endpoint='asignar_recinto_usuario')
@login_required
def asignar_recinto_usuario():
    """
    API para asignar/desasignar un recinto a un trabajador (solo administradores)
    """
    try:
        # Verificar permisos usando sistema dinámico
        if not (current_user.is_superadmin() or current_user.has_page_permission('/gestion-usuarios-recinto')):
            return jsonify({'success': False, 'error': 'No tiene permisos para realizar esta acción'}), 403
        
        data = request.get_json()
        
        trabajador_id = data.get('trabajador_id')
        recinto_id = data.get('recinto_id')
        asignar = data.get('asignar', True)  # True para asignar, False para desasignar
        
        # Validar datos
        if not trabajador_id or not recinto_id:
            return jsonify({'success': False, 'error': 'Faltan datos requeridos'}), 400
        
        # Verificar que el trabajador existe
        trabajador = Trabajador.query.get(trabajador_id)
        if not trabajador:
            return jsonify({'success': False, 'error': 'Trabajador no encontrado'}), 404
        
        # Verificar que el recinto existe
        recinto = Recinto.query.filter_by(id=recinto_id, activo=True).first()
        if not recinto:
            return jsonify({'success': False, 'error': 'Recinto no encontrado'}), 404
        
        # Validaciones de permisos solo para administradores normales (NO para SUPERADMIN)
        if not current_user.is_superadmin():
            # Verificar que el administrador tiene acceso a gestionar este recinto
            if not AdministradorRecinto.tiene_acceso_recinto(current_user.id, recinto_id):
                return jsonify({'success': False, 'error': 'No tiene permisos para gestionar este recinto'}), 403
            
            # Verificar que el trabajador pertenece a un recinto gestionado por este admin
            recintos_admin = AdministradorRecinto.obtener_recintos_administrador(current_user.id)
            recinto_ids_admin = [asignacion.recinto_id for asignacion in recintos_admin]
            
            if trabajador.recinto_id not in recinto_ids_admin:
                return jsonify({'success': False, 'error': 'No puede gestionar trabajadores de este recinto'}), 403
        
        # Realizar la asignación o desasignación
        if asignar:
            success, message = TrabajadorRecinto.asignar_recinto(trabajador_id, recinto_id)
        else:
            success, message = TrabajadorRecinto.desasignar_recinto(trabajador_id, recinto_id)
        
        result = {
            'success': success,
            'message': message,
            'trabajador': trabajador.nombre,
            'recinto': recinto.nombre
        }
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500


@controllers_bp.route('/api/matriz-usuarios', methods=['GET'], endpoint='matriz_usuarios')
@login_required
def matriz_usuarios():
    """
    API para obtener la matriz actualizada de trabajadores vs recintos para un administrador
    """
    try:
        # Verificar permisos usando sistema dinámico
        if not (current_user.is_superadmin() or current_user.has_page_permission('/gestion-usuarios-recinto')):
            return jsonify({'success': False, 'error': 'No tiene permisos'}), 403
        
        trabajadores, estructura, asignaciones = TrabajadorRecinto.obtener_matriz_por_administrador(current_user.id)
        
        # Convertir estructura a formato JSON serializable
        estructura_json = {}
        for sector, tipos in estructura.items():
            estructura_json[sector.id] = {
                'nombre': sector.nombre,
                'tipos': {}
            }
            for tipo, recintos in tipos.items():
                estructura_json[sector.id]['tipos'][tipo.id] = {
                    'nombre': tipo.nombre,
                    'recintos': [{'id': r.id, 'nombre': r.nombre} for r in recintos]
                }
        
        return jsonify({
            'success': True,
            'trabajadores': [{'id': t.id, 'nombre': t.nombre, 'email': t.email} for t in trabajadores],
            'estructura': estructura_json,
            'asignaciones': asignaciones
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error interno: {str(e)}'}), 500

# ==================================================================================
# Rutas CRUD para Fases
@controllers_bp.route('/fases', endpoint='ruta_fases')
@login_required
def fases():
    """Página principal de gestión de fases"""
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/fases')):
        flash('No tiene permisos para acceder a esta página', 'error')
        return redirect(url_for('main.dashboard'))
    
    # Obtener datos según nivel de permisos
    if current_user.is_superadmin():
        fases = Fase.query.all()
    else:
        # Para fases, al ser datos generales del sistema, se muestran todos
        # independientemente del recinto (a menos que se implemente lógica específica)
        fases = Fase.query.all()
    
    return render_template('fases.html', fases=fases)

@controllers_bp.route('/add_fase', methods=['POST'], endpoint='add_fase')
@login_required
def add_fase():
    """Agregar nueva fase"""
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/fases')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    from flask import session
    
    # Verificar token para prevenir doble envío
    form_token = request.form.get('form_token')
    session_token = session.get('last_add_form_token')
    
    if form_token and form_token == session_token:
        flash('Operación ya procesada', 'warning')
        return redirect(url_for('controllers.ruta_fases'))
    
    try:
        nombre = request.form['name']
        nueva_fase = Fase(nombre=nombre)
        db.session.add(nueva_fase)
        db.session.commit()
        
        # Guardar token para prevenir doble procesamiento
        if form_token:
            session['last_add_form_token'] = form_token
            
        flash('Fase agregada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al agregar fase: {e}', 'error')
    return redirect(url_for('controllers.ruta_fases'))

@controllers_bp.route('/update_fase/<int:id>', methods=['POST'], endpoint='update_fase')
@login_required
def update_fase(id):
    """Actualizar fase existente"""
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/fases')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    from flask import session
    
    # Verificar token para prevenir doble envío
    form_token = request.form.get('form_token')
    session_token = session.get('last_form_token')
    
    if form_token and form_token == session_token:
        flash('Operación ya procesada', 'warning')
        return redirect(url_for('controllers.ruta_fases'))
    
    try:
        fase = Fase.query.get_or_404(id)
        fase.nombre = request.form['name']
        db.session.commit()
        
        # Guardar token para prevenir doble procesamiento
        if form_token:
            session['last_form_token'] = form_token
            
        flash('Fase actualizada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar fase: {e}', 'error')
    return redirect(url_for('controllers.ruta_fases'))

@controllers_bp.route('/eliminar_fase/<int:id>', methods=['POST'], endpoint='eliminar_fase')
@login_required
def eliminar_fase(id):
    """Eliminar fase"""
    # Verificar permisos
    if not (current_user.is_superadmin() or current_user.has_page_permission('/fases')):
        flash('No tiene permisos para realizar esta acción', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        fase = Fase.query.get_or_404(id)
        db.session.delete(fase)
        db.session.commit()
        flash('Fase eliminada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar fase: {e}', 'error')
    return redirect(url_for('controllers.ruta_fases'))

@controllers_bp.route('/test_cascade.html')
def test_cascade():
    """Test de selects en cascada"""
    from flask import send_from_directory
    return send_from_directory('.', 'test_cascade.html')

@controllers_bp.route('/prueba-menu', endpoint='ruta_prueba_menu')
@login_required
def prueba_menu():
    """Página de prueba para verificar el funcionamiento del menú Bootstrap"""
    return render_template('prueba-menu.html')

@controllers_bp.route('/test-header-debug', endpoint='test_header_debug')
@login_required
def test_header_debug():
    """Página de debug para analizar el menú Bootstrap en detalle"""
    return render_template('test_header_menu_debug.html')
